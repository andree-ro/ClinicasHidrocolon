import openpyxl
import pymysql
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QColor, QIcon, QPixmap
from PyQt5.uic import loadUi
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QMainWindow, QMessageBox, QAbstractItemView, QTableWidget, QPushButton, QSizePolicy, \
    QHBoxLayout, QWidget, QTableWidgetItem, QCheckBox, QLineEdit, QDialog, QFileDialog
from PyQt5.QtWidgets import QApplication, QMainWindow, QCalendarWidget, QVBoxLayout, QWidget, QLabel
from PyQt5.QtCore import QDate
from datetime import datetime
from .column_selector_farmacia import ColumnSelectorDialog, FarmaciaColumnManager
# Al inicio del archivo ventanaFuncional.py, agregar esta importación si no está:
from PyQt5.QtWidgets import QMessageBox

from reportlab.pdfgen import canvas

import sql_structures
from .ModificarMedicamentos import *
from .Jornadas import *
from .ModificarTerapia import *
from .Carrito import *
from .ModificarPacientes import *
from .ModificarCombos import *
from .descuentosMedi import *
from .ModificarUsuarios import *
from .ModificarCierre import *
from .medi_combos import DialogoCombo
from .pago_dividido import *
from .medi_ter import *
from .medi_jorda import *
from .Extras import *
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, legal
from reportlab.lib.units import inch
import os
from sql_structures.manager import Manager

class VentanaFuncional(QMainWindow):
    switch_window = QtCore.pyqtSignal(str)
    _contra = " "

    _dinero_total = 0.0
    _diferencia_efectivo = 0.0
    _usuario = ''
    _dato = ''
    _reiniciar_carrito = " "
    _id_detalle = 0
    _instancia_actual = None
    _metodo_pago_guardado = "efectivo"  # ← NUEVA LÍNEA
    _es_pago_dividido = False  # ← NUEVA LÍNEA
    _monto_efectivo_dividido = 0.0  # ← NUEVA LÍNEA
    _monto_tarjeta_dividido = 0.0   # ← NUEVA LÍNEA

    def __init__(self):
        super(VentanaFuncional, self).__init__()

        loadUi('C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\Hidrocolon.ui', self)

        self.debug_radio_buttons()

        # Configurar un QTimer para actualizar la etiqueta periódicamente
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.actualizar_datos_carrito)
        self.timer.start(1000)  # Actualización cada 1 segundo (1000 ms)

        # AGREGAR ESTAS LÍNEAS:
        self.farmacia_column_manager = FarmaciaColumnManager(self.bd_farmacia, self)
        self.farmacia_column_manager.apply_column_visibility()
        self.setup_column_selector_button()

        self.factura = self.label_30.text()
        self.resultadoProductos = 0.0
        self.resultadot = 0.0
        self.id_c = 0
        self.resultado = 0.0
        self.cabeza = ""
        self.new_value = ''
        self.selected_items = []
        self.nombre_medi = ""
        self.presentacio_medi = ""
        self.laboratorio_medi = ""
        self.existencias_medi = ""
        self.tarjeta_medi = ""
        self.efectivo_medi = ""
        self.id_nombre = ""
        self.medicamento = AgregarMedi()
        self.jornada = jornadas()
        self.terapias = AgregarTerapia()
        self.paciente = AgregarPacientes()
        self.carrito = Metodos_carrito()
        self.combo = AgregarCombos()
        self.descu = DescuentoMedi(self)
        self.usuario_in = AgregarUsuarios()
        self.cierre = AgregarCierre()
        self.mana = Manager()

        self.cargarTablacarrito()

        # self.calendar.selectionChanged.connect(self.filtro_dia)

        self.lineEdit_3.returnPressed.connect(self.resto)
        # self.Buscar_medicina.returnPressed.connect(self.busqueda_farmacia)
        self.Buscar_medicina.textChanged.connect(self.busqueda_farmacia)
        self.Buscar_jornada.textChanged.connect(self.busqueda_jornada)
        self.Buscar_terapia.textChanged.connect(self.busqueda_terapia)
        self.Buscar_paciente.textChanged.connect(self.busqueda_paciente)
        self.Buscar_combo.textChanged.connect(self.busqueda_combo)
        self.Buscar_usuario.textChanged.connect(self.busqueda_usuario)

        # self.bd_farmacia.setColumnWidth(0, 1)
        self.bd_farmacia.setColumnWidth(0, 75)
        self.bd_farmacia.setColumnWidth(1, 240)
        self.bd_farmacia.setColumnWidth(2, 110)
        self.bd_farmacia.setColumnWidth(3, 100)
        self.bd_farmacia.setColumnWidth(4, 95)
        self.bd_farmacia.setColumnWidth(5, 175)
        self.bd_farmacia.setColumnWidth(6, 70)
        self.bd_farmacia.setColumnWidth(7, 95)
        self.bd_farmacia.setColumnWidth(8, 120)

        self.bd_cierre.setColumnWidth(0, 140)
        self.bd_cierre.setColumnWidth(1, 110)
        self.bd_cierre.setColumnWidth(2, 90)
        self.bd_cierre.setColumnWidth(3, 130)
        self.bd_cierre.setColumnWidth(4, 130)
        self.bd_cierre.setColumnWidth(5, 150)
        self.bd_cierre.setColumnWidth(6, 120)

        self.bd_terapias.setColumnWidth(0, 25)
        self.bd_terapias.setColumnWidth(1, 240)
        self.bd_terapias.setColumnWidth(2, 110)
        self.bd_terapias.setColumnWidth(3, 110)
        self.bd_terapias.setColumnWidth(4, 120)

        self.bd_jornadas.setColumnWidth(0, 25)
        self.bd_jornadas.setColumnWidth(1, 240)
        self.bd_jornadas.setColumnWidth(2, 110)
        self.bd_jornadas.setColumnWidth(3, 110)
        self.bd_terapias.setColumnWidth(4, 120)

        self.bd_vitacora.setColumnWidth(0, 100)
        self.bd_vitacora.setColumnWidth(1, 120)
        self.bd_vitacora.setColumnWidth(2, 450)
        self.bd_vitacora.setColumnWidth(3, 120)
        self.bd_vitacora.setColumnWidth(4, 120)

        self.btn_modulo_farmacia.clicked.connect(self.show_page_farmacia)
        self.btn_modulo_terapias.clicked.connect(self.show_page_terapias)
        self.btn_modulo_jornadas.clicked.connect(self.show_page_jornadas)
        self.btn_modulo_pacientes.clicked.connect(self.show_page_pacientes)
        self.btn_modulo_combos.clicked.connect(self.show_page_combos)
        self.btn_modulo_usuarios.clicked.connect(self.show_page_usuarios)
        self.btn_modulo_cierre.clicked.connect(self.show_page_cierre)
        self.btn_cumple.clicked.connect(self.show_page_cumple)
        self.btn_cita.clicked.connect(self.show_page_cita)
        self.btn_vitacora.clicked.connect(self.show_page_vitacora)
        self.btn_modulo_banco.clicked.connect(self.show_page_fina)

        self.btn_actualizar_medi.clicked.connect(self.IniciarMod)
        self.bd_farmacia.cellClicked.connect(self.click_tabla_medicamento)

        self.btn_actualizar_terapia.clicked.connect(self.IniciarTerapia)

        self.btn_actualizar_jornada.clicked.connect(self.IniciarJornadas)

        self.btn_actualizar_paciente.clicked.connect(self.IniciarPacientes)

        self.btn_actualizar_combo.clicked.connect(self.IniciarCombos)

        self.btn_agregar_usuario.clicked.connect(self.IniciarUsuarios)

        self.btn_pagodiv.clicked.connect(self.mostrar_dialogo_pago_dividido)

        self.btn_descuentos.clicked.connect(self.IniciarContra_des)
        self.bnt_devu.clicked.connect(self.Devolu)
        self.bnt_devu_3.clicked.connect(self.Devolu_total)
        self.btn_in_efe.clicked.connect(self.Iniciar_ing_efec)
        self.btn_in_tar.clicked.connect(self.Iniciar_ing_tar)
        self.btn_in_trasfe.clicked.connect(self.Iniciar_ing_trasf)
        self.btn_in_gastos.clicked.connect(self.Iniciar_ing_gastos)
        self.btn_in_crudo.clicked.connect(self.Iniciar_ing_crudo)
        self.btn_in_cheque.clicked.connect(self.Iniciar_ing_cheques)
        self.btn_Repor.clicked.connect(self.Iniciar_ing_reporte)
        self.btn_in_comision.clicked.connect(self.Iniciar_ing_comision)

        self.btn_enviar_datos.clicked.connect(self.iniciarDatos)

        self.carrito_medi.clicked.connect(self.IniciarEX)

        self.btn_carrito_farmacia.clicked.connect(self.show_page_carrito)
        self.btn_carrito_terapias.clicked.connect(self.show_page_carrito)
        self.btn_carrito_jornadas.clicked.connect(self.show_page_carrito)
        self.btn_carrito_combos.clicked.connect(self.show_page_carrito)
        # self.carrito_medi.clicked.connect(self.agregar_medicamento_a_carrito)
        self.carrito_jorda.clicked.connect(self.registrar_jornada)
        self.carrito_terapia.clicked.connect(self.registrar_terapia)
        # self.carrito_terapia.clicked.connect(self.agregar_terapia_a_carrito)
        self.carrito_combo.clicked.connect(self.agregar_combo_a_carrito)

        # Botones recargar tablas
        self.btn_modulo_jornadas.clicked.connect(self.cargarTablaJornadas)
        self.bd_jornadas.cellClicked.connect(self.click_tabla_jornadas)
        self.btn_cita.clicked.connect(self.cargarTablaPacientes_cita)
        self.btn_cumple.clicked.connect(self.cargarTablaPacientes_cumple)
        self.btn_modulo_terapias.clicked.connect(self.cargarTablaTerapias)
        self.bd_terapias.cellClicked.connect(self.click_tabla_terapias)
        self.bd_terapias.cellChanged.connect(self.on_cell_changed_tera)
        self.bd_cierre.cellClicked.connect(self.click_tabla_cierre)
        self.bd_pacientes_citas.cellChanged.connect(self.click_tabla_pacientes_citas)
        # self.bd_pacientes_citas.cellChanged.connect(self.on_cell_changed_citas)
        self.bd_pacientes_citas.itemChanged.connect(self.on_item_changed)

        self.btn_modulo_combos.clicked.connect(self.cargarTablaTerapias)
        self.bd_combos.cellClicked.connect(self.click_tabla_combo)
        self.bd_combos.cellChanged.connect(self.on_cell_changed_combo)

        self.btn_modulo_usuarios.clicked.connect(self.cargarTablaUsuario)
        self.bd_usuario.cellClicked.connect(self.click_tabla_usuario)
        self.bd_usuario.cellChanged.connect(self.on_cell_changed_usu)

        self.bd_farmacia.cellChanged.connect(self.on_cell_changed_far)
        self.bd_jornadas.cellChanged.connect(self.on_cell_changed_jorda)
        self.bd_pacientes.cellClicked.connect(self.click_tabla_pacientes)
        self.bd_pacientes.cellChanged.connect(self.on_cell_changed_paci)

        self.btn_carrito_farmacia.clicked.connect(self.cargarTablacarrito)
        self.bd_carrito.cellClicked.connect(self.click_tabla_carrito)
        self.btn_carrito_jornadas.clicked.connect(self.cargarTablacarrito)
        self.btn_carrito_terapias.clicked.connect(self.cargarTablacarrito)
        self.btn_carrito_combos.clicked.connect(self.cargarTablacarrito)

        self.btn_limpiar_tabla.clicked.connect(self.borrar_tabla)
        self.btn_enviar_datos.clicked.connect(self.ingresar_cierre)

        self.radio_efectivo.toggled.connect(self.maleselected)
        self.radio_tarjeta.toggled.connect(self.femaleselected)

        self.radio_efectivo.toggled.connect(self.suma_total_pagos)
        self.radio_tarjeta.toggled.connect(self.suma_total_pagos)

        self.btn_busca_medi.clicked.connect(self.busqueda_farmacia)
        self.btn_busca_jorda.clicked.connect(self.busqueda_jornada)

        self.btn_modulo_cierre.clicked.connect(self.cargarTablaCierre)
        self.btn_modulo_cierre.clicked.connect(self.monto_e)

        self.btn_dia.clicked.connect(self.filtro_dia)
        self.btn_semana.clicked.connect(self.filtro_semana)
        self.btn_mes.clicked.connect(self.filtro_mes)
        self.btn_anio.clicked.connect(self.filtro_anio)

        self.btn_excel_farmacia_2.clicked.connect(self.IniciarExt)

        self.pushButton_20.clicked.connect(self.generar_pdf_farmacia)
        self.pushButton_28.clicked.connect(self.generar_pdf_terapias)
        self.pushButton_19.clicked.connect(self.generar_pdf_jornadas)
        self.pushButton_18.clicked.connect(self.generar_pdf_paciente)
        self.pushButton_27.clicked.connect(self.generar_pdf_combos)
        self.pushButton_21.clicked.connect(self.generar_pdf_cierre)

        self.bd_farmacia.cellDoubleClicked.connect(self.click_tabla_medicamento)

        self.sesion.clicked.connect(self.FinalizarMod)

        #        self.btn_agregar_datos_factura.clicked.connect(self.generar_pdf_comprobante)

        self.btn_excel_farmacia.clicked.connect(self.excel_farmacia)
        self.btb_excel_jornadas.clicked.connect(self.excel_jornadas)
        self.btn_excel_pacientes.clicked.connect(self.excel_pacientes)
        self.btn_excel_combos.clicked.connect(self.excel_combos)
        self.btn_excel_terapias.clicked.connect(self.excel_terapias)
        self.btn_pdf_pacientes.clicked.connect(self.pdf_pacientes)

        self.btn_pdf_bitacora.clicked.connect(self.ventana_fecha)

        self.btn_bitacora_ventas.clicked.connect(self.generar_pdf_bitacora_ventas)

        # Variables para paciente vinculado
        self.paciente_actual = None

        # Crear botón para vincular paciente
        self.btn_vincular_paciente = QPushButton("Vincular Paciente")
        self.btn_vincular_paciente.clicked.connect(self.vincular_paciente_venta)

        # Crear label para mostrar paciente actual
        self.label_paciente_actual = QLabel("Sin paciente vinculado")

        VentanaFuncional._instancia_actual = self

        # =============================================================
        # SECCIÓN DE UBICACIÓN Y ESTILOS DE BOTONES DEL PACIENTE
        # =============================================================

        try:
            # Buscar un contenedor existente donde agregar los botones
            if hasattr(self, 'page_carrito'):
                contenedor = self.page_carrito
            elif hasattr(self, 'frame_carrito'):
                contenedor = self.frame_carrito
            elif hasattr(self, 'widget_carrito'):
                contenedor = self.widget_carrito
            else:
                contenedor = self.centralwidget if hasattr(self, 'centralwidget') else self

            # Crear layout para los nuevos botones
            if not hasattr(contenedor, 'layout') or contenedor.layout() is None:
                nuevo_layout = QVBoxLayout(contenedor)
            else:
                nuevo_layout = contenedor.layout()

            # Crear frame para organizar los elementos del paciente
            self.frame_paciente = QFrame()
            self.frame_paciente.setFrameStyle(QFrame.StyledPanel)

            # Layout del frame del paciente
            paciente_layout = QVBoxLayout(self.frame_paciente)

            # Agregar label del paciente
            paciente_layout.addWidget(self.label_paciente_actual)

            # Layout horizontal para los botones
            botones_layout = QHBoxLayout()
            botones_layout.addWidget(self.btn_vincular_paciente)

            # Agregar botón para ver historial
            self.btn_ver_historial = QPushButton("Ver Historial")
            self.btn_ver_historial.clicked.connect(self.ver_historial_paciente)
            botones_layout.addWidget(self.btn_ver_historial)

            paciente_layout.addLayout(botones_layout)

            # Agregar el frame al contenedor principal
            if isinstance(nuevo_layout, QVBoxLayout):
                nuevo_layout.addWidget(self.frame_paciente)
            elif isinstance(nuevo_layout, QHBoxLayout):
                nuevo_layout.addWidget(self.frame_paciente)
            else:
                self.frame_paciente.setParent(contenedor)
                self.frame_paciente.move(10, 10)
                self.frame_paciente.resize(300, 100)

            print("Botones de paciente agregados correctamente")

        except Exception as e:
            print(f"Error al agregar botones: {e}")
            # Plan B: Posicionamiento manual
            try:
                self.frame_paciente = QFrame(self)
                self.frame_paciente.setFrameStyle(QFrame.StyledPanel)
                self.frame_paciente.setGeometry(10, 10, 300, 100)

                paciente_layout = QVBoxLayout(self.frame_paciente)
                paciente_layout.addWidget(self.label_paciente_actual)

                botones_layout = QHBoxLayout()
                botones_layout.addWidget(self.btn_vincular_paciente)

                self.btn_ver_historial = QPushButton("Ver Historial")
                self.btn_ver_historial.clicked.connect(self.ver_historial_paciente)
                botones_layout.addWidget(self.btn_ver_historial)

                paciente_layout.addLayout(botones_layout)

                self.frame_paciente.show()
                print("Botones agregados con posicionamiento manual")

            except Exception as e2:
                print(f"Error en plan B: {e2}")

        # =============================================================
        # APLICAR ESTILOS ELEGANTES A LOS BOTONES
        # =============================================================

        # Estilos CSS que coinciden con tu interfaz
        estilo_boton_principal = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #5a9fd4, stop:1 #306ba3);
                    border: 2px solid #2d5d87;
                    border-radius: 8px;
                    color: white;
                    font-weight: bold;
                    font-size: 12px;
                    padding: 8px 16px;
                    min-height: 25px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #6bb6ff, stop:1 #4d8fd1);
                    border: 2px solid #4d8fd1;
                }
                QPushButton:pressed {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #306ba3, stop:1 #1e4d73);
                }
            """

        estilo_boton_secundario = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #7fb069, stop:1 #5a8049);
                    border: 2px solid #4a6b3a;
                    border-radius: 8px;
                    color: white;
                    font-weight: bold;
                    font-size: 12px;
                    padding: 8px 16px;
                    min-height: 25px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #8bc475, stop:1 #6a9654);
                    border: 2px solid #6a9654;
                }
                QPushButton:pressed {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #5a8049, stop:1 #3d5831);
                }
            """

        estilo_label_paciente = """
                QLabel {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #e8f4fd, stop:1 #d1e9f7);
                    border: 2px solid #4d8fd1;
                    border-radius: 8px;
                    color: #2d5d87;
                    font-weight: bold;
                    font-size: 11px;
                    padding: 6px 12px;
                    margin: 2px;
                }
            """

        estilo_frame = """
                QFrame {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #f0f8ff, stop:1 #e6f3ff);
                    border: 2px solid #b3d9f7;
                    border-radius: 12px;
                    margin: 5px;
                    padding: 5px;
                }
            """

        # Aplicar estilos a los botones y label
        self.btn_vincular_paciente.setStyleSheet(estilo_boton_principal)
        self.btn_vincular_paciente.setText("👤 Vincular Paciente")
        self.btn_vincular_paciente.setMinimumSize(140, 35)

        if hasattr(self, 'btn_ver_historial'):
            self.btn_ver_historial.setStyleSheet(estilo_boton_secundario)
            self.btn_ver_historial.setText("📋 Ver Historial")
            self.btn_ver_historial.setMinimumSize(120, 35)

        self.label_paciente_actual.setStyleSheet(estilo_label_paciente)

        if hasattr(self, 'frame_paciente'):
            # Reducir ANCHO máximo del frame
            self.frame_paciente.setMaximumWidth(280)  # Mucho más estrecho
            self.frame_paciente.setMinimumWidth(280)
            self.frame_paciente.setMaximumHeight(120)  # Altura normal

            # Layout más compacto horizontalmente
            if self.frame_paciente.layout():
                layout = self.frame_paciente.layout()
                layout.setSpacing(6)
                layout.setContentsMargins(6, 8, 6, 8)

            # Reorganizar botones en VERTICAL para ahorrar ancho
            if hasattr(self, 'btn_ver_historial'):
                # Buscar el layout de botones y cambiarlo a vertical
                botones_layout = None
                for i in range(layout.count()):
                    item = layout.itemAt(i)
                    if item and isinstance(item, QHBoxLayout):
                        botones_layout = item
                        break

                if botones_layout:
                    # Cambiar a layout vertical
                    while botones_layout.count():
                        child = botones_layout.takeAt(0)
                        if child.widget():
                            child.widget().setParent(None)

                    # Crear nuevo layout vertical para botones
                    botones_v_layout = QVBoxLayout()
                    botones_v_layout.setSpacing(4)

                    # Botones más pequeños y estrechos
                    self.btn_vincular_paciente.setMinimumSize(100, 25)
                    self.btn_vincular_paciente.setMaximumSize(150, 25)
                    self.btn_vincular_paciente.setText("👤 Vincular")  # Texto más corto

                    self.btn_ver_historial.setMinimumSize(100, 25)
                    self.btn_ver_historial.setMaximumSize(150, 25)
                    self.btn_ver_historial.setText("📋 Historial")  # Texto más corto

                    botones_v_layout.addWidget(self.btn_vincular_paciente)
                    botones_v_layout.addWidget(self.btn_ver_historial)

                    # Reemplazar en el layout principal
                    layout.addLayout(botones_v_layout)

            # Label más estrecho con texto en múltiples líneas
            font_size_compacto = """
                QLabel {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #e8f4fd, stop:1 #d1e9f7);
                    border: 2px solid #4d8fd1;
                    border-radius: 6px;
                    color: #2d5d87;
                    font-weight: bold;
                    font-size: 9px;
                    padding: 4px 6px;
                    margin: 1px;
                    max-width: 250px;
                }
            """
            self.label_paciente_actual.setStyleSheet(font_size_compacto)
            self.label_paciente_actual.setWordWrap(True)  # Permitir múltiples líneas
            self.label_paciente_actual.setMaximumWidth(250)


    def generar_pdf_bitacora_ventas(self):
        try:
            # Obtener la ruta del escritorio
            # desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
            nombre_archivo = os.path.join(desktop, "Bitacora-Ventas-Anulacion.pdf")

            encabezados = ["Producto", "Cantidad", "Total", "Fecha", "Accion"]
            ancho_personalizado, alto_personalizado = letter
            pdf = SimpleDocTemplate(nombre_archivo, pagesize=(ancho_personalizado, alto_personalizado),
                                    topMargin=0.15 * inch)
            elementos = []

            # Agregar logo
            logo = Image('logo.png')
            logo.drawHeight = 1.5 * inch
            logo.drawWidth = 1.5 * inch
            elementos.append(logo)

            # Estilo y título
            styles = getSampleStyleSheet()
            estilo_titulo = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1,
                                           spaceAfter=30)
            titulo = Paragraph("Bitacora De Ventas y Devoluciones", estilo_titulo)
            elementos.append(titulo)
            elementos.append(Spacer(1, 20))

            # Estilo de la tabla
            estilo_tabla = TableStyle(
                [('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                 ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                 ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                 ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                 ('FONTSIZE', (0, 0), (-1, 0), 14),
                 ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                 ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                 ('GRID', (0, 0), (-1, -1), 1, colors.black),
                 ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                 ('ALIGN', (0, 0), (-1, 0), 'CENTER')])

            # Obtener datos de la base de datos
            datos_tabla = [encabezados] + list(self.mana.bitacora_ventas())
            ancho_columnas = [2 * inch, 1.75 * inch, 1.30 * inch, 1.25 * inch, 1 * inch]
            tabla = Table(datos_tabla, colWidths=ancho_columnas)
            tabla.setStyle(estilo_tabla)

            # Agregar tabla al documento
            elementos.append(tabla)

            # Crear el archivo PDF
            pdf.build(elementos)
            QMessageBox.about(self, 'Aviso', f'PDF Generado correctamente')
        except Exception as e:
            print(e)

    def pdf_pacientes(self):
        from datetime import datetime
        import os
        from reportlab.lib.pagesizes import landscape, legal
        from reportlab.pdfgen import canvas

        try:
            today = datetime.today()
            # Obtener la fecha actual en formato "YYYY-MM-DD"
            fecha_actual = f"{today.year}{today.month:02d}{today.day:02d}"

            # Nombre del archivo PDF usando la fecha
            nombre_pdf = f"Reporte_Cita_Pacientes_{fecha_actual}.pdf"

            # Obtener la ruta del escritorio
            desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
            #desktop = "C:\\Users\\VICTUS\\OneDrive\\Escritorio"  # Para Windows/Linux
            # Si estás en un sistema en español, usa:
            # desktop = os.path.join(os.path.expanduser('~'), 'Escritorio')

            # Crear la carpeta "Reportes Cita Pacientes" en el escritorio
            carpeta = os.path.join(desktop, "Reportes Cita Pacientes")
            if not os.path.exists(carpeta):
                os.makedirs(carpeta)

            # Ruta completa donde se guardará el archivo
            ruta_pdf = os.path.join(carpeta, nombre_pdf)

            # Crear el PDF en la ruta especificada
            pdf = canvas.Canvas(ruta_pdf, pagesize=landscape(legal))
            width, height = landscape(legal)
            logo_path = "logo.png"

            def dibujar_encabezado(pdf, width, height, headers):
                # Dibujar logo
                image_width = 75
                image_height = 75
                pdf.drawImage(logo_path, 50, height - image_height - 10, width=image_width, height=image_height)

                # Título del reporte
                pdf.setFont("Helvetica-Bold", 20)
                pdf.drawCentredString(width / 2, height - 60, "Reporte de Citas de Pacientes")

                # Encabezados de la tabla (omitiendo primera y última columna)
                x = 30
                y = height - 100
                row_height = 30
                col_width = 126

                pdf.setFillColorRGB(0.1, 0.4, 0.7)
                pdf.setStrokeColorRGB(0, 0, 0)
                pdf.rect(x, y - row_height, col_width * (len(headers) - 2), row_height, fill=1)
                pdf.setFillColorRGB(1, 1, 1)

                # Reducir el tamaño de fuente de los encabezados a 10
                pdf.setFont("Helvetica-Bold", 10)

                for col, header in enumerate(headers):
                    # Omitir primera y última columna
                    if col == 0 or col == len(headers) - 1:
                        continue
                    pdf.drawCentredString(x + (col - 1) * col_width + col_width / 2, y - row_height / 2, header)

                return x, y - 2 * row_height

            # Obtener encabezados (omitiendo la primera y última columna)
            headers = [self.bd_pacientes_citas.horizontalHeaderItem(i).text() for i in
                       range(self.bd_pacientes_citas.columnCount())]

            # Dibujar encabezado en la primera página
            x, y = dibujar_encabezado(pdf, width, height, headers)

            # Configuraciones iniciales
            row_height = 30
            col_width = 126

            # Iterar por todas las filas
            for row in range(self.bd_pacientes_citas.rowCount()):
                # Verificar si necesitamos una nueva página
                if y - row_height < 50:  # Margen inferior
                    pdf.showPage()
                    # Redibujar encabezados en la nueva página
                    x, y = dibujar_encabezado(pdf, width, height, headers)

                # Dibujar fila
                pdf.setFillColorRGB(0, 0, 0)
                for col in range(self.bd_pacientes_citas.columnCount()):
                    # Omitir primera y última columna
                    if col == 0 or col == self.bd_pacientes_citas.columnCount() - 1:
                        continue

                    # Dibujar rectángulo de celda
                    pdf.rect(x + (col - 1) * col_width, y - row_height, col_width, row_height, stroke=1, fill=0)

                    # Obtener y dibujar texto
                    item = self.bd_pacientes_citas.item(row, col)
                    if item is not None:
                        text = item.text()
                        pdf.drawCentredString(x + (col - 1) * col_width + col_width / 2, y - row_height / 2, text)

                # Mover coordenada Y hacia abajo
                y -= row_height

            # Guardar el PDF
            pdf.save()

            QMessageBox.about(self, 'Aviso', f'PDF Generado correctamente en: {ruta_pdf}')
        except Exception as e:
            QMessageBox.about(self, 'Error', str(e))

    def ventana_fecha(self):
        try:
            from .pdf_bitacora import PDFBitacora
            self.PDFBitacora = PDFBitacora()
            self.PDFBitacora.show()
        except Exception as e:
            QMessageBox.about(self, 'Error', e)

    def obtener_conexion(self):
        return pymysql.connect(
            host = "localhost",
            user = "root",
            password = "2332",
            db = "bdhidrocolon"
        )

    def obtener_datos(self, tabla):
        conexion = self.obtener_conexion()
        datos = []
        with conexion.cursor() as cursor:
            cursor.execute(f"SELECT * FROM {tabla}")
            resultados = cursor.fetchall()
            for row in resultados:
                datos.append(row)
        conexion.close()
        return datos

    def excel_farmacia(self):
        datos = self.obtener_datos("medicamentos")
        libro = openpyxl.Workbook()
        hoja = libro.active

        hoja["A1"] = "Id"
        hoja["B1"] = "Codigo"
        hoja["C1"] = "Nombre"
        hoja["D1"] = "Presentacion"
        hoja["E1"] = "Laboratorio"
        hoja["F1"] = "Existencias"
        hoja["G1"] = "Fecha"
        hoja["H1"] = "Tarjeta"
        hoja["I1"] = "Efectivo"

        for i, row in enumerate(datos):
            hoja[f"A{i + 2}"] = row[0]
            hoja[f"B{i + 2}"] = row[1]
            hoja[f"C{i + 2}"] = row[2]
            hoja[f"D{i + 2}"] = row[3]
            hoja[f"E{i + 2}"] = row[4]
            hoja[f"F{i + 2}"] = row[5]
            hoja[f"G{i + 2}"] = row[6]
            hoja[f"H{i + 2}"] = row[7]
            hoja[f"I{i + 2}"] = row[8]

        # Obtener ruta del escritorio
        desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
        archivo_path = os.path.join(desktop, "medicamentos_Excel.xlsx")

        # Guardar archivo
        libro.save(archivo_path)
        QMessageBox.information(self, "Aviso", f"Excel de farmacia creado correctamente en: {archivo_path}")

    def excel_jornadas(self):
        datos = self.obtener_datos("jornadas")
        libro = openpyxl.Workbook()
        hoja = libro.active

        hoja["A1"] = "Id"
        hoja["B1"] = "Nombre"
        hoja["C1"] = "Tarjeta"
        hoja["D1"] = "Efectivo"

        for i, row in enumerate(datos):
            hoja[f"A{i + 2}"] = row[0]
            hoja[f"B{i + 2}"] = row[1]
            hoja[f"C{i + 2}"] = row[2]
            hoja[f"D{i + 2}"] = row[3]

        # Obtener ruta del escritorio
        desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
        archivo_path = os.path.join(desktop, "Jornadas_Excel.xlsx")

        # Guardar archivo
        libro.save(archivo_path)
        QMessageBox.information(self, "Aviso", f"Excel de Jornadas creado correctamente en: {archivo_path}")

    def excel_pacientes(self):
        datos = self.obtener_datos("paciente")
        libro = openpyxl.Workbook()
        hoja = libro.active

        hoja["A1"] = "Id"
        hoja["B1"] = "Nombre"
        hoja["C1"] = "Apellido"
        hoja["D1"] = "Telefono"
        hoja["E1"] = "DPI"
        hoja["F1"] = "Cita"
        hoja["G1"] = "Cumpleaños"

        for i, row in enumerate(datos):
            hoja[f"A{i + 2}"] = row[0]
            hoja[f"B{i + 2}"] = row[1]
            hoja[f"C{i + 2}"] = row[2]
            hoja[f"D{i + 2}"] = row[3]
            hoja[f"E{i + 2}"] = row[4]
            hoja[f"F{i + 2}"] = row[5]
            hoja[f"G{i + 2}"] = row[6]

        # Obtener ruta del escritorio
        desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
        archivo_path = os.path.join(desktop, "pacientes_Excel.xlsx")

        # Guardar archivo
        libro.save(archivo_path)
        QMessageBox.information(self, "Aviso", f"Excel de pacientes creado correctamente en: {archivo_path}")

    def excel_combos(self):
        datos = self.obtener_datos("promociones")
        libro = openpyxl.Workbook()
        hoja = libro.active

        hoja["A1"] = "Id"
        hoja["B1"] = "Nombre"
        hoja["C1"] = "Tarjeta"
        hoja["D1"] = "Efectivo"

        for i, row in enumerate(datos):
            hoja[f"A{i + 2}"] = row[0]
            hoja[f"B{i + 2}"] = row[1]
            hoja[f"C{i + 2}"] = row[2]
            hoja[f"D{i + 2}"] = row[3]

        # Obtener ruta del escritorio
        desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
        archivo_path = os.path.join(desktop, "Combos_Excel.xlsx")

        # Guardar archivo
        libro.save(archivo_path)
        QMessageBox.information(self, "Aviso", f"Excel de combos creado correctamente en: {archivo_path}")

    def excel_terapias(self):
        datos = self.obtener_datos("terapias")
        libro = openpyxl.Workbook()
        hoja = libro.active

        hoja["A1"] = "Id"
        hoja["B1"] = "Nombre"
        hoja["C1"] = "Tarjeta"
        hoja["D1"] = "Efectivo"

        for i, row in enumerate(datos):
            hoja[f"A{i + 2}"] = row[0]
            hoja[f"B{i + 2}"] = row[1]
            hoja[f"C{i + 2}"] = row[2]
            hoja[f"D{i + 2}"] = row[3]

        # Obtener ruta del escritorio
        desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
        archivo_path = os.path.join(desktop, "terapias_Excel.xlsx")

        # Guardar archivo
        libro.save(archivo_path)
        QMessageBox.information(self, "Aviso", f"Excel de terapias creado correctamente en: {archivo_path}")

    def IniciarMod(self):
        try:
            self.cargarTablacarrito()
            self.switch_window.emit('medicina')
            self.cargarTablacarrito()
        except Exception as e:
            print(e)

    def IniciarContra_des(self):
        try:
            # Crear nueva instancia pasando referencia a self
            self.descu = DescuentoMedi(self)
            self.switch_window.emit('descu_contra')
        except Exception as e:
            print(e)

    def FinalizarMod(self):
        try:
            self.borrar_tabla()
            self.show_page()
            self.switch_window.emit('sesion')
        except Exception as e:
            print(e)

    def IniciarTerapia(self):
        self.cargarTablacarrito()
        self.switch_window.emit('terapia')
        self.cargarTablacarrito()

    def IniciarJornadas(self):
        manager = sql_structures.Manager()
        dato = manager.print_table('carrito')
        self.bd_carrito.setRowCount(len(dato))
        self.label_3.setStyleSheet('font: 9pt "Nunito";')
        self.label_3.setText(str(len(dato)))
        self.switch_window.emit('jornada')

    def IniciarPacientes(self):
        self.switch_window.emit('paciente')
        manager = sql_structures.Manager()
        dato = manager.print_table('carrito')
        self.bd_carrito.setRowCount(len(dato))

    def IniciarCombos(self):
        self.switch_window.emit('combo')
        manager = sql_structures.Manager()
        dato = manager.print_table('carrito')
        self.bd_carrito.setRowCount(len(dato))
        self.label_90.setStyleSheet('font: 9pt "Nunito";')
        self.label_90.setText(str(len(dato)))

    def IniciarDes(self):
        from .descuentosMedi import DescuentoMedi
        self.DescuentoMedi = DescuentoMedi(self)  # Pasar referencia a self
        self.DescuentoMedi.show()

    def IniciarExt(self):
        self.Extras = VentanaGestionExtras()

        self.Extras.show()

    def Iniciar_ing_efec(self):
        from .efectivo import CashRegisterApp
        self.efec = CashRegisterApp()
        self.efec.show()

    def Iniciar_ing_tar(self):
        from .tarjeta import IngresoTarjetaApp
        self.tar = IngresoTarjetaApp()
        self.tar.show()

    def Iniciar_ing_trasf(self):
        from .trasfere import IngresoTransferenciaApp
        self.tras = IngresoTransferenciaApp()
        self.tras.show()

    def Iniciar_ing_gastos(self):
        from .gastos import GastosApp
        self.gastos = GastosApp()
        self.gastos.show()

    def Iniciar_ing_crudo(self):
        from .cierre_crudo import CierreApp
        self.cierre = CierreApp()
        self.cierre.show()

    def Iniciar_ing_cheques(self):
        from .cheques import RegistroCheques
        self.cheque = RegistroCheques()
        self.cheque.show()

    def Iniciar_ing_reporte(self):
        from .Reporte import AplicacionReporteFinanciero
        self.reporte = AplicacionReporteFinanciero()
        self.reporte.show()

    def Iniciar_ing_comision(self):
        from .comision import VentasComisionesApp
        self.comi = VentasComisionesApp()
        self.comi.show()

    def Devolu(self):
        self.switch_window.emit('contra')

    def Devolu_total(self):
        # from.codigoAnulacion import CodigoAnulacion
        # self.CodigoAnulacion = CodigoAnulacion()
        # self.CodigoAnulacion.show()
        self.switch_window.emit('contra_total')

    def iniciarDatos(self):
        try:
            from .datos_cliente import DatosCliente
            self.DatosCliente = DatosCliente()

            # AUTO-COMPLETAR DATOS SI HAY PACIENTE VINCULADO
            if self.paciente_actual:
                try:
                    # Obtener información completa del paciente
                    manager = sql_structures.Manager()
                    paciente_info = manager.obtener_paciente_por_id(self.paciente_actual['id'])

                    if paciente_info:
                        nombre_completo = f"{paciente_info[1]} {paciente_info[2]}"
                        telefono = str(paciente_info[3]) if len(paciente_info) > 3 and paciente_info[3] else ""
                        dpi = str(paciente_info[4]) if len(paciente_info) > 4 and paciente_info[4] else ""

                        # USAR LOS NOMBRES CORRECTOS DE LOS CAMPOS CON CONVERSIÓN A STRING

                        # Nombre completo
                        self.DatosCliente.le_agNombrePaciente.setText(str(nombre_completo))
                        print(f"✅ Nombre completado: {nombre_completo}")

                        # Teléfono (convertir a string)
                        self.DatosCliente.le_agApellido.setText(telefono)
                        print(f"✅ Teléfono completado: {telefono}")

                        # NIT (usar DPI convertido a string)
                        self.DatosCliente.le_NIT.setText(dpi)
                        print(f"✅ NIT completado: {dpi}")

                        # Dirección por defecto
                        self.DatosCliente.le_agDpi_2.setText("Ciudad")
                        print(f"✅ Dirección completada: Ciudad de Guatemala")

                        # Mensaje de confirmación
                        self.mostrar_notificacion_sutil(
                            f"📋 Datos de {self.paciente_actual['nombre_completo']} auto-completados")

                except Exception as e:
                    print(f"Error al auto-completar datos: {e}")
                    import traceback
                    traceback.print_exc()

            # Código original
            self.label_28.setText("")
            self.label_29.setText("")
            self.label_30.setText("")
            self.label_33.setText("")
            self.DatosCliente.show()
            self.bd_carrito.update()
            self.bd_carrito.repaint()
            self.limpiar_tabla(self.bd_carrito)

        except Exception as e:
            print(f"Error en iniciarDatos: {e}")
            import traceback
            traceback.print_exc()

    def obtener_datos_paciente_para_factura(self):
        """Obtiene los datos del paciente actual formateados para facturación"""
        if not self.paciente_actual:
            return None

        try:
            manager = sql_structures.Manager()
            paciente_info = manager.obtener_paciente_por_id(self.paciente_actual['id'])

            if paciente_info:
                return {
                    'nombre_completo': f"{paciente_info[1]} {paciente_info[2]}",
                    'telefono': paciente_info[3] if len(paciente_info) > 3 else "",
                    'dpi': paciente_info[4],
                    'direccion': ""  # Por ahora vacía, se puede agregar después
                }
        except Exception as e:
            print(f"Error al obtener datos del paciente: {e}")

        return None

    def IniciarEX(self):
        try:
            from .ModificarExistencia import Existencias
            self.Existencias = Existencias()
            self.Existencias.show()
            self.Existencias.recibir2(str(self.id_venta))
        except Exception as e:
            print('errorrrrrr' + e)

    def IniciarUsuarios(self):
        self.switch_window.emit('usuarios')

    def show_page_farmacia(self):
        self.stackedWidget.setCurrentWidget(self.page_farmacia)
        self.FrameFarm.setStyleSheet("""QFrame {background-color: #88cc14;}""")
        self.FrameTera.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameJorna.setStyleSheet("""QFrame {background-color: white;}""")
        self.FramePac.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCombo.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: white;}""")
        self.cargarTablaFarmacia()

    def show_page(self):
        self.stackedWidget.setCurrentWidget(self.page_new)

    def show_page_terapias(self):
        self.stackedWidget.setCurrentWidget(self.page_terapias)
        self.FrameTera.setStyleSheet("""QFrame {background-color: #88cc14;}""")
        self.FrameFarm.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameJorna.setStyleSheet("""QFrame {background-color: white;}""")
        self.FramePac.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCombo.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: white;}""")

    def show_page_jornadas(self):
        self.stackedWidget.setCurrentWidget(self.page_jornadas)
        self.FrameJorna.setStyleSheet("""QFrame {background-color: #88cc14;}""")
        self.FrameTera.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameFarm.setStyleSheet("""QFrame {background-color: white;}""")
        self.FramePac.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCombo.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: white;}""")

    def show_page_pacientes(self):
        self.stackedWidget.setCurrentWidget(self.page_pacientes)
        self.FramePac.setStyleSheet("""QFrame {background-color: #88cc14;}""")
        self.FrameTera.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameFarm.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameJorna.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCombo.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: white;}""")
        self.cargarTablaPacientes()

    def show_page_combos(self):
        self.stackedWidget.setCurrentWidget(self.page_combos)
        self.FrameCombo.setStyleSheet("""QFrame {background-color: #88cc14;}""")
        self.FramePac.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameTera.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameFarm.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameJorna.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: white;}""")
        self.cargarTablaCombo()

    def show_page_usuarios(self):
        self.stackedWidget.setCurrentWidget(self.page_usuarios)
        self.FrameCombo.setStyleSheet("""QFrame {background-color: white;}""")
        self.FramePac.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameTera.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameFarm.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameJorna.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: #88cc14;}""")
        self.cargarTablaUsuario()

    def show_page_cierre(self):
        self.stackedWidget.setCurrentWidget(self.page_cierre_inventario)
        self.FrameCombo.setStyleSheet("""QFrame {background-color: white;}""")
        self.FramePac.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameTera.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameFarm.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameJorna.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: #88cc14;}""")

    def show_page_cita(self):
        self.stackedWidget.setCurrentWidget(self.page_citas)
        self.FramePac.setStyleSheet("""QFrame {background-color: #88cc14;}""")
        self.FrameTera.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameFarm.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameJorna.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCombo.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: white;}""")
        self.cargarTablaPacientes()

    def show_page_cumple(self):
        self.stackedWidget.setCurrentWidget(self.page_cumple)
        self.FramePac.setStyleSheet("""QFrame {background-color: #88cc14;}""")
        self.FrameTera.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameFarm.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameJorna.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCombo.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: white;}""")
        self.cargarTablaPacientes()

    def show_page_fina(self):
        self.stackedWidget.setCurrentWidget(self.page_fina)
        self.FramePac.setStyleSheet("""QFrame {background-color: #88cc14;}""")
        self.FrameTera.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameFarm.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameJorna.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCombo.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameUsuarios.setStyleSheet("""QFrame {background-color: white;}""")
        self.FrameCierre.setStyleSheet("""QFrame {background-color: white;}""")
        #self.cargarTablaPacientes()


    def show_page_vitacora(self):
        self.stackedWidget.setCurrentWidget(self.page_vitacora)
        self.cargarTablavitacora()

    def show_page_carrito(self):
        self.stackedWidget.setCurrentWidget(self.page_carrito)
        self.radio_tarjeta.setChecked(True)
        self.femaleselected(True)

    def check_expiration_date(self, fecha_vencimiento, casa_medica):
        """
        Verifica si el medicamento está próximo a vencer según la casa médica
        """
        try:
            # Convertir la fecha de string a QDate
            fecha_venc = QDate.fromString(fecha_vencimiento, "yyyy-MM-dd")
            fecha_actual = QDate.currentDate()

            # Calcular días hasta vencimiento
            dias_hasta_vencimiento = fecha_actual.daysTo(fecha_venc)

            # Verificar según la casa médica
            if casa_medica.lower() == "hidrocolon":
                return dias_hasta_vencimiento <= 30  # 1 mes
            else:
                return dias_hasta_vencimiento <= 120  # 4 meses
        except Exception as e:
            print(f"Error al verificar fecha: {e}")
            return False

    def cargarTablaFarmacia(self, mostrar_mensaje_stock=True, mostrar_mensaje_vencimiento=True):
        """Cargar tabla farmacia con selector de columnas - VERSIÓN SIMPLE"""
        try:
            # 1. PREPARAR TABLA CON NUEVAS COLUMNAS
            if hasattr(self, 'farmacia_column_manager'):
                self.farmacia_column_manager.apply_column_visibility()

            # 2. USAR TU MÉTODO EXISTENTE QUE YA FUNCIONA
            self.cargarTablaFarmacia_sin()

            # 3. AGREGAR LAS COLUMNAS ADICIONALES
            self.agregar_columnas_adicionales()

            print("✅ Tabla de farmacia cargada con selector de columnas")

        except Exception as e:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.about(self, 'Aviso', f'Error de carga de tabla: {str(e)}')
            print(f"⚠️ Error en cargarTablaFarmacia: {e}")

    def agregar_columnas_adicionales(self):
        """Agregar las columnas adicionales de medicamentos"""
        try:
            import sql_structures
            from PyQt5.QtWidgets import QTableWidgetItem

            manager = sql_structures.Manager()

            # ✅ AQUÍ VA EL QUERY ACTUALIZADO CON COSTO
            query = """
                    SELECT id, \
                           nombre, \
                           presentacion, \
                           laboratorio, \
                           existencias, \
                           fecha, \
                           tarjeta, \
                           efectivo,
                           indicacion, \
                           contra, \
                           dosis, \
                           comision, \
                           costo
                    FROM medicamentos
                    ORDER BY nombre \
                    """
            manager.cursor.execute(query)
            datos_completos = manager.cursor.fetchall()

            # Solo proceder si tenemos datos
            if not datos_completos:
                return

            # Verificar si tenemos el gestor de columnas
            if not hasattr(self, 'farmacia_column_manager'):
                return

            visible_columns = self.farmacia_column_manager.visible_columns
            column_mapping = self.farmacia_column_manager.column_mapping

            # Procesar cada fila
            for i, fila_datos in enumerate(datos_completos):
                if i >= self.bd_farmacia.rowCount():
                    break

                # Agregar datos para cada columna según visibilidad
                for j, valor in enumerate(fila_datos):
                    # ✅ MAPEO ACTUALIZADO CON COSTO
                    columnas_bd = ['id', 'nombre', 'presentacion', 'laboratorio', 'existencias',
                                   'fecha', 'tarjeta', 'efectivo', 'indicacion', 'contra', 'dosis', 'costo', 'comision']

                    if j < len(columnas_bd):
                        col_key = columnas_bd[j]

                        # Solo establecer si la columna es visible
                        if visible_columns.get(col_key, False):
                            col_info = column_mapping.get(col_key)
                            if col_info:
                                cell_value = str(valor) if valor is not None else ""
                                self.bd_farmacia.setItem(i, col_info['index'], QTableWidgetItem(cell_value))

            print("✅ Columnas adicionales agregadas exitosamente")

        except Exception as e:
            print(f"⚠️ Error agregando columnas adicionales: {e}")

    def cargarTablaFarmacia_sin(self):
        try:
            manager = sql_structures.Manager()
            dato = manager.print_table_farmacia()
            self.bd_farmacia.setRowCount(len(dato))

            # Variables para controlar si ya se mostró el mensaje
            mostrar_mensaje_stock = True
            mostrar_mensaje_vencimiento = True

            # Crear el controlador de resaltado si no existe
            if not hasattr(self, 'row_highlighter'):
                self.row_highlighter = RowHighlighter(self.bd_farmacia)
            else:
                # No es necesario limpiar ya que mantendremos los estados
                pass

            for fila, listaItem in enumerate(dato):
                # Crear botones y layout
                self.bd_farmacia.blockSignals(True)
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                    QPushButton:pressed{background-color: #88cc14}"""), icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                     QPushButton:pressed{background-color:  #FE2C55}"""), icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                btn_tres = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #4e84f2}
                                                                     QPushButton:pressed{background-color:  #4e84f2}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\detalle.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)
                layout.setSpacing(0)
                layout.addWidget(btn_uno)
                layout.addWidget(btn_dos)
                layout.addWidget(btn_tres)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_farmacia.setCellWidget(fila, self.bd_farmacia.columnCount() - 1, widget)
                self.boton_uno_coneccion_far(btn_uno, fila, self.bd_farmacia)
                self.boton_dos_connecion_far(btn_dos, fila, self.bd_farmacia)
                self.boton_tres_connecion_far(btn_tres, fila, self.bd_farmacia)

            for i in range(len(dato)):
                for j in range(8):  # Asumiendo 8 columnas de datos
                    self.bd_farmacia.setItem(i, j, QTableWidgetItem(str(dato[i][j])))

                # Verificar condiciones de alerta
                stock = int(dato[i][4])  # Stock
                casa_medica = str(dato[i][3])  # Casa médica
                fecha_vencimiento = str(dato[i][5])  # Fecha de vencimiento
                nombre_medicamento = str(dato[i][2])  # Nombre del medicamento

                # Verificar stock bajo
                if stock <= 10:
                    self.row_highlighter.highlight_row(i, True)  # True para alerta de stock
                    if mostrar_mensaje_stock:
                        #QMessageBox.about(self, 'Aviso', 'Stock bajo -- Revisar')
                        mostrar_mensaje_stock = False  # Ya no mostrar más mensajes de stock

                # Verificar fecha de vencimiento
                if self.check_expiration_date(fecha_vencimiento, casa_medica):
                    if i not in self.row_highlighter.highlighted_rows:
                        self.row_highlighter.highlight_row(i, False)  # False para alerta de vencimiento
                    if mostrar_mensaje_vencimiento:
                        #QMessageBox.about(self, 'Aviso', 'Medicamentos próximos a vencer -- Revisar')
                        mostrar_mensaje_vencimiento = False  # Ya no mostrar más mensajes de vencimiento

            self.bd_farmacia.blockSignals(False)
            # Conectar señal para manejar cambios del usuario
            self.bd_farmacia.itemChanged.connect(self.on_item_changed)

            # Restaurar los resaltados después de cargar los datos
            self.row_highlighter.restore_highlights()

        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def click_tabla_medicamento(self, row, column):
        """Manejar clics en tabla medicamentos - VERSIÓN SIMPLE"""
        try:
            import sql_structures
            self.id_c = 0
            manager = sql_structures.Manager()
            item = self.bd_farmacia.item(row, column)

            if not item:
                return

            value = item.text()
            columns_ingreso = ['id', 'nombre', 'presentacion', 'laboratorio', 'existencias',
                               'fecha', 'tarjeta', 'efectivo', 'indicacion', 'contra', 'dosis', 'comision']

            header_item = self.bd_farmacia.horizontalHeaderItem(column)
            if not header_item:
                return

            column_name = header_item.text()

            # Mapeo de nombres de columnas (EXTENDIDO)
            if column_name == 'ID':
                self.id_c = manager.dar_id_normal('medicamentos', value, 'id')
                self.cabeza = 'id'
            elif column_name == 'Nombre':
                self.id_c = manager.dar_id_normal('medicamentos', value, 'nombre')
                self.cabeza = 'nombre'
            elif column_name == 'Presentación':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'presentacion')
                self.cabeza = 'presentacion'
            elif column_name == 'Laboratorio':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'laboratorio')
                self.cabeza = 'laboratorio'
            elif column_name == 'Existencias':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'existencias')
                self.cabeza = 'existencias'
            elif column_name == 'Fecha de Vencimiento':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'fecha')
                self.cabeza = 'fecha'
            elif column_name == 'Precio Tarjeta':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'tarjeta')
                self.cabeza = 'tarjeta'
            elif column_name == 'Precio Efectivo':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'efectivo')
                self.cabeza = 'efectivo'
            # NUEVAS COLUMNAS
            elif column_name == 'Indicación':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'indicacion')
                self.cabeza = 'indicacion'
            elif column_name == 'Contraindicación':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'contra')
                self.cabeza = 'contra'
            elif column_name == 'Dosis':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'dosis')
                self.cabeza = 'dosis'
            elif column_name == 'Comisión':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'comision')
                self.cabeza = 'comision'
            elif column_name == 'Costo':
                self.id_c = manager.get('medicamentos', columns_ingreso, value, 'costo')
                self.cabeza = 'costo'

            # Obtener ID de la fila para otras operaciones
            id_item = self.bd_farmacia.item(row, 0)
            if id_item:
                self.id_venta = id_item.text()

        except Exception as e:
            print(f"⚠️ Error en click_tabla_medicamento: {e}")

    def agregar_medicamento_a_carrito(self, cantidad, id):
        from .ModificarExistencia import Existencias
        self.Existencias = Existencias()
        try:
            manager = sql_structures.Manager()

            carrito_actual = manager.get_carrito("carrito", "medicamentos_id", id)

            pre = manager.get_presentacion("medicamentos", "id", id)
            print(id)
            print(pre)
            item = manager.get_carrito_medic("medicamentos", 'id', id, "presentacion", pre)
            print(item)
            if not item:
                print("No se encontró el medicamento")
                return

            existencias = item[0][1]  # item[0][1] es existencias
            if existencias < cantidad:
                print(f"No hay suficientes existencias. Disponible: {existencias}, Solicitado: {cantidad}")
                return

            nombre = item[0][0]  # Nombre
            tarjeta = item[0][2]  # Precio tarjeta
            efectivo = item[0][3]  # Precio efectivo

            # Reducir existencias ANTES de actualizar el carrito
            nuevas_existencias = existencias - cantidad
            self.medicamento.actualizarMedicamentor(id, nuevas_existencias, "Existencias")

            if carrito_actual:
                print(312456)
                cantidad_actual = carrito_actual[0][2]  # existencias en el carrito
                print(cantidad_actual)
                nueva_cantidad = cantidad_actual + cantidad
                print(nueva_cantidad)
                nuevo_tarjeta = nueva_cantidad * tarjeta
                nuevo_efectivo = nueva_cantidad * efectivo
                # Actualizar el registro existente en el carrito
                self.carrito.actualizar_a_carrito_nuevo(id, nuevo_tarjeta, nuevo_efectivo, nueva_cantidad)

            else:
                precio_tarjeta = cantidad * tarjeta
                precio_efectivo = cantidad * efectivo

                self.carrito.agregar_a_carrito(nombre, precio_efectivo, cantidad, precio_tarjeta, id,
                                               -1, -1, -1, -1, -1)

            # CONSUMIR EXTRAS ASOCIADOS AL MEDICAMENTO
            resultado_extras = self.consumir_extras_medicamento(id, cantidad, nombre)
            print(resultado_extras)
            if not resultado_extras[0]:
                print(f"Advertencia con extras: {resultado_extras[1]}")
            else:
                print(f"Extras procesados: {resultado_extras[1]}")

            self.cargarTablacarrito()
            # Actualizar totales después de agregar
            self.actualizar_totales_carrito()
            self.Existencias.close()

        except Exception as e:
            print(f"Error al agregar al carrito: {str(e)}")

    def consumir_extras_medicamento(self, medicamento_id, cantidad_vendida, nombre_medicamento):
        """Consume los extras asociados a un medicamento cuando se agrega al carrito"""

        # Configuración de base de datos
        db_config = {
            'host': "127.0.0.1",
            'user': "root",
            'password': "2332",  # Cambia por tu contraseña
            'database': "bdhidrocolon",
            'charset': 'utf8mb4'
        }

        try:
            import pymysql
            connection = pymysql.connect(**db_config)
            cursor = connection.cursor()

            # 1. OBTENER EXTRAS ASOCIADOS AL MEDICAMENTO
            extras_asociados = []
            try:
                cursor.execute("SELECT extra_nombre FROM medicamento_extras WHERE medicamento_id = %s",
                               (medicamento_id,))
                extras_asociados = [row[0] for row in cursor.fetchall()]
            except Exception as e:
                # La tabla medicamento_extras puede no existir aún
                print(f"Tabla medicamento_extras no encontrada: {e}")
                cursor.close()
                connection.close()
                return True, "Sin extras (tabla no encontrada)"

            if not extras_asociados:
                cursor.close()
                connection.close()
                return True, "Sin extras asociados"

            print(f"Extras asociados al medicamento {medicamento_id}: {extras_asociados}")

            # 2. CONSUMIR CADA EXTRA
            extras_consumidos = []
            extras_sin_stock = []

            for extra_nombre in extras_asociados:
                for _ in range(cantidad_vendida):  # Consumir según cantidad vendida
                    # Verificar stock actual
                    cursor.execute("SELECT id, cantidad FROM extras WHERE nombre=%s AND activo=1", (extra_nombre,))
                    resultado = cursor.fetchone()

                    if resultado:
                        extra_id, stock_actual = resultado
                        if stock_actual > 0:
                            # Reducir stock en 1
                            nuevo_stock = stock_actual - 1
                            cursor.execute("UPDATE extras SET cantidad=%s WHERE id=%s", (nuevo_stock, extra_id))

                            # Registrar movimiento
                            try:
                                cursor.execute('''
                                    INSERT INTO movimientos_stock_extras (extra_id, tipo, cantidad, observaciones)
                                    VALUES (%s, 'salida', 1, %s)
                                ''', (extra_id, f'Venta medicamento: {nombre_medicamento} (ID: {medicamento_id})'))
                            except Exception as mov_error:
                                print(f"Error registrando movimiento: {mov_error}")

                            if extra_nombre not in extras_consumidos:
                                extras_consumidos.append(extra_nombre)
                        else:
                            if extra_nombre not in extras_sin_stock:
                                extras_sin_stock.append(extra_nombre)
                    else:
                        if extra_nombre not in extras_sin_stock:
                            extras_sin_stock.append(f"{extra_nombre} (no encontrado)")

            connection.commit()
            cursor.close()
            connection.close()

            # 3. PREPARAR MENSAJE DE RESULTADO
            mensaje_partes = []
            if extras_consumidos:
                mensaje_partes.append(f"Extras consumidos: {', '.join(extras_consumidos)}")
            if extras_sin_stock:
                mensaje_partes.append(f"Sin stock: {', '.join(extras_sin_stock)}")

            mensaje = " | ".join(mensaje_partes) if mensaje_partes else "Sin extras procesados"

            return len(extras_sin_stock) == 0, mensaje

        except Exception as e:
            print(f"Error al consumir extras: {e}")
            return False, f"Error al consumir extras: {e}"

    def agregar_medicamento_regalo_a_carrito(self, cantidad, id):
        """Método específico para agregar medicamentos de regalo al carrito con precio 0"""
        from .ModificarExistencia import Existencias
        self.Existencias = Existencias()

        self.db_config = {
            'host': "127.0.0.1",
            'user': "root",
            'password': "2332",  # Cambia por tu contraseña
            'database': "bdhidrocolon",
            'charset': 'utf8mb4'
        }

        try:
            manager = sql_structures.Manager()

            carrito_actual = manager.get_carrito("carrito", "medicamentos_id", id)

            pre = manager.get_presentacion("medicamentos", "id", id)
            item = manager.get_carrito_medic("medicamentos", 'id', id, "presentacion", pre)

            if not item:
                print("No se encontró el medicamento")
                return

            existencias = item[0][1]  # existencias disponibles
            if existencias < cantidad:
                print(f"No hay suficientes existencias. Disponible: {existencias}, Solicitado: {cantidad}")
                return

            nombre = item[0][0]   # Agregar indicador de regalo
            # Precios siempre en 0 para regalos
            tarjeta = 0
            efectivo = 0

            # Reducir existencias
            nuevas_existencias = existencias - cantidad
            self.medicamento.actualizarMedicamentor(id, nuevas_existencias, "Existencias")

            if carrito_actual:
                # Si ya existe en carrito, actualizar cantidad pero mantener precio 0
                cantidad_actual = carrito_actual[0][1]
                nueva_cantidad = cantidad_actual + cantidad

                self.carrito.actualizar_a_carrito_nuevo(id, 0, 0, nueva_cantidad)

            else:
                # Agregar nuevo medicamento de regalo
                self.carrito.agregar_a_carrito(nombre, 0, cantidad, 0, id,
                                               -1, -1, -1, -1, -1
                                               )

            self.cargarTablacarrito()
            self.Existencias.close()
            print(f"Medicamento agregado: {nombre} - Cantidad: {cantidad}")

        except Exception as e:
            print(f"Error al agregar medicamento de regalo al carrito: {str(e)}")

    def procesar_registro_combos(self, jornada_id, requiere_medicamentos, medicamentos):
        try:
            manager = sql_structures.Manager()
            item = manager.get_carrito_jo("jornadas", "id", jornada_id)

            # Agregar la jornada al carrito
            self.carrito.agregar_a_carrito(item[0][0], item[0][2], 1, item[0][1], -1, jornada_id,
                                           -1, -1, -1, -1)

            # Si el combo requiere medicamentos, agregarlos como regalo
            if requiere_medicamentos and medicamentos:
                for med in medicamentos:
                    # Usar el método específico para regalos
                    self.agregar_medicamento_regalo_a_carrito(med['cantidad'], med['id'])

            QMessageBox.information(self, "Éxito", "Jornada registrada correctamente")
            # Solo cargar una vez al final
            self.cargarTablacarrito()

        except Exception as e:
            print(f"Error al procesar registro de jornada: {e}")
            QMessageBox.critical(self, "Error", "Error al procesar el registro de la jornada")

    def modificar_existencias (self, cantidad, id):
        from .ModificarExistencia import Existencias
        self.Existencias = Existencias()
        try:
            manager = sql_structures.Manager()

            item = manager.get_carrito("medicamentos", "id", id)
            if not item:
                print("No se encontró el medicamento")
                return

            existencias = item[0][1]  # item[0][1] es existencias
            if existencias < cantidad:
                print(f"No hay suficientes existencias. Disponible: {existencias}, Solicitado: {cantidad}")
                return

            nombre = item[0][0]  # Nombre
            tarjeta = item[0][2]  # Precio tarjeta
            efectivo = item[0][3]  # Precio efectivo

            self.carrito.agregar_a_carrito(nombre, 0, cantidad, 0, id, -1, -1, -1, -1, -1)

            nuevas_existencias = existencias - cantidad
            self.medicamento.actualizarMedicamentor(id, nuevas_existencias, "existencias")

        except Exception as e:
            print(f"Error al agregar al carrito: {str(e)}")

    def busqueda_farmacia(self, letras):
        try:
            self.limpiar_tabla(self.bd_farmacia)
            manager = sql_structures.Manager()
            data = manager.busqueda_medicina(letras)
            self.bd_farmacia.setRowCount(len(data))
            for fila, listaItem in enumerate(data):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)
                layout.setSpacing(0)
                layout.addWidget(btn_uno)
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_farmacia.setCellWidget(fila, self.bd_farmacia.columnCount() - 1, widget)
                self.boton_uno_coneccion_far(btn_uno, fila, self.bd_farmacia)
                self.boton_dos_connecion_far(btn_dos, fila, self.bd_farmacia)
                for columna, item in enumerate(listaItem):
                    self.bd_farmacia.setItem(fila, columna, QTableWidgetItem(str(item)))  # insertar items
                    self.bd_farmacia.item(fila, columna).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)  # alinear items
        except Exception as e:
            print(e)
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def on_item_changed(self, item):
        try:
            row = item.row()
            col = item.column()

            # Verificar que el índice de fila sea válido
            if row < 0 or row >= self.bd_farmacia.rowCount():
                print(f"Índice de fila inválido: {row}")
                return

            # Verificar que el índice de columna sea válido
            if col < 0 or col >= self.bd_farmacia.columnCount():
                print(f"Índice de columna inválido: {col}")
                return

        except Exception as e:
            print(f"Error en on_item_changed: {e}")



    # ___________________________________________________________

    def cargarTablaJornadas(self):
        try:
            manager = sql_structures.Manager()
            dato = manager.print_table('jornadas')
            self.bd_jornadas.setRowCount(len(dato))
            for fila, listaItem in enumerate(dato):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_jornadas.setCellWidget(fila, self.bd_jornadas.columnCount() - 1, widget)
                self.boton_uno_coneccion_jorda(btn_uno, fila, self.bd_jornadas)
                self.boton_dos_connecion_jorda(btn_dos, fila, self.bd_jornadas)
                for i in range(len(dato)):
                    self.bd_jornadas.setItem(i, 0, QTableWidgetItem(str(dato[i][0])))
                    self.bd_jornadas.setItem(i, 1, QTableWidgetItem(str(dato[i][1])))
                    self.bd_jornadas.setItem(i, 2, QTableWidgetItem(str(dato[i][3])))
                    self.bd_jornadas.setItem(i, 3, QTableWidgetItem(str(dato[i][2])))
        except Exception as e:
            print(e)
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def click_tabla_jornadas(self, row, column):
        try:
            manager = sql_structures.Manager()
            item = self.bd_jornadas.item(row, column)
            value = item.text()
            columns_ingreso = ['id', 'nombre', 'tarjeta', 'efectivo']
            header_item = self.bd_jornadas.horizontalHeaderItem(column)
            column_name = header_item.text()
            if column_name == 'ID':
                self.id_c = manager.get('jornadas', columns_ingreso, value, 'id')
                self.cabeza = header_item.text()
            elif column_name == 'Nombre':
                self.id_c = manager.get('jornadas', columns_ingreso, value, 'nombre')
                self.cabeza = header_item.text()
            elif column_name == 'Tarjeta':
                self.id_c = manager.get('jornadas', columns_ingreso, value, 'tarjeta')
                self.cabeza = header_item.text()
            elif column_name == 'Efectivo':
                self.id_c = manager.get('jornadas', columns_ingreso, value, 'efectivo')
                self.cabeza = header_item.text()

        except Exception as e:
            print(e)

    def agregar_jornada_a_carrito(self):
        manager = sql_structures.Manager()
        item = []
        item = manager.get_carrito_jo("jornadas", "id", self.id_c)
        self.carrito.agregar_a_carrito(item[0][0],  item[0][2], 0, item[0][1],  -1, -1,
                                       -1, self.id_c, -1, -1)
        self.cargarTablacarrito()

    def busqueda_jornada(self, letras):
        try:
            self.limpiar_tabla(self.bd_jornadas)
            manager = sql_structures.Manager()
            data = manager.busqueda("jornadas", letras)
            self.bd_jornadas.setRowCount(len(data))
            for fila, listaItem in enumerate(data):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_jornadas.setCellWidget(fila, self.bd_jornadas.columnCount() - 1, widget)
                self.boton_uno_coneccion_far(btn_uno, fila, self.bd_jornadas)
                self.boton_dos_connecion_far(btn_dos, fila, self.bd_jornadas)
                for columna, item in enumerate(listaItem):
                    self.bd_jornadas.setItem(fila, columna, QTableWidgetItem(str(item)))  # insertar items
                    self.bd_jornadas.item(fila, columna).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)  # alinear items
        except Exception as e:
            print(e)
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def registrar_jornada(self):
        try:
            dialogo = Dialogo_jorda(self)

            if dialogo.exec_() == QDialog.Accepted:
                requiere_medicamentos = dialogo.chk_requiere_medicamentos.isChecked()
                medicamentos = dialogo.medicamentos_seleccionados

                self.procesar_registro_combos(self.id_c, requiere_medicamentos, medicamentos)

        except Exception as e:
            print(f"Error al registrar jornada: {e}")
            QMessageBox.critical(self, "Error", "Error al registrar la jornada")

    def procesar_registro_combos(self, jornada_id, requiere_medicamentos, medicamentos):
        try:
            manager = sql_structures.Manager()
            item = manager.get_carrito_jo("jornadas", "id", jornada_id)

            # Agregar la jornada al carrito
            self.carrito.agregar_a_carrito(item[0][0], item[0][2], 1, item[0][1], -1, -1,
                                           -1, jornada_id, -1, -1)

            # Si el combo requiere medicamentos, agregarlos como regalo
            if requiere_medicamentos and medicamentos:
                for med in medicamentos:
                    # Usar el método específico para regalos
                    self.agregar_medicamento_regalo_a_carrito(med['cantidad'], med['id'])

            QMessageBox.information(self, "Éxito", "Jornada registrada correctamente")
            # Solo cargar una vez al final
            self.cargarTablacarrito()

        except Exception as e:
            print(f"Error al procesar registro de jornada: {e}")
            QMessageBox.critical(self, "Error", "Error al procesar el registro de la jornada")


    # ___________________________________________________________

    def cargarTablacarrito(self, modo=None):
        try:
            manager = sql_structures.Manager()

            # Seleccionar el tipo de datos según el modo
            if modo == 'efectivo':
                dato = manager.print_table_efectivo()
            elif modo == 'tarjeta':
                dato = manager.print_table_tarjeta()
            elif modo == 'individual':
                dato = manager.print_table_carrito_individual()
            else:
                # Modo por defecto - mostrar precios individuales
                dato = manager.print_table_carrito_individual()

            self.bd_carrito.setRowCount(len(dato))
            for fila, listaItem in enumerate(dato):
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                         QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                btn_tres = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 30px} QPushButton:hover{background-color:  #2c72fe}
                                                                    QPushButton:pressed{background-color:  #2c72fe}"""),
                                            icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\1.png')
                btn_cuatro = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 30px} QPushButton:hover{background-color:  #2c72fe}
                                                                        QPushButton:pressed{background-color:  #2c72fe}"""),
                                              icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\2.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)
                layout.setSpacing(0)
                layout.addWidget(btn_dos)
                layout.addWidget(btn_tres)
                layout.addWidget(btn_cuatro)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_carrito.setCellWidget(fila, self.bd_carrito.columnCount() - 1, widget)
                self.boton_dos_connecion_carrito(btn_dos, fila, self.bd_carrito)
                self.boton_tres_connecion_carrito(btn_tres, fila, self.bd_carrito)
                self.boton_cuatro_connecion_carrito(btn_cuatro, fila, self.bd_carrito)
                btn_dos.setToolTip("Eliminar")

                for i in range(len(dato)):
                    self.bd_carrito.setItem(i, 0, QTableWidgetItem(str(dato[i][0])))
                    self.bd_carrito.setItem(i, 1, QTableWidgetItem(str(dato[i][1])))
                    self.bd_carrito.setItem(i, 2, QTableWidgetItem(str(dato[i][2])))
                    self.bd_carrito.setItem(i, 3, QTableWidgetItem(str(dato[i][3])))
                    self.bd_carrito.setItem(i, 4, QTableWidgetItem(str(dato[i][4])))

                if len(dato) == 0:
                    self.label_3.setText(str(0))
                    self.label_90.setText(str(0))
                    self.label.setText(str(0))
                    self.label_85.setText(str(0))
                else:
                    self.label_3.setStyleSheet('font: 10pt "MS Shell Dlg 2";')
                    self.label_3.setText(str(len(dato)))
                    self.label_90.setStyleSheet('font: 9pt "Nunito";')
                    self.label_90.setText(str(len(dato)))
                    self.label.setStyleSheet('font: 9pt "Nunito";')
                    self.label.setText(str(len(dato)))
                    self.label_85.setStyleSheet('font: 9pt "Nunito";')
                    self.label_85.setText(str(len(dato)))

            # Actualizar totales automáticamente
            self.actualizar_totales_carrito()

        except Exception as e:
            print(f"Error tabla carrito: {e}")
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def cargarTablacarrito_tarjeta(self):
        try:
            manager = sql_structures.Manager()
            # dato = manager.print_table('carrito')
            dato = manager.print_table_tarjeta()
            self.bd_carrito.setRowCount(len(dato))
            for fila, listaItem in enumerate(dato):

                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciad
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_carrito.setCellWidget(fila, self.bd_carrito.columnCount() - 1, widget)
                self.boton_dos_connecion_carrito(btn_dos, fila, self.bd_carrito)
                for i in range(len(dato)):
                    self.bd_carrito.setItem(i, 0, QTableWidgetItem(str(dato[i][0])))
                    self.bd_carrito.setItem(i, 1, QTableWidgetItem(str(dato[i][1])))
                    self.bd_carrito.setItem(i, 2, QTableWidgetItem(str(dato[i][2])))
                    self.bd_carrito.setItem(i, 3, QTableWidgetItem(str(dato[i][3])))
                    self.bd_carrito.setItem(i, 4, QTableWidgetItem(str(dato[i][4])))
                if len(dato) == 0:
                    self.label_3.setText(str(0))
                    self.label_90.setText(str(0))
                    self.label.setText(str(0))
                    self.label_85.setText(str(0))
                else:
                    self.label_3.setStyleSheet('font: 10pt "MS Shell Dlg 2";')
                    self.label_3.setText(str(len(dato)))
                    self.label_90.setStyleSheet('font: 9pt "Nunito";')
                    self.label_90.setText(str(len(dato)))
                    self.label.setStyleSheet('font: 9pt "Nunito";')
                    self.label.setText(str(len(dato)))
                    self.label_85.setStyleSheet('font: 9pt "Nunito";')
                    self.label_85.setText(str(len(dato)))

        except Exception as e:
            print(f"Error tabla carrito: {e}")
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def actualizar_datos_carrito(self):
        consulta = sql_structures.Manager()
        numero = consulta.contar_datos()
        self.label.setText(str(numero[0][0]))
        self.label_3.setText(str(numero[0][0]))
        self.label_90.setText(str(numero[0][0]))
        self.label_85.setText(str(numero[0][0]))

        # Actualizar totales automáticamente
        self.actualizar_totales_carrito()

    def click_tabla_carrito(self, row, column):
        try:
            # Primero validamos que la tabla tenga datos
            if self.bd_carrito.rowCount() == 0:
                print("La tabla está vacía")
                return

            # Validamos que row y column sean válidos
            if row < 0 or column < 0:
                print("Índices no válidos")
                return

            # Validamos que el item exista
            item = self.bd_carrito.item(row, column)
            if item is None:
                print(f"No hay item en la posición ({row}, {column})")
                return

            # Validamos que el header exista
            header_item = self.bd_carrito.horizontalHeaderItem(column)
            if header_item is None:
                print(f"No hay header en la columna {column}")
                return

            manager = sql_structures.Manager()
            value = item.text()
            columns_ingreso = ['id', 'nombre', 'existencias', 'tarjeta', 'efectivo',
                               'medicamentos_id', 'terapias_id', 'promociones_id',
                               'jordas_id', 'ultrasonidos_id', 'consumibles_id']
            column_name = header_item.text()

            # Usar un diccionario para simplificar la lógica
            column_mapping = {
                'ID': 'id',
                'Nombre': 'nombre',
                'Existencias': 'existencias',
                'Tarjeta': 'tarjeta',
                'Efectivo': 'efectivo'
            }

            if column_name in column_mapping:
                self.id_c = manager.get('carrito', columns_ingreso, value, column_mapping[column_name])
                self.cabeza = column_name

            else:
                print(f"Columna no reconocida: {column_name}")

        except Exception as e:
            print(f"Error: {str(e)}")

    def mostrar_dialogo_pago_dividido(self):
        try:
            # Calcular total del carrito
            total = 0
            for row in range(self.bd_carrito.rowCount()):
                total += float(self.bd_carrito.item(row, 4).text())

            if total <= 0:
                QMessageBox.warning(self, "Aviso", "No hay productos en el carrito")
                return

            # Mostrar diálogo de pago dividido
            dialogo = DialogoPagoDividido(self, total)
            if dialogo.exec_() == QDialog.Accepted:
                efectivo, tarjeta = dialogo.obtener_montos()
                if efectivo is not None and tarjeta is not None:
                    self.procesar_pago_dividido(efectivo, tarjeta)
            self.label_28.setText("")
            self.label_29.setText("")
            self.label_30.setText("")

        except Exception as e:
            QMessageBox.critical(self, "Error", "Error al procesar el pago dividido")

    def procesar_pago_dividido(self, monto_efectivo, monto_tarjeta):
        try:
            # VentanaFuncional._diferencia_efectivo = monto_tarjeta - monto_efectivo
            total_carrito = 0
            for row in range(self.bd_carrito.rowCount()):
                total_carrito += float(self.bd_carrito.item(row, 4).text())

            # Validar montos
            if abs((monto_efectivo + monto_tarjeta) - total_carrito) > 0.01:
                raise ValueError("La suma de los pagos no coincide con el total")

            # Actualizar cada item en el carrito
            for row in range(self.bd_carrito.rowCount()):
                manager = sql_structures.Manager()
                value = self.bd_carrito.item(row, 0).text()
                precio_item = float(self.bd_carrito.item(row, 4).text())

                # Calcular proporción para este item
                proporcion = precio_item / total_carrito
                efectivo_item = round(monto_efectivo * proporcion, 2)
                tarjeta_item = round(monto_tarjeta * proporcion, 2)

                self.ingresar_cierre()

            # Recargar la tabla y mostrar mensaje de éxito
            self.cargarTablacarrito()
            QMessageBox.information(self, "Éxito",
                                    f"Pago dividido procesado correctamente!\n"
                                    f"Efectivo: Q.{monto_efectivo:.2f}\n"
                                    f"Tarjeta: Q.{monto_tarjeta:.2f}")

            VentanaFuncional._diferencia_efectivo = monto_efectivo
            from .datos_cliente import DatosCliente
            self.DatosCliente = DatosCliente()
            self.DatosCliente.show()

        except Exception as e:
            QMessageBox.critical(self, "Error",
                                 "No se pudo procesar el pago dividido.\n"
                                 "Por favor, verifique los montos e intente nuevamente.")

    # ___________________________________________________________

    def cargarTablaTerapias(self):
        try:
            print(1)
            manager = sql_structures.Manager()
            dato = manager.print_table('terapias')
            self.bd_terapias.setRowCount(len(dato))

            for fila, listaItem in enumerate(dato):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)

                self.bd_terapias.setCellWidget(fila, self.bd_terapias.columnCount() - 1, widget)
                self.boton_uno_coneccion_tera(btn_uno, fila, self.bd_terapias)
                self.boton_dos_connecion_tera(btn_dos, fila, self.bd_terapias)

                for i in range(len(dato)):
                    self.bd_terapias.setItem(i, 0, QTableWidgetItem(str(dato[i][0])))
                    self.bd_terapias.setItem(i, 1, QTableWidgetItem(str(dato[i][1])))
                    self.bd_terapias.setItem(i, 2, QTableWidgetItem(str(dato[i][2])))
                    self.bd_terapias.setItem(i, 3, QTableWidgetItem(str(dato[i][3])))
                print(1)# alinear items
        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def click_tabla_terapias(self, row, column):
        try:
            manager = sql_structures.Manager()
            item = self.bd_terapias.item(row, column)
            value = item.text()
            columns_ingreso = ['id', 'nombre', 'tarjeta', 'efectivo']
            header_item = self.bd_terapias.horizontalHeaderItem(column)
            column_name = header_item.text()
            if column_name == 'ID':
                self.id_c = manager.get('terapias', columns_ingreso, value, 'id')
                self.cabeza = header_item.text()
            elif column_name == 'Nombre':
                self.id_c = manager.get('terapias', columns_ingreso, value, 'nombre')
                self.cabeza = header_item.text()
            elif column_name == 'Tarjeta':
                self.id_c = manager.get('terapias', columns_ingreso, value, 'tarjeta')
                self.cabeza = header_item.text()
            elif column_name == 'Efectivo':
                self.id_c = manager.get('terapias', columns_ingreso, value, 'efectivo')
                self.cabeza = header_item.text()
        except Exception as e:
             pass

    def agregar_terapia_a_carrito(self):
        manager = sql_structures.Manager()
        item = []
        item = manager.get_carrito_jo("terapias", "id", self.id_c)
        self.carrito.agregar_a_carrito(item[0][0],  item[0][2], 1, item[0][1],  -1, self.id_c,
                                -1, -1, -1, -1)
        self.cargarTablacarrito()

    def busqueda_terapia(self, letras):
        try:
            self.limpiar_tabla(self.bd_terapias)
            manager = sql_structures.Manager()
            data = manager.busqueda("terapias", letras)
            self.bd_terapias.setRowCount(len(data))
            for fila, listaItem in enumerate(data):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_terapias.setCellWidget(fila, self.bd_terapias.columnCount() - 1, widget)
                self.boton_uno_coneccion_far(btn_uno, fila, self.bd_terapias)
                self.boton_dos_connecion_far(btn_dos, fila, self.bd_terapias)
                for columna, item in enumerate(listaItem):
                    self.bd_terapias.setItem(fila, columna, QTableWidgetItem(str(item)))  # insertar items
                    self.bd_terapias.item(fila, columna).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)  # alinear items
        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def registrar_terapia(self):
        try:
            dialogo = DialogoCombo(self)

            if dialogo.exec_() == QDialog.Accepted:
                requiere_medicamentos = dialogo.chk_requiere_medicamentos.isChecked()
                medicamentos = dialogo.medicamentos_seleccionados
                self.procesar_registro_terapia(self.id_c, requiere_medicamentos, medicamentos)

        except Exception as e:
            print(f"Error al registrar combo: {e}")
            QMessageBox.critical(self, "Error", "Error al registrar el combo")

    def procesar_registro_terapia(self, jornada_id, requiere_medicamentos, medicamentos):
        try:
            manager = sql_structures.Manager()
            item = manager.get_carrito_jo("terapias", "id", jornada_id)

            self.carrito.agregar_a_carrito(item[0][0], item[0][2], 1, item[0][1], -1, jornada_id,
                                               -1, -1, -1, -1)

            # Si el combo requiere medicamentos, procesarlos
            if requiere_medicamentos and medicamentos:
                for med in medicamentos:
                    self.agregar_medicamento_regalo_a_carrito(med['cantidad'], med['id'])


            QMessageBox.information(self, "Éxito", "Terapia registrado correctamente")
            self.cargarTablacarrito()

        except Exception as e:
            print(f"Error al procesar registro de combo: {e}")
            QMessageBox.critical(self, "Error", "Error al procesar el registro del combo")

    def cargarTablaPacientes(self):
        try:
            manager = sql_structures.Manager()
            dato = manager.print_table('paciente')
            self.bd_pacientes.setRowCount(len(dato))
            for fila, listaItem in enumerate(dato):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_pacientes.setCellWidget(fila, self.bd_pacientes.columnCount() - 1, widget)
                self.boton_uno_coneccion_paci(btn_uno, fila, self.bd_pacientes)
                self.boton_dos_connecion_paci(btn_dos, fila, self.bd_pacientes)
                for i in range(len(dato)):
                    self.bd_pacientes.setItem(i, 0, QTableWidgetItem(str(dato[i][0])))
                    self.bd_pacientes.setItem(i, 1, QTableWidgetItem(str(dato[i][1])))
                    self.bd_pacientes.setItem(i, 2, QTableWidgetItem(str(dato[i][2])))
                    self.bd_pacientes.setItem(i, 3, QTableWidgetItem(str(dato[i][3])))
                    self.bd_pacientes.setItem(i, 4, QTableWidgetItem(str(dato[i][4])))
                    self.bd_pacientes.setItem(i, 5, QTableWidgetItem(str(dato[i][5])))
                    self.bd_pacientes.setItem(i, 6, QTableWidgetItem(str(dato[i][6])))
                    self.bd_pacientes.setItem(i, 7, QTableWidgetItem(str(dato[i][7])))

        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def cargarTablaPacientes_cita(self):
        try:
            import datetime
            manager = sql_structures.Manager()
            hoy = datetime.date.today()
            manana = hoy + datetime.timedelta(days=1)
            dato = manager.print_table_citas('paciente', manana, hoy)
            id = manager.print_table_cita(manana)
            # print(dato, manana, id)
            # Obtener la fecha de hoy
            # Obtener la fecha de mañana


            self.bd_pacientes_citas.setRowCount(len(id))

            # Crear una estructura para almacenar los estados de los checkboxes
            checkbox_states = {}

            # Guardar los estados de los checkboxes antes de recargar la tabla
            for fila in range(self.bd_pacientes_citas.rowCount()):
                checkbox_widget = self.bd_pacientes_citas.cellWidget(fila, self.bd_pacientes_citas.columnCount() - 1)
                if checkbox_widget is not None and isinstance(checkbox_widget, QWidget):
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox is not None:
                        checkbox_states[fila] = checkbox.isChecked()
                    else:
                        checkbox_states[fila] = False
                else:
                    checkbox_states[fila] = False

            # Recargar los datos en la tabla
            #for fila, listaItem in enumerate(id):
            #    for col, valor in enumerate(listaItem):
            #        self.bd_pacientes_citas.setItem(fila, col, QTableWidgetItem(str(valor)))

            for i in range(len(id)):
                self.bd_pacientes_citas.setItem(i, 0, QTableWidgetItem(str(id[i][1])))
                self.bd_pacientes_citas.setItem(i, 1, QTableWidgetItem(str(id[i][1])))
                self.bd_pacientes_citas.setItem(i, 2, QTableWidgetItem(str(id[i][2])))
                self.bd_pacientes_citas.setItem(i, 3, QTableWidgetItem(str(id[i][3])))
                self.bd_pacientes_citas.setItem(i, 4, QTableWidgetItem(str(id[i][4])))
                self.bd_pacientes_citas.setItem(i, 5, QTableWidgetItem(str(id[i][6])))
                #self.bd_pacientes.setItem(i, 6, QTableWidgetItem(str(dato[i][6])))
                self.bd_pacientes_citas.setItem(i, 6, QTableWidgetItem(str(id[i][8])))

                # Crear un checkbox para cada fila
                checkbox = QCheckBox()

                # Restaurar el estado del checkbox desde la estructura
                checkbox.setChecked(checkbox_states.get(i, False))

                # Centrar el checkbox en la celda
                checkbox_layout = QHBoxLayout()
                checkbox_layout.setContentsMargins(0, 0, 0, 0)
                checkbox_layout.setAlignment(Qt.AlignCenter)
                checkbox_widget = QWidget()
                checkbox_layout.addWidget(checkbox)
                checkbox_widget.setLayout(checkbox_layout)

                # Agregar el checkbox en la última columna
                self.bd_pacientes_citas.setCellWidget(i, self.bd_pacientes_citas.columnCount() - 1, checkbox_widget)
        except Exception as e:
            print(e)
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def cargarTablaPacientes_cumple(self):
        try:
            import datetime
            manager = sql_structures.Manager()
            fecha_actual = datetime.date.today()
            id = manager.print_table_cumple_mes(fecha_actual)
            self.bd_pacientes_cumple.setRowCount(len(id))
            for i in range(len(id)):
                self.bd_pacientes_cumple.setItem(i, 0, QTableWidgetItem(str(id[i][0])))
                self.bd_pacientes_cumple.setItem(i, 1, QTableWidgetItem(str(id[i][1])))
                self.bd_pacientes_cumple.setItem(i, 2, QTableWidgetItem(str(id[i][2])))
                self.bd_pacientes_cumple.setItem(i, 3, QTableWidgetItem(str(id[i][3])))
                self.bd_pacientes_cumple.setItem(i, 4, QTableWidgetItem(str(id[i][4])))
                self.bd_pacientes_cumple.setItem(i, 5, QTableWidgetItem(str(id[i][6])))

        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def click_tabla_pacientes(self, row, column):
        try:
            manager = sql_structures.Manager()
            item = self.bd_pacientes.item(row, column)
            value = item.text()
            columns_ingreso = ['id', 'nombre', 'apellido', 'telefono', 'dpi', 'cita', 'cumpleaños']
            header_item = self.bd_pacientes.horizontalHeaderItem(column)
            column_name = header_item.text()
            if column_name == 'ID':
                self.id_c = manager.dar_id_normal('paciente', value, 'id')
                self.cabeza = header_item.text()
            elif column_name == 'Apellido':
                self.id_c = manager.dar_id_normal('paciente', value, 'apellido')
                self.cabeza = header_item.text()
            elif column_name == 'Teléfono':
                self.id_c = manager.dar_id_normal('paciente', value, 'telefono')
                self.cabeza = 'telefono'
            elif column_name == 'DPI':
                self.id_c = manager.dar_id_normal('paciente', value, 'dpi')
                self.cabeza = 'dpi'
            elif column_name == 'Fecha De Primera Cita':
                self.id_c = manager.dar_id_normal('paciente', value, 'fecha')
                self.cabeza = 'fecha'
            elif column_name == 'Cita':
                self.id_c = manager.dar_id_normal('paciente', value, 'cita')
                self.cabeza = 'cita'
            elif column_name == 'Cumpleaños':
                self.id_c = manager.dar_id_normal('paciente', value, 'cumpleaños')
                self.cabeza = 'cumpleaños'
            elif column_name == 'Cumpleaños':
                self.id_c = manager.dar_id_normal('paciente', value, 'cumpleaños')
                self.cabeza = 'cumpleaños'

            else:
                self.id_c = manager.dar_id_normal('paciente', value, 'nombre')
                self.cabeza = header_item.text()
        except Exception as e:
            pass

    def click_tabla_pacientes_citas(self, row, column):
        try:
            manager = sql_structures.Manager()
            item = self.bd_pacientes_citas.item(row, column)
            value = item.text()
            columns_ingreso = ['id', 'nombre', 'apellido', 'telefono', 'dpi', 'cita', 'cumpleaños']
            header_item = self.bd_pacientes_citas.horizontalHeaderItem(column)
            column_name = header_item.text()

            if column_name == 'Observaciones':
                # self.id_c = manager.dar_id_normal('paciente', value, 'observaciones')
                self.cabeza = 'observaciones'
                self.id_nombre = self.bd_pacientes_citas.item(row, 0).text()
                self.id_c = manager.get_id_name_pa(self.id_nombre)
        except Exception as e:
            print(e)

    def busqueda_paciente(self, letras):
        try:
            self.limpiar_tabla(self.bd_pacientes)
            manager = sql_structures.Manager()
            data = manager.busqueda("paciente", letras)
            self.bd_pacientes.setRowCount(len(data))
            for fila, listaItem in enumerate(data):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_pacientes.setCellWidget(fila, self.bd_pacientes.columnCount() - 1, widget)
                self.boton_uno_coneccion_far(btn_uno, fila, self.bd_pacientes)
                self.boton_dos_connecion_far(btn_dos, fila, self.bd_pacientes)
                for columna, item in enumerate(listaItem):
                    self.bd_pacientes.setItem(fila, columna, QTableWidgetItem(str(item)))  # insertar items
                    self.bd_pacientes.item(fila, columna).setTextAlignment(
                        Qt.AlignCenter | Qt.AlignVCenter)  # alinear items
        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    # _______________________________________________________________

    def cargarTablaCombo(self):
        try:
            manager = sql_structures.Manager()
            dato = manager.print_table('promociones')
            self.bd_combos.setRowCount(len(dato))
            for fila, listaItem in enumerate(dato):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_combos.setCellWidget(fila, self.bd_combos.columnCount() - 1, widget)
                self.boton_uno_coneccion_combo(btn_uno, fila, self.bd_combos)
                self.boton_dos_connecion_combo(btn_dos, fila, self.bd_combos)
                for columna, item in enumerate(listaItem):
                    self.bd_combos.setItem(fila, columna, QTableWidgetItem(str(item)))  # insertar items
                    self.bd_combos.item(fila, columna).setTextAlignment(
                        Qt.AlignCenter | Qt.AlignVCenter)  # alinear items
        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def click_tabla_combo(self, row, column):
        try:
            manager = sql_structures.Manager()
            item = self.bd_combos.item(row, column)
            value = item.text()
            columns_ingreso = ['id', 'nombre', 'tarjeta', 'efectivo']
            header_item = self.bd_combos.horizontalHeaderItem(column)
            column_name = header_item.text()
            if column_name == 'ID':
                self.id_c = manager.dar_id_normal('promociones', value, 'id')
                self.cabeza = header_item.text()
            elif column_name == 'Terapia':
                self.id_c = manager.dar_id_normal('promociones', value, 'nombre')
                self.cabeza = 'nombre'
            elif column_name == 'Tarjeta':
                self.id_c = manager.dar_id_normal('promociones', value, 'tarjeta')
                self.cabeza = header_item.text()
            elif column_name == 'Efectivo':
                self.id_c = manager.dar_id_normal('promociones', value, 'efectivo')
                self.cabeza = header_item.text()
        except Exception as e:
            pass

    def agregar_combo_a_carrito(self):
        try:
            dialogo = DialogoCombo(self)

            if dialogo.exec_() == QDialog.Accepted:
                requiere_medicamentos = dialogo.chk_requiere_medicamentos.isChecked()
                medicamentos = dialogo.medicamentos_seleccionados

                self.procesar_registro_combo(self.id_c, requiere_medicamentos, medicamentos)

        except Exception as e:
            print(f"Error al registrar combo: {e}")
            QMessageBox.critical(self, "Error", "Error al registrar el combo")

    def procesar_registro_combo(self, jornada_id, requiere_medicamentos, medicamentos):
        try:
            manager = sql_structures.Manager()
            item = manager.get_carrito_jo("promociones", "id", jornada_id)

            self.carrito.agregar_a_carrito(item[0][0], item[0][2], 1, item[0][1], -1, jornada_id,
                                               -1, -1, -1, -1)

            # Si el combo requiere medicamentos, procesarlos
            if requiere_medicamentos and medicamentos:
                for med in medicamentos:
                    self.agregar_medicamento_regalo_a_carrito(med['cantidad'], med['id'])


            QMessageBox.information(self, "Éxito", "Combo registrado correctamente")
            self.cargarTablacarrito()

        except Exception as e:
            print(f"Error al procesar registro de combo: {e}")
            QMessageBox.critical(self, "Error", "Error al procesar el registro del combo")

    def busqueda_combo(self, letras):
        try:
            self.limpiar_tabla(self.bd_combos)
            manager = sql_structures.Manager()
            data = manager.busqueda("promociones", letras)
            self.bd_combos.setRowCount(len(data))
            for fila, listaItem in enumerate(data):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_combos.setCellWidget(fila, self.bd_combos.columnCount() - 1, widget)
                self.boton_uno_coneccion_far(btn_uno, fila, self.bd_combos)
                self.boton_dos_connecion_far(btn_dos, fila, self.bd_combos)
                for columna, item in enumerate(listaItem):
                    self.bd_combos.setItem(fila, columna, QTableWidgetItem(str(item)))  # insertar items
                    self.bd_combos.item(fila, columna).setTextAlignment(
                        Qt.AlignCenter | Qt.AlignVCenter)  # alinear items
        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    # _______________________________________________________________

    def cargarTablaUsuario(self):
        try:
            manager = sql_structures.Manager()
            dato = manager.print_table('usuario')
            self.bd_usuario.setRowCount(len(dato))
            for fila, listaItem in enumerate(dato):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_usuario.setCellWidget(fila, self.bd_usuario.columnCount() - 1, widget)
                self.boton_uno_coneccion_usu(btn_uno, fila, self.bd_usuario)
                self.boton_dos_connecion_usu(btn_dos, fila, self.bd_usuario)
                for i in range(len(dato)):
                    self.bd_usuario.setItem(i, 0, QTableWidgetItem(str(dato[i][0])))
                    self.bd_usuario.setItem(i, 1, QTableWidgetItem(str(dato[i][1])))
                    self.bd_usuario.setItem(i, 2, QTableWidgetItem(str(dato[i][2])))
                    self.bd_usuario.setItem(i, 3, QTableWidgetItem(str(dato[i][3])))

        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def click_tabla_usuario(self, row, column):
        manager = sql_structures.Manager()
        item = self.bd_usuario.item(row, column)
        value = item.text()
        columns_ingreso = ['id', 'Usuario', 'Contraseña', 'Rol', 'permisos_id']
        header_item = self.bd_usuario.horizontalHeaderItem(column)
        column_name = header_item.text()

        if column_name == 'Usuario':
            self.id_c = manager.get('usuario', columns_ingreso, value, 'Usuario')
            self.cabeza = header_item.text()
        elif column_name == 'Contraseña':
            self.id_c = manager.get('usuario', columns_ingreso, value, 'Contraseña')
            self.cabeza = header_item.text()
        elif column_name == 'Rol':
            self.id_c = manager.get('usuario', columns_ingreso, value, 'Rol')
            self.cabeza = header_item.text()

    def busqueda_usuario(self, letras):
        try:
            self.limpiar_tabla(self.bd_usuario)
            manager = sql_structures.Manager()
            data = manager.busqueda_usu("usuario", letras)
            self.bd_usuario.setRowCount(len(data))
            for fila, listaItem in enumerate(data):
                btn_uno = self.creabotones(
                    estilo=(u"""QPushButton{background-color: white; color: white} QPushButton:hover{background-color: #88cc14} 
                                                    QPushButton:pressed{background-color: #88cc14}"""),
                    icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\editar.png')
                btn_dos = self.creabotones(estilo=(u"""QPushButton{background-color: white; color: white; border-radius: 25px} QPushButton:hover{background-color:  #FE2C55}
                                                                                     QPushButton:pressed{background-color:  #FE2C55}"""),
                                           icono='C:\\Users\\andre\\OneDrive\\Escritorio\\Sistema-Hidrocolon-main\\views\\InterfaceImages\\eliminarr.png')
                layout = QHBoxLayout()
                layout.setContentsMargins(0, 0, 0, 0)  # espaciado
                layout.setSpacing(0)  # espaciado
                layout.addWidget(btn_uno)  # agregar botones al layout
                layout.addWidget(btn_dos)
                widget = QWidget()
                widget.setLayout(layout)
                self.bd_usuario.setCellWidget(fila, self.bd_usuario.columnCount() - 1, widget)
                self.boton_uno_coneccion_far(btn_uno, fila, self.bd_usuario)
                self.boton_dos_connecion_far(btn_dos, fila, self.bd_usuario)
                for columna, item in enumerate(listaItem):
                    self.bd_usuario.setItem(fila, columna, QTableWidgetItem(str(item)))  # insertar items
                    self.bd_usuario.item(fila, columna).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)  # alinear items
        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    # ______________________________________________________________

    def cargarTablaCierre(self):
        try:
            manager = sql_structures.Manager()
            dato = manager.print_table_cierre()
            self.bd_cierre.setRowCount(len(dato))
            for i in range(len(dato)):
                self.bd_cierre.setItem(i, 0, QTableWidgetItem(str(dato[i][0])))
                self.bd_cierre.setItem(i, 1, QTableWidgetItem(str(dato[i][1])))
                self.bd_cierre.setItem(i, 2, QTableWidgetItem(str(dato[i][2])))
                self.bd_cierre.setItem(i, 3, QTableWidgetItem(str(dato[i][3])))
                self.bd_cierre.setItem(i, 4, QTableWidgetItem(str(dato[i][4])))
                self.bd_cierre.setItem(i, 5, QTableWidgetItem(str(dato[i][5])))
                self.bd_cierre.setItem(i, 6, QTableWidgetItem(str(dato[i][6])))

        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def click_tabla_cierre(self, row, column):
        manager = sql_structures.Manager()
        item = self.bd_cierre.item(row, column)
        value = item.text()
        columns_ingreso = ['id', 'nombre', "cantidad", "efectivo", "tarjeta", "monto", "fecha", "usuario", "carrito_id"]
        header_item = self.bd_cierre.horizontalHeaderItem(column)
        column_name = header_item.text()

        if column_name == 'Detalle':
            VentanaFuncional._contra = manager.get_carrito_devu('cierre', columns_ingreso, value, 'nombre')
        elif column_name == 'Cantidad':
            VentanaFuncional._contra = manager.get_carrito_devu('cierre', columns_ingreso, value, 'cantidad')
        elif column_name == 'Efectivo':
            VentanaFuncional._contra = manager.get_carrito_devu('cierre', columns_ingreso, value, 'efectivo')
        elif column_name == 'Tarjeta':
            VentanaFuncional._contra = manager.get_carrito_devu('cierre', columns_ingreso, value, 'tarjeta')
        elif column_name == 'Monto Total':
            VentanaFuncional._contra = manager.get_carrito_devu('cierre', columns_ingreso, value, 'monto')
        elif column_name == 'Fecha de Venta':
            VentanaFuncional._contra = manager.get_carrito_devu('cierre', columns_ingreso, value, 'fecha')
        elif column_name == 'Usuario Responsable':
            VentanaFuncional._contra = manager.get_carrito_devu('cierre', columns_ingreso, value, 'usuario')
        elif column_name == 'Codigo_de_venta':
            VentanaFuncional._contra = manager.get_carrito_devu('cierre', columns_ingreso, value, 'carrito_id')

    def cargarTablavitacora(self):
        try:
            manager = sql_structures.Manager()
            dato = manager.print_table('vitacora')
            self.bd_vitacora.setRowCount(len(dato))
            for i in range(len(dato)):
                self.bd_vitacora.setItem(i, 0, QTableWidgetItem(str(dato[i][1])))
                self.bd_vitacora.setItem(i, 1, QTableWidgetItem(str(dato[i][2])))
                self.bd_vitacora.setItem(i, 2, QTableWidgetItem(str(dato[i][3])))
                self.bd_vitacora.setItem(i, 3, QTableWidgetItem(str(dato[i][4])))
                self.bd_vitacora.setItem(i, 4, QTableWidgetItem(str(dato[i][5])))
        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    # ___________________________________________________________

    def creabotones(self, estilo=None, titulo=None, icono=None, tooltip=None):
        boton_nuevo = QPushButton()
        boton_nuevo.setStyleSheet(estilo)
        boton_nuevo.setIcon(QIcon(icono))
        boton_nuevo.setToolTip(tooltip)
        boton_nuevo.setText(titulo)
        sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Ignored)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(boton_nuevo.sizePolicy().hasHeightForWidth())
        boton_nuevo.setSizePolicy(sizePolicy)
        return boton_nuevo
# ___________________________________________________________

    def boton_uno_coneccion_tera(self, b_uno, fila, tabla):
        b_uno.clicked.connect(lambda: tabla.selectRow(fila))
        b_uno.clicked.connect(lambda: self.boton_uno_accion_tera(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_dos_connecion_tera(self, b_dos, fila, tabla):
        b_dos.clicked.connect(lambda: tabla.selectRow(fila))  # selecciona la fina
        b_dos.clicked.connect(lambda: self.boton_dos_accion_tera(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_uno_accion_tera(self, ide):
        self.terapias.actualizarTerapia(ide, self.new_value, self.cabeza)

    def boton_dos_accion_tera(self, ide):
        self.terapias.eliminar_terapias(ide)
        self.cargarTablaTerapias()

    def on_cell_changed_tera(self, row, column):
        self.new_value = self.bd_terapias.item(row, column).text()

    # ___________________________________________________________

    def boton_uno_coneccion_far(self, b_uno, fila, tabla):
        b_uno.clicked.connect(lambda: tabla.selectRow(fila))
        b_uno.clicked.connect(lambda: self.boton_uno_accion_far(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_dos_connecion_far(self, b_dos, fila, tabla):
        b_dos.clicked.connect(lambda: tabla.selectRow(fila))  # selecciona la fina
        b_dos.clicked.connect(lambda: self.boton_dos_accion_far(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_tres_connecion_far(self, b_dos, fila, tabla):
        b_dos.clicked.connect(lambda: tabla.selectRow(fila))  # selecciona la fina
        b_dos.clicked.connect(lambda: self.boton_tres_accion_far(tabla.item(fila, 0).text()))

    def boton_uno_accion_far(self, ide):
        """Abre la ventana de agregar medicamento en modo edición"""
        try:
            # Importar la ventana de agregar medicamento
            from .ModificarMedicamentos import AgregarMedi

            # Crear instancia de la ventana
            self.ventana_editar_medicamento = AgregarMedi()

            # Cargar los datos del medicamento
            self.ventana_editar_medicamento.cargar_medicamento_para_editar(ide)

            # Conectar señal para recargar la tabla cuando se cierre
            self.ventana_editar_medicamento.switch_window.connect(
                lambda: self.cargarTablaFarmacia()
            )

            # Mostrar la ventana
            self.ventana_editar_medicamento.show()

        except Exception as e:
            print(f"Error al abrir ventana de edición: {e}")
            QMessageBox.critical(self, "Error", f"No se pudo abrir la ventana de edición: {str(e)}")

    def boton_dos_accion_far(self, ide):
        self.medicamento.eliminarMedicamento(ide)
        self.cargarTablaFarmacia()

    def boton_tres_accion_far(self, id):
        try:
            VentanaFuncional._id_detalle = id
            print(1)
            from .detalles import Detalles
            self.Detalles = Detalles()
            self.Detalles.show()
        except Exception as e:
            print(e)


    def on_cell_changed_far(self, row, column):
        self.new_value = self.bd_farmacia.item(row, column).text()

# ___________________________________________________________

    def boton_uno_coneccion_paci(self, b_uno, fila, tabla):
        b_uno.clicked.connect(lambda: tabla.selectRow(fila))
        b_uno.clicked.connect(lambda: self.boton_uno_accion_paci(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_dos_connecion_paci(self, b_dos, fila, tabla):
        b_dos.clicked.connect(lambda: tabla.selectRow(fila))  # selecciona la fina
        b_dos.clicked.connect(lambda: self.boton_dos_accion_paci(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_dos_connecion_paci_cita(self,  fila, tabla):
        tabla.selectRow(fila)  # selecciona la fina
        self.boton_uno_accion_paci_cita(tabla.item(fila, 0).text())  # obtiene el id

    def boton_uno_accion_paci(self, ide):
        self.paciente.actualizarPaciente(ide, self.new_value, self.cabeza)

    def boton_dos_accion_paci(self, ide):
        self.paciente.eliminar_paciente(ide)
        self.cargarTablaPacientes()

    def on_cell_changed_paci(self, row, column):
        self.new_value = self.bd_pacientes.item(row, column).text()

    def on_item_changed(self, item):
        self.new_value = item.text()
        row = item.row()
        column = item.column()
        self.on_cell_changed_citas(row, column)

    def on_cell_changed_citas(self, row, column):
        try:
            #self.new_value = self.bd_pacientes_citas.item(row, column).text()
            self.current_row = row  # Guardamos la fila actual como variable de clase
            self.bd_pacientes_citas.keyPressEvent = self.handle_key_press
            print(self.new_value)
        except Exception as e:
            print(e)

    def handle_key_press(self, event):
        try:
            manager = sql_structures.Manager()
            if event.key() in (Qt.Key_Return, Qt.Key_Enter):
                fila_actual = self.bd_pacientes_citas.currentRow()
                if fila_actual >= 0:
                    self.id_nombre = self.bd_pacientes_citas.item(self.current_row, 0).text()
                    self.id_c = manager.get_id_name_pa(self.id_nombre)
                    self.boton_uno_accion_paci_cita(self.id_c)
            # Importante: mantener el comportamiento original de la tabla
            QTableWidget.keyPressEvent(self.bd_pacientes_citas, event)
        except Exception as e:
            print(e)

    def boton_uno_accion_paci_cita(self, ide):
        print(ide, self.new_value, self.cabeza)
        self.paciente.actualizarPaciente(ide, self.new_value, self.cabeza)

# ___________________________________________________________

    def boton_uno_coneccion_jorda(self, b_uno, fila, tabla):
        b_uno.clicked.connect(lambda: tabla.selectRow(fila))
        b_uno.clicked.connect(lambda: self.boton_uno_accion_jorda(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_dos_connecion_jorda(self, b_dos, fila, tabla):
        b_dos.clicked.connect(lambda: tabla.selectRow(fila))  # selecciona la fina
        b_dos.clicked.connect(lambda: self.boton_dos_accion_jorda(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_uno_accion_jorda(self, ide):
        self.jornada.actualizarjornadas(ide, self.new_value, self.cabeza)

    def boton_dos_accion_jorda(self, ide):
        self.jornada.eliminar_jornadas(ide)
        self.cargarTablaJornadas()

    def on_cell_changed_jorda(self, row, column):
        self.new_value = self.bd_jornadas.item(row, column).text()
# ___________________________________________________________

    def boton_dos_connecion_carrito(self, b_dos, fila, tabla):
        b_dos.clicked.connect(lambda: tabla.selectRow(fila))  # selecciona la fina
        b_dos.clicked.connect(lambda: self.boton_dos_accion_carrito(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_dos_accion_carrito(self, ide):
        try:
            manager = sql_structures.Manager()
            id = manager.get_id_2("carrito", ide)
            print(f"ide {id}")
            cantidad = manager.get_cantidad_carrito(id)
            print(1)
            nombre = manager.get_name_carrito(id)
            print(2)
            id_medi = manager.get_idddd(nombre[0][0])
            print(3)
            id_medicamento = manager.obtener_id_carrito(id)
            print(4)

            print(5)
            if id_medi:
                pre = manager.get_presentacion("medicamentos", "id", id_medicamento)
                # item = manager.get_carrito_medic("medicamentos", "id", id_medi[0][0], "nombre", nombre)
                item = manager.get_carrito_medic("medicamentos", "id", id_medicamento, 'presentacion', pre)
                print(item)
                if item:
                    existencias = item[0][1]  # item[0][1] es existencias
                    nuevas_existencias = existencias + cantidad[0][0]
                    self.medicamento.actualizarMedicamentor(id_medicamento, nuevas_existencias, "existencias")
                else:
                    pass
            else:
                pass
            self.carrito.eliminar_a_carritoo(id)
            self.cargarTablacarrito()

            # Actualizar totales después de eliminar
            self.actualizar_totales_carrito()

        except Exception as E:
            print(E)

    #def boton_tres_connecion_carrito(self, b_dos, fila, tabla):
    #    b_dos.clicked.connect(lambda: tabla.selectRow(fila))  # selecciona la fila
    #    b_dos.clicked.connect(lambda: self.boton_tres_accion_carrito(tabla.item(fila, 0).text(), fila))

    def get_boton_from_layout(self, fila, index):
        cell_widget = self.bd_carrito.cellWidget(fila, self.bd_carrito.columnCount() - 1)
        if cell_widget:
            layout = cell_widget.layout()
            if layout and layout.count() > index:
                return layout.itemAt(index).widget()
        return None

    def boton_tres_connecion_carrito(self, b_tres, fila, tabla):
        def on_efectivo_clicked():
            try:
                tabla.selectRow(fila)
                id_carrito = tabla.item(fila, 0).text()
                self.boton_tres_accion_carrito(id_carrito, fila)
                b_tres.setStyleSheet("background-color: #4CAF50; color: white;")
                b_tarjeta = self.get_boton_from_layout(fila, 2)
                if b_tarjeta:
                    b_tarjeta.setStyleSheet("background-color: white;")
            except Exception as e:
                print(f"[ERROR on_efectivo_clicked] {e}")

        b_tres.clicked.connect(on_efectivo_clicked)

    def boton_tres_accion_carrito(self, ide, fila):
        try:
            print(f"[EFECTIVO] Aplicando precio efectivo al ID {ide} en fila {fila}")
            manager = sql_structures.Manager()
            idee = manager.dar_id_normal('carrito', ide, 'nombre')
            manager.aplicar_precio_efectivo(idee)

            # Recargar con modo individual para respetar precios de cada item
            self.cargarTablacarrito(modo='individual')

            # Actualizar totales después del cambio
            self.actualizar_totales_carrito()

        except Exception as e:
            print(f"[ERROR boton_tres_accion_carrito] {e}")

    def boton_cuatro_connecion_carrito(self, b_cuatro, fila, tabla):
        def on_tarjeta_clicked():
            try:
                tabla.selectRow(fila)
                id_carrito = tabla.item(fila, 0).text()
                self.boton_cuatro_accion_carrito(id_carrito, fila)
                b_cuatro.setStyleSheet("background-color: #4CAF50; color: white;")
                b_efectivo = self.get_boton_from_layout(fila, 1)
                if b_efectivo:
                    b_efectivo.setStyleSheet("background-color: white;")
            except Exception as e:
                print(f"[ERROR on_tarjeta_clicked] {e}")

        b_cuatro.clicked.connect(on_tarjeta_clicked)

    def boton_cuatro_accion_carrito(self, ide, fila):
        try:
            print(f"[TARJETA] Aplicando precio tarjeta al ID {ide} en fila {fila}")
            manager = sql_structures.Manager()
            idee = manager.dar_id_normal('carrito', ide, 'nombre')
            manager.aplicar_precio_tarjeta(idee)

            # Recargar con modo individual para respetar precios de cada item
            self.cargarTablacarrito(modo='individual')

            # Actualizar totales después del cambio
            self.actualizar_totales_carrito()

        except Exception as e:
            print(f"[ERROR boton_cuatro_accion_carrito] {e}")

    def actualizar_total_carrito(self):
        """Actualiza el total del carrito"""
        try:
            self.porcentaje = DescuentoMedi.get_porcentaje()
            self.cantidad = DescuentoMedi.get_cantidad()
            total = 0

            for fila in range(self.bd_carrito.rowCount()):
                precio_total = self.bd_carrito.item(fila, 4)
                if precio_total and precio_total.text():
                    total += float(precio_total.text())

            total_descuento_cantidad = total - self.cantidad
            total_descuento_porcentaje_pibote = total*self.porcentaje
            total_descuento_porcentaje = total - total_descuento_porcentaje_pibote

            # Actualizar las etiquetas de totales
            self.label_28.setText(f"{total:.2f}")  # Ajusta estos labels según los que uses para mostrar totales
            self.label_30.setText(f"{total_descuento_cantidad:.2f}")
            self.label_30.setText(f"{total_descuento_porcentaje:.2f}")

        except Exception as e:
            print(f"Error al actualizar total: {e}")

    def on_cell_changed_carrito(self, row, column):
        self.new_value = self.bd_carrito.item(row, column).text()

        # ___________________________________________________________

    def boton_uno_coneccion_combo(self, b_uno, fila, tabla):
        b_uno.clicked.connect(lambda: tabla.selectRow(fila))
        b_uno.clicked.connect(lambda: self.boton_uno_accion_combo(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_dos_connecion_combo(self, b_dos, fila, tabla):
        b_dos.clicked.connect(lambda: tabla.selectRow(fila))  # selecciona la fina
        b_dos.clicked.connect(lambda: self.boton_dos_accion_combo(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_uno_accion_combo(self, ide):
        self.combo.actualizarpromociones(ide, self.new_value, self.cabeza)

    def boton_dos_accion_combo(self, ide):
        self.combo.eliminarpromociones(ide)
        self.cargarTablaCombo()

    def on_cell_changed_combo(self, row, column):
        self.new_value = self.bd_combos.item(row, column).text()

    # ___________________________________________________________

    def boton_uno_coneccion_usu(self, b_uno, fila, tabla):
        b_uno.clicked.connect(lambda: tabla.selectRow(fila))
        b_uno.clicked.connect(lambda: self.boton_uno_accion_usu(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_dos_connecion_usu(self, b_dos, fila, tabla):
        b_dos.clicked.connect(lambda: tabla.selectRow(fila))  # selecciona la fina
        b_dos.clicked.connect(lambda: self.boton_dos_accion_usu(tabla.item(fila, 0).text()))  # obtiene el id

    def boton_uno_accion_usu(self, ide):
        self.usuario_in.update_user(ide, self.new_value, self.cabeza)

    def boton_dos_accion_usu(self, ide):
        try:
            self.usuario_in.delete_user(ide)
            self.cargarTablaUsuario()
        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de carga de tabla')

    def on_cell_changed_usu(self, row, column):
        self.new_value = self.bd_usuario.item(row, column).text()

    # ___________________________________________________________

    def femaleselected(self, selected):
        try:
            # Obtener valores iniciales de porcentaje y cantidad
            self.porcentajet = DescuentoMedi.get_porcentaje()
            self.cantidadt = DescuentoMedi.get_cantidad()

            self.sumat = 0.0
            self.resultadot = 0.0
            self.descuentot = 0.0

            # Caso: Descuento basado en porcentaje
            if self.porcentajet > 0:
                if selected:
                    self.cargarTablacarrito(modo='tarjeta')
                    self.btn_pagodiv.show()
                    self.frame_51.hide()

                    # Calcular suma y aplicar descuento
                    manager = sql_structures.Manager()
                    id = manager.get_dinero_tarjeta()
                    for i in id:
                        self.sumat += float(i[0])

                    self.descuentot = self.sumat * self.porcentajet
                    self.resultadot = self.sumat - self.descuentot

                    # Actualizar etiquetas
                    self.label_28.setText(str(self.sumat))
                    self.label_29.setText(str(self.descuentot))
                    self.label_30.setText(str(self.resultadot))

            # Caso: Descuento basado en cantidad fija
            elif self.cantidadt > 0:
                if selected:
                    self.cargarTablacarrito(modo='tarjeta')
                    self.btn_pagodiv.show()
                    self.frame_51.hide()

                    # Calcular suma y aplicar descuento
                    manager = sql_structures.Manager()
                    id = manager.get_dinero_tarjeta()
                    for i in id:
                        self.sumat += float(i[0])

                    self.descuentot = self.cantidadt
                    self.resultadot = self.sumat - self.descuentot

                    # Actualizar etiquetas
                    self.label_28.setText(str(self.sumat))
                    self.label_29.setText(str(self.descuentot))
                    self.label_30.setText(str(self.resultadot))

            # Caso: Sin descuentos
            else:
                if selected:
                    self.cargarTablacarrito(modo='tarjeta')
                    self.btn_pagodiv.show()
                    self.frame_51.hide()

                    # Calcular solo la suma
                    manager = sql_structures.Manager()
                    id = manager.get_dinero_tarjeta()
                    for i in id:
                        self.sumat += float(i[0])

                    self.resultadot = self.sumat

                    # Actualizar etiquetas
                    self.label_28.setText(str(self.sumat))
                    self.label_29.setText("0.0")
                    self.label_30.setText(str(self.resultadot))

            # Actualizar totales después de todos los cambios
            self.actualizar_totales_carrito()

        except Exception as e:
            QMessageBox.about(self, 'Aviso Pago Tarjeta', 'Ingrese un medicamento/terapia.')

    def maleselected(self, selected):
        try:
            # Obtener valores iniciales de porcentaje y cantidad
            self.porcentaje = DescuentoMedi.get_porcentaje()
            self.cantidad = DescuentoMedi.get_cantidad()

            self.suma = 0.0
            self.resultado = 0.0
            self.descuento = 0.0

            # Caso: Descuento basado en porcentaje
            if self.porcentaje > 0:
                if selected:
                    self.cargarTablacarrito(modo='efectivo')
                    self.btn_pagodiv.hide()
                    self.frame_51.show()

                    # Calcular suma y aplicar descuento
                    manager = sql_structures.Manager()
                    id = manager.get_dinero_efectivo()
                    for i in id:
                        self.suma += float(i[0])

                    self.descuento = self.suma * self.porcentaje
                    self.resultado = self.suma - self.descuento

                    # Actualizar etiquetas
                    self.label_28.setText(str(self.suma))
                    self.label_29.setText(str(self.descuento))
                    self.label_30.setText(str(self.resultado))

            # Caso: Descuento basado en cantidad fija
            elif self.cantidad > 0:
                if selected:
                    self.cargarTablacarrito(modo='efectivo')
                    self.btn_pagodiv.hide()
                    self.frame_51.show()

                    # Calcular suma y aplicar descuento
                    manager = sql_structures.Manager()
                    id = manager.get_dinero_efectivo()
                    for i in id:
                        self.suma += float(i[0])

                    self.descuento = self.cantidad
                    self.resultado = self.suma - self.descuento

                    # Actualizar etiquetas
                    self.label_28.setText(str(self.suma))
                    self.label_29.setText(str(self.descuento))
                    self.label_30.setText(str(self.resultado))

            # Caso: Sin descuentos
            else:
                if selected:
                    self.cargarTablacarrito(modo='efectivo')
                    self.btn_pagodiv.hide()
                    self.frame_51.show()

                    # Calcular solo la suma
                    manager = sql_structures.Manager()
                    id = manager.get_dinero_efectivo()
                    for i in id:
                        self.suma += float(i[0])

                    self.resultado = self.suma

                    # Actualizar etiquetas
                    self.label_28.setText(str(self.suma))
                    self.label_29.setText("0.0")
                    self.label_30.setText(str(self.resultado))

            # Actualizar totales después de todos los cambios
            self.actualizar_totales_carrito()

        except Exception as e:
            QMessageBox.about(self, 'Aviso Pago Efectivo', 'Ingrese un medicamento/terapia.')

    def suma_total_pagos(self):
        VentanaFuncional._dinero_total = str(self.label_30.text())

    def resto(self):
        try:
            dado = self.lineEdit_3.text()
            vuelto = int(dado) - int(self.resultado)
            self.label_33.setText(str(vuelto))
            self.lineEdit_3.clear()
        except Exception as e:
            QMessageBox.about(self, 'Aviso', 'Error de agregado!')

    def bloqueo(self, rol, usuario):
        VentanaFuncional._usuario = usuario
        self.usuario = usuario
        if rol == 1:
            #self.bd_farmacia.hideColumn(0)
            self.bd_terapias.hideColumn(0)
            self.bd_combos.hideColumn(0)
            self.bd_pacientes_cumple.hideColumn(0)
            self.bd_pacientes_citas.hideColumn(0)
            self.bd_pacientes.hideColumn(0)
            self.bd_jornadas.hideColumn(0)
            self.label_13.setText(usuario)
            self.usuario = self.label_13.text()
            self.btn_modulo_usuarios.show()
            self.btn_modulo_cierre.show()

            self.btn_actualizar_medi.show()
            ### los comentados son del boton de agregar paciente
            # self.btn_actualizar_paciente.hide()
            self.btn_actualizar_combo.show()
            self.btn_actualizar_terapia.show()
            self.btn_actualizar_jornada.show()
            self.frame_139.show()
            # self.frame_133.hide()
            self.frame_137.show()
            self.frame_152.show()
            self.frame_154.show()
            self.FrameCierre.show()
            self.FrameUsuarios.show()
            self.frame_30.show()
            self.frame_31.show()

            self.btn_modulo_pacientes.show()
            self.btn_actualizar_paciente.show()
            self.frame_139.show()

            self.frame_29.show()

            self.bd_farmacia.showColumn(7)
            self.bd_pacientes.showColumn(7)
            self.bd_combos.showColumn(4)
            self.bd_terapias.showColumn(4)
            self.bd_jornadas.showColumn(4)
            # self.bd_farmacia.hideColumn(0)
        elif rol == 2:
            self.bd_farmacia.hideColumn(8)
            #self.bd_farmacia.hideColumn(0)
            self.bd_terapias.hideColumn(0)
            self.bd_combos.hideColumn(0)
            self.bd_pacientes_cumple.hideColumn(0)
            self.bd_pacientes_citas.hideColumn(0)
            self.bd_pacientes.hideColumn(0)
            self.bd_jornadas.hideColumn(0)
            self.btn_modulo_usuarios.setHidden(True)
            self.btn_modulo_cierre.hide()
            self.btn_modulo_pacientes.hide()
            self.label_13.setText(usuario)
            self.usuario = self.label_13.text()
            self.btn_actualizar_medi.hide()
            self.bd_farmacia.hideColumn(7)
            self.bd_pacientes.hideColumn(7)
            self.btn_actualizar_paciente.hide()
            self.bd_combos.hideColumn(4)
            self.btn_actualizar_combo.hide()
            self.bd_terapias.hideColumn(4)
            self.btn_actualizar_terapia.hide()
            self.bd_jornadas.hideColumn(4)
            self.btn_actualizar_jornada.hide()
            self.frame_139.hide()
            self.frame_137.hide()
            self.frame_152.hide()
            self.frame_154.hide()
            self.FrameCierre.hide()
            self.FrameUsuarios.hide()
            self.frame_30.hide()
            self.frame_31.hide()
            self.frame_29.hide()
            self.btn_in_cheque.hide()
            self.btn_Repor.hide()
            # self.bd_farmacia.hideColumn(0)
        elif rol == 3:
            self.bd_farmacia.hideColumn(8)
            #self.bd_farmacia.hideColumn(0)
            self.bd_terapias.hideColumn(0)
            self.bd_combos.hideColumn(0)
            self.bd_pacientes_cumple.hideColumn(0)
            self.bd_pacientes_citas.hideColumn(0)
            self.bd_pacientes.hideColumn(0)
            self.bd_jornadas.hideColumn(0)
            self.btn_modulo_usuarios.setHidden(True)
            self.btn_modulo_cierre.hide()
            self.label_13.setText(usuario)
            self.usuario = self.label_13.text()
            self.btn_actualizar_medi.hide()
            #self.bd_farmacia.hideColumn(7)
            self.bd_pacientes.hideColumn(8)
            self.bd_combos.hideColumn(4)
            self.btn_actualizar_combo.hide()
            self.bd_terapias.hideColumn(4)
            self.btn_actualizar_terapia.hide()
            self.bd_jornadas.hideColumn(4)
            self.btn_actualizar_jornada.hide()
            self.frame_139.hide()
            #self.frame_133.hide()
            self.frame_137.hide()
            self.frame_152.hide()
            #self.frame_154.hide()
            self.FrameCierre.hide()
            self.FrameUsuarios.hide()
            self.frame_30.hide()
            self.frame_31.hide()
            self.btn_in_cheque.hide()
            self.btn_Repor.hide()
            # self.bd_farmacia.hideColumn(0)

    def limpiar_tabla(self, tabla):
        tabla.setRowCount(0)

    def borrar_tabla(self):
        try:
            self.label_33.setText('')
            from .descuentosMedi import DescuentoMedi
            manager = sql_structures.Manager()

            print("Iniciando borrado de tabla...")

            # Verificar si hay elementos en el carrito antes de procesarlos
            try:
                articulos_venta = manager.print_table_name_carrito("carrito")
                print(f"Artículos encontrados: {len(articulos_venta) if articulos_venta else 0}")
            except Exception as e:
                print(f"Error al obtener artículos del carrito: {e}")
                articulos_venta = []

            # Solo procesar si hay artículos
            if articulos_venta:
                for i, articulo in enumerate(articulos_venta):
                    try:
                        print(f"Procesando artículo {i + 1}: {articulo}")

                        # Verificar que el artículo tenga el formato esperado
                        if not articulo or len(articulo) == 0:
                            print(f"Artículo vacío, saltando...")
                            continue

                        nombre_articulo = articulo[0]

                        # Obtener ID del artículo
                        try:
                            id_e = manager.get_id_name(nombre_articulo)
                            if not id_e or len(id_e) == 0:
                                print(f"No se encontró ID para {nombre_articulo}")
                                continue
                            id_elemento = id_e[0][0]
                            print(f"ID del elemento: {id_elemento}")
                        except Exception as e:
                            print(f"Error al obtener ID del artículo {nombre_articulo}: {e}")
                            continue

                        # Obtener cantidad
                        try:
                            cantidad = manager.get_cantidad_carrito(id_elemento)
                            if not cantidad or len(cantidad) == 0:
                                print(f"No se encontró cantidad para {nombre_articulo}")
                                continue
                            cantidad_valor = cantidad[0][0]
                            print(f"Cantidad: {cantidad_valor}")
                        except Exception as e:
                            print(f"Error al obtener cantidad: {e}")
                            continue

                        # Obtener nombre del carrito
                        try:
                            nombre = manager.get_name_carrito(id_elemento)
                            if not nombre or len(nombre) == 0:
                                print(f"No se encontró nombre en carrito para ID {id_elemento}")
                                continue
                            nombre_carrito = nombre[0][0]
                            print(f"Nombre del carrito: {nombre_carrito}")
                        except Exception as e:
                            print(f"Error al obtener nombre del carrito: {e}")
                            continue

                        # Verificar si es un medicamento
                        try:
                            id_medi = manager.get_idddd(nombre_carrito)
                            print(f"ID medicamento: {id_medi}")

                            if id_medi and len(id_medi) > 0:
                                # Es un medicamento, restaurar existencias
                                try:
                                    id_medicamento = manager.obtener_id_carrito(id_elemento)
                                    print(f"ID medicamento en carrito: {id_medicamento}")

                                    # Obtener presentación
                                    pre = manager.get_presentacion("medicamentos", "id", id_medicamento)
                                    print(f"Presentación: {pre}")

                                    # Obtener item del medicamento
                                    item = manager.get_carrito_medic("medicamentos", "id", id_medicamento,
                                                                     'presentacion', pre)
                                    print(f"Item medicamento: {item}")

                                    if item and len(item) > 0:
                                        existencias_actuales = item[0][1]
                                        nuevas_existencias = existencias_actuales + cantidad_valor
                                        print(
                                            f"Restaurando existencias: {existencias_actuales} + {cantidad_valor} = {nuevas_existencias}")

                                        self.medicamento.actualizarMedicamentor(id_medicamento, nuevas_existencias,
                                                                                "existencias")
                                        print(f"Existencias restauradas para medicamento ID {id_medicamento}")
                                    else:
                                        print(f"No se encontró item del medicamento")
                                except Exception as e:
                                    print(f"Error al restaurar existencias del medicamento: {e}")
                            else:
                                print(f"No es un medicamento: {nombre_carrito}")

                        except Exception as e:
                            print(f"Error al verificar si es medicamento: {e}")

                    except Exception as e:
                        print(f"Error al procesar artículo {i + 1}: {e}")
                        continue

            # Limpiar la tabla del carrito
            try:
                print("Eliminando registros del carrito...")
                manager.delete_table("carrito")
                print("Tabla de carrito limpiada")
            except Exception as e:
                print(f"Error al limpiar tabla de carrito: {e}")
                raise

            # Resetear descuentos
            try:
                print("Reseteando descuentos...")
                DescuentoMedi.reset_descuentos()
                print("Descuentos reseteados")
            except Exception as e:
                print(f"Error al resetear descuentos: {e}")

            # Recargar la tabla del carrito
            try:
                print("Recargando tabla del carrito...")
                self.cargarTablacarrito()
                print("Tabla recargada")
            except Exception as e:
                print(f"Error al recargar tabla: {e}")

            # Resetear etiquetas
            try:
                print("Reseteando etiquetas...")
                self.label_28.setText("0.00")
                self.label_29.setText("0.00")
                self.label_30.setText("0.00")
                print("Etiquetas reseteadas")
            except Exception as e:
                print(f"Error al resetear etiquetas: {e}")

            # Actualizar totales
            try:
                print("Actualizando totales...")
                self.actualizar_totales_carrito()
                print("Totales actualizados")
            except Exception as e:
                print(f"Error al actualizar totales: {e}")

            print("Borrado de tabla completado exitosamente")

        except Exception as e:
            print(f"Error crítico en borrar_tabla: {e}")
            import traceback
            traceback.print_exc()

            # En caso de error crítico, al menos intentar limpiar la tabla
            try:
                manager = sql_structures.Manager()
                manager.delete_table("carrito")
                self.cargarTablacarrito()
                self.label_28.setText("0.00")
                self.label_29.setText("0.00")
                self.label_30.setText("0.00")
                QMessageBox.warning(self, "Advertencia",
                                    "Se limpió la tabla pero hubo errores al restaurar existencias")
            except Exception as e2:
                print(f"Error al hacer limpieza de emergencia: {e2}")
                QMessageBox.critical(self, "Error", f"Error crítico al limpiar tabla: {str(e)}")

    def borrar_tabla_normal(self):
        from .descuentosMedi import DescuentoMedi
        manager = sql_structures.Manager()
        manager.delete_table("carrito")
        self.cargarTablacarrito()

        # Resetear todas las etiquetas y descuentos
        self.label_28.setText("0.00")
        self.label_29.setText("0.00")
        self.label_30.setText("0.00")
        DescuentoMedi.reset_descuentos()

        # Actualizar totales para asegurar que todo esté en 0
        self.actualizar_totales_carrito()

    def ingresar_cierre(self):
        try:
            print("🔍 INICIO: ingresar_cierre")
            import datetime
            manager = sql_structures.Manager()

            # DEBUGGING: Verificar carrito
            id = manager.print_table("carrito")
            print(f"🔍 Items en carrito: {len(id) if id else 0}")
            print(f"🔍 Datos del carrito: {id}")

            if not id or len(id) == 0:
                print("⚠️ Carrito vacío")
                QMessageBox.warning(self, "Carrito Vacío", "No hay productos en el carrito para procesar.")
                return

            # DEBUGGING: Verificar paciente
            print(f"🔍 Paciente actual: {self.paciente_actual}")

            # Capturar items del carrito para el historial
            items_para_historial = []
            for item in id:
                print(f"🔍 Procesando item: {item}")
                medicamento_id = self.obtener_medicamento_id_por_nombre(item[1]) if item[1] else None
                print(f"🔍 Medicamento ID obtenido: {medicamento_id}")

                items_para_historial.append({
                    'medicamento_id': medicamento_id,
                    'nombre': item[1],
                    'cantidad': item[2],
                    'precio_unitario': float(item[4]) / item[2] if item[2] > 0 else 0,
                    'precio_total': float(item[4])
                })

            print(f"🔍 Items para historial: {items_para_historial}")

            # TU LÓGICA ORIGINAL DE VENTA
            print("🔍 Iniciando procesamiento de venta...")
            tiempo = datetime.date.today()
            id_ca = manager.get_ultimo_carrito()
            print(f"🔍 ID carrito obtenido: {id_ca}")
            self.usuario = self.label_13.text()
            print(f"🔍 Usuario: {self.usuario}")
            id_ca = int(id_ca[0]) + 1
            print(f"🔍 Nuevo ID carrito: {id_ca}")

            metodo_pago = self.obtener_metodo_pago_seleccionado()
            print(f"🔍 Método de pago: {metodo_pago}")
            # GUARDAR método de pago para el comprobante
            VentanaFuncional._metodo_pago_guardado = metodo_pago
            print(f"🔍 Método de pago guardado EN CLASE: {VentanaFuncional._metodo_pago_guardado}")

            # Procesar cada item
            for i in id:
                print(f"🔍 Procesando venta del item: {i}")
                try:
                    if metodo_pago == "efectivo":
                        monto = float(i[4])
                        print(f"🔍 Registrando venta efectivo: {monto}")
                        self.cierre.agregarcierre(i[1], i[2], 0, i[4], monto, tiempo, self.usuario, id_ca)
                    elif metodo_pago == "tarjeta":
                        monto = float(i[3])
                        print(f"🔍 Registrando venta tarjeta: {monto}")
                        # Actualizar precio para tarjeta
                        for hist_item in items_para_historial:
                            if hist_item['nombre'] == i[1]:
                                hist_item['precio_total'] = float(i[3])
                                hist_item['precio_unitario'] = float(i[3]) / i[2] if i[2] > 0 else 0
                        self.cierre.agregarcierre(i[1], i[2], i[3], 0, monto, tiempo, self.usuario, id_ca)
                    else:
                        raise ValueError("Método de pago no válido")
                    print(f"✅ Item procesado correctamente")
                except Exception as e:
                    print(f"❌ Error al procesar item {i}: {e}")
                    raise e

            print("🔍 Venta principal completada, procesando historial...")

            # REGISTRAR EN HISTORIAL
            if self.paciente_actual and items_para_historial:
                print(f"🔍 Registrando historial para paciente ID: {self.paciente_actual['id']}")
                try:
                    success = manager.registrar_compra_en_historial(
                        self.paciente_actual['id'],
                        id_ca,
                        items_para_historial,
                        metodo_pago,
                        self.usuario
                    )
                    print(f"🔍 Resultado historial: {success}")

                    if success:
                        self.mostrar_notificacion_sutil(
                            f"✅ Compra registrada en historial de {self.paciente_actual['nombre_completo']}")
                        print("✅ Historial registrado correctamente")
                    else:
                        print("⚠️ Error al registrar historial")

                except Exception as e:
                    print(f"❌ Excepción en historial: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print("ℹ️ No hay paciente vinculado o no hay items para historial")

            # Limpiar paciente
            if self.paciente_actual:
                print("🔍 Limpiando paciente vinculado")
                self.limpiar_paciente_vinculado()

            print("✅ COMPLETADO: ingresar_cierre")
            QMessageBox.information(self, "Venta Completada", "✅ Venta procesada correctamente.")

        except Exception as e:
            print(f"❌ ERROR CRÍTICO en ingresar_cierre: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Error al procesar la venta: {str(e)}")

    def obtener_metodo_pago_seleccionado(self):
        """Obtiene el método de pago seleccionado en los radio buttons"""
        try:
            print(f"🔍 Verificando radio buttons:")
            print(f"🔍 radio_efectivo existe: {hasattr(self, 'radio_efectivo')}")
            print(f"🔍 radio_tarjeta existe: {hasattr(self, 'radio_tarjeta')}")

            if hasattr(self, 'radio_efectivo') and self.radio_efectivo.isChecked():
                print(f"🔍 Efectivo seleccionado")
                return "efectivo"
            elif hasattr(self, 'radio_tarjeta') and self.radio_tarjeta.isChecked():
                print(f"🔍 Tarjeta seleccionada")
                return "tarjeta"
            else:
                # Verificar si alguno está marcado por defecto
                if hasattr(self, 'radio_efectivo'):
                    efectivo_checked = self.radio_efectivo.isChecked()
                    print(f"🔍 Efectivo checked: {efectivo_checked}")
                if hasattr(self, 'radio_tarjeta'):
                    tarjeta_checked = self.radio_tarjeta.isChecked()
                    print(f"🔍 Tarjeta checked: {tarjeta_checked}")

                print(f"🔍 Ningún método seleccionado, usando efectivo por defecto")
                return "efectivo"
        except Exception as e:
            print(f"Error al obtener método de pago: {e}")
            return "efectivo"

    def obtener_medicamento_id_por_nombre(self, nombre_producto):
        """Obtiene el ID del medicamento por su nombre (si es un medicamento)"""
        try:
            manager = sql_structures.Manager()
            # Usar tu función existente para obtener ID
            id_result = manager.get_idddd(nombre_producto)
            if id_result and len(id_result) > 0:
                return id_result[0][0]  # Retornar el ID del medicamento
            return None  # No es un medicamento o no se encontró
        except Exception as e:
            print(f"Error al obtener ID de medicamento: {e}")
            return None

    def monto_e(self):
        try:
            manager = sql_structures.Manager()
            id = manager.print_table("carrito")
            montos = manager.get_monto_cierre("efectivo")
            monto_e = 0.0
            for i in montos:
                monto_e = monto_e + float(i[0])

            montos_t = manager.get_monto_cierre("tarjeta")
            monto_t = 0.0
            for i in montos_t:
                monto_t = monto_t + float(i[0])

            monto_total = monto_t + monto_e
            impuesto_sin_redondeo = monto_total * 0.16
            impuesto = round(impuesto_sin_redondeo, 2)
            self.label_44.setText(str(monto_e))
            self.label_46.setText(str(monto_t))
            self.label_47.setText(str(monto_total))
            self.label_59.setText(str(impuesto))
            netas = monto_total - impuesto
            self.label_65.setText(str(netas))

        except Exception as e:
            print(f"Error: {e}")

    def filtro_dia(self):
        try:
            current_date = self.calendar.selectedDate()
            manager = sql_structures.Manager()
            filter_date = current_date.toString("yyyy-MM-dd")
            id = manager.print_table_dia_cierre(filter_date)
            self.bd_cierre.setRowCount(len(id))
            for i in range(len(id)):
                self.bd_cierre.setItem(i, 0, QTableWidgetItem(str(id[i][0])))
                self.bd_cierre.setItem(i, 1, QTableWidgetItem(str(id[i][1])))
                self.bd_cierre.setItem(i, 2, QTableWidgetItem(str(id[i][2])))
                self.bd_cierre.setItem(i, 3, QTableWidgetItem(str(id[i][3])))
                self.bd_cierre.setItem(i, 4, QTableWidgetItem(str(id[i][4])))
                self.bd_cierre.setItem(i, 5, QTableWidgetItem(str(id[i][5])))
                self.bd_cierre.setItem(i, 6, QTableWidgetItem(str(id[i][6])))
            montos = manager.get_montos_dia("efectivo", filter_date)
            monto_e = 0.0
            for i in montos:
                monto_e = monto_e + float(i[0])

            montos_t = manager.get_montos_dia("tarjeta", filter_date)
            monto_t = 0.0
            for i in montos_t:
                monto_t = monto_t + float(i[0])

            monto_total = monto_t + monto_e

            self.label_44.setText(str(monto_e))
            self.label_46.setText(str(monto_t))
            self.label_47.setText(str(monto_total))
            impuesto = monto_total * 0.16
            self.label_59.setText(str(impuesto))
            netas = monto_total - impuesto
            self.label_65.setText(str(netas))
        except Exception as e:
            print(f"Error: {e}")

    def filtro_semana(self):
        try:
            current_date = self.calendar.selectedDate()
            manager = sql_structures.Manager()
            filter_date = current_date.toString("yyyy-MM-dd")
            id = manager.print_table_semana("cierre", filter_date)
            self.bd_cierre.setRowCount(len(id))
            for i in range(len(id)):
                self.bd_cierre.setItem(i, 0, QTableWidgetItem(str(id[i][0])))
                self.bd_cierre.setItem(i, 1, QTableWidgetItem(str(id[i][1])))
                self.bd_cierre.setItem(i, 2, QTableWidgetItem(str(id[i][2])))
                self.bd_cierre.setItem(i, 3, QTableWidgetItem(str(id[i][3])))
                self.bd_cierre.setItem(i, 4, QTableWidgetItem(str(id[i][4])))
                self.bd_cierre.setItem(i, 5, QTableWidgetItem(str(id[i][5])))
                self.bd_cierre.setItem(i, 6, QTableWidgetItem(str(id[i][6])))
            montos = manager.get_montos_semana("efectivo", filter_date)
            monto_e = 0.0
            for i in montos:
                monto_e = monto_e + float(i[0])

            montos_t = manager.get_montos_semana("tarjeta", filter_date)
            monto_t = 0.0
            for i in montos_t:
                monto_t = monto_t + float(i[0])

            monto_total = monto_t + monto_e

            self.label_44.setText(str(monto_e))
            self.label_46.setText(str(monto_t))
            self.label_47.setText(str(monto_total))
            impuesto = monto_total * 0.16
            self.label_59.setText(str(impuesto))
            netas = monto_total - impuesto
            self.label_65.setText(str(netas))
        except Exception as e:
            print(f"Error: {e}")

    def filtro_mes(self):
        try:
            current_date = self.calendar.selectedDate()
            manager = sql_structures.Manager()
            filter_date = current_date.toString("yyyy-MM-dd")
            id = manager.print_table_mes("cierre", filter_date)
            self.bd_cierre.setRowCount(len(id))
            for i in range(len(id)):
                self.bd_cierre.setItem(i, 0, QTableWidgetItem(str(id[i][0])))
                self.bd_cierre.setItem(i, 1, QTableWidgetItem(str(id[i][1])))
                self.bd_cierre.setItem(i, 2, QTableWidgetItem(str(id[i][2])))
                self.bd_cierre.setItem(i, 3, QTableWidgetItem(str(id[i][3])))
                self.bd_cierre.setItem(i, 4, QTableWidgetItem(str(id[i][4])))
                self.bd_cierre.setItem(i, 5, QTableWidgetItem(str(id[i][5])))
                self.bd_cierre.setItem(i, 6, QTableWidgetItem(str(id[i][6])))
            montos = manager.get_montos_mes("efectivo", filter_date)
            monto_e = 0.0
            for i in montos:
                monto_e = monto_e + float(i[0])

            montos_t = manager.get_montos_mes("tarjeta", filter_date)
            monto_t = 0.0
            for i in montos_t:
                monto_t = monto_t + float(i[0])

            monto_total = monto_t + monto_e

            self.label_44.setText(str(monto_e))
            self.label_46.setText(str(monto_t))
            self.label_47.setText(str(monto_total))
            impuesto = monto_total * 0.16
            self.label_59.setText(str(impuesto))
            netas = monto_total - impuesto
            self.label_65.setText(str(netas))
        except Exception as e:
            print(f"Error: {e}")

    def filtro_anio(self):
        try:
            current_date = self.calendar.selectedDate()
            manager = sql_structures.Manager()
            filter_date = current_date.toString("yyyy-MM-dd")
            id = manager.print_table_año("cierre", filter_date)
            self.bd_cierre.setRowCount(len(id))
            for i in range(len(id)):
                self.bd_cierre.setItem(i, 0, QTableWidgetItem(str(id[i][0])))
                self.bd_cierre.setItem(i, 1, QTableWidgetItem(str(id[i][1])))
                self.bd_cierre.setItem(i, 2, QTableWidgetItem(str(id[i][2])))
                self.bd_cierre.setItem(i, 3, QTableWidgetItem(str(id[i][3])))
                self.bd_cierre.setItem(i, 4, QTableWidgetItem(str(id[i][4])))
                self.bd_cierre.setItem(i, 5, QTableWidgetItem(str(id[i][5])))
                self.bd_cierre.setItem(i, 6, QTableWidgetItem(str(id[i][6])))
            montos = manager.get_montos_año("efectivo", filter_date)
            monto_e = 0.0
            for i in montos:
                monto_e = monto_e + float(i[0])

            montos_t = manager.get_montos_año("tarjeta", filter_date)
            monto_t = 0.0
            for i in montos_t:
                monto_t = monto_t + float(i[0])

            monto_total = monto_t + monto_e
            self.label_44.setText(str(monto_e))
            self.label_46.setText(str(monto_t))
            self.label_47.setText(str(monto_total))
            impuesto = monto_total * 0.16
            self.label_59.setText(str(impuesto))
            netas = monto_total - impuesto
            self.label_65.setText(str(netas))
        except Exception as e:
            print(f"Error: {e}")

    def debug_radio_buttons(self):
        """Función temporal para encontrar los radio buttons"""
        print("🔍 TODOS LOS OBJETOS CON 'radio' EN EL NOMBRE:")
        for attr_name in dir(self):
            if 'radio' in attr_name.lower():
                print(f"  - {attr_name}")

        print("🔍 TODOS LOS OBJETOS CON 'efectivo' O 'tarjeta':")
        for attr_name in dir(self):
            if 'efectivo' in attr_name.lower() or 'tarjeta' in attr_name.lower():
                print(f"  - {attr_name}")

        print("🔍 VERIFICANDO ESTADO ACTUAL:")
        if hasattr(self, 'radio_efectivo'):
            print(f"  - radio_efectivo.isChecked(): {self.radio_efectivo.isChecked()}")
        if hasattr(self, 'radio_tarjeta'):
            print(f"  - radio_tarjeta.isChecked(): {self.radio_tarjeta.isChecked()}")

    def generar_pdf_farmacia(self):
        try:
            # Obtener la ruta del escritorio
            print(1)
            desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
            print(1)
            nombre_archivo = os.path.join(desktop, "ExistenciasMedicamentos.pdf")
            print(1)
            encabezados = ["Nombre", "Presentación", "Laboratorio", "Existencias", "Tarjeta", "Efectivo"]
            ancho_personalizado, alto_personalizado = letter
            pdf = SimpleDocTemplate(nombre_archivo, pagesize=(ancho_personalizado, alto_personalizado),
                                    topMargin=0.15 * inch)
            elementos = []

            # Agregar logo
            #logo = Image('logo.png')
             #logo.drawHeight = 1.5 * inch
            # logo.drawWidth = 1.5 * inch
            #elementos.append(logo)

            # Estilo y título
            styles = getSampleStyleSheet()
            estilo_titulo = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1,
                                           spaceAfter=30)
            titulo = Paragraph("Reporte de Existencias de Medicamentos", estilo_titulo)
            elementos.append(titulo)
            elementos.append(Spacer(1, 20))

            # Estilo de la tabla
            estilo_tabla = TableStyle(
                [('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                 ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                 ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                 ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                 ('FONTSIZE', (0, 0), (-1, 0), 14),
                 ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                 ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                 ('GRID', (0, 0), (-1, -1), 1, colors.black),
                 ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                 ('ALIGN', (0, 0), (-1, 0), 'CENTER')])

            # Obtener datos de la base de datos
            datos_tabla = [encabezados] + list(self.mana.obtener_datos_desde_mysql_medi("medicamentos"))
            ancho_columnas = [1.75 * inch, 1.75 * inch, 1.30 * inch, 1.25 * inch, 1 * inch, 0.75 * inch, 0.75 * inch]
            tabla = Table(datos_tabla, colWidths=ancho_columnas)
            tabla.setStyle(estilo_tabla)

            # Agregar tabla al documento
            elementos.append(tabla)

            # Crear el archivo PDF
            pdf.build(elementos)

            # Abrir el PDF generado
            # os.system(f"start {nombre_archivo}")

        except Exception as e:
            print(e)

    def generar_pdf_terapias(self):
        try:
            # Obtener la ruta del escritorio
            desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
            nombre_archivo = os.path.join(desktop, "Reporte-Terapias.pdf")

            encabezados = ["ID", "Nombre", "Tarjeta", "Efectivo"]

            ancho_personalizado, alto_personalizado = letter
            pdf = SimpleDocTemplate(nombre_archivo, pagesize=(ancho_personalizado, alto_personalizado),
                                    topMargin=0.15 * inch)

            elementos = []

            # logo = Image('logo.png')
            # logo.drawHeight = 1.5 * inch
            # logo.drawWidth = 1.5 * inch
            # elementos.append(logo)

            styles = getSampleStyleSheet()
            estilo_titulo = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                alignment=1,
                spaceAfter=30
            )

            titulo = Paragraph("Reporte de Terapias", estilo_titulo)
            elementos.append(titulo)

            elementos.append(Spacer(1, 20))

            estilo_tabla = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER')
            ])

            # Obtener los datos de la base de datos
            datos_tabla = [encabezados] + list(self.mana.obtener_datos_desde_mysql("terapias"))
            ancho_columnas = [0.5 * inch, 4 * inch, 1.30 * inch, 1.25 * inch]
            tabla = Table(datos_tabla, colWidths=ancho_columnas)
            tabla.setStyle(estilo_tabla)

            # Agregar la tabla a los elementos
            elementos.append(tabla)

            # Crear el archivo PDF
            pdf.build(elementos)

            # Abrir el archivo PDF generado
            # os.system(f"start {nombre_archivo}")

        except Exception as e:
            print(e)

    def generar_pdf_jornadas(self):
        try:
            # Obtener la ruta del escritorio
            desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
            nombre_archivo = os.path.join(desktop, "Reporte-Jornadas-Existentes.pdf")

            encabezados = ["ID", "Nombre", "Efectivo", "Tarjeta"]

            ancho_personalizado, alto_personalizado = letter
            pdf = SimpleDocTemplate(nombre_archivo, pagesize=(ancho_personalizado, alto_personalizado),
                                    topMargin=0.15 * inch)

            elementos = []

            # logo = Image('logo.png')
            # logo.drawHeight = 1.5 * inch
            # logo.drawWidth = 1.5 * inch
            # elementos.append(logo)

            styles = getSampleStyleSheet()
            estilo_titulo = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                alignment=1,
                spaceAfter=30
            )

            titulo = Paragraph("Reporte de Jornadas Existentes", estilo_titulo)
            elementos.append(titulo)

            elementos.append(Spacer(1, 20))

            estilo_tabla = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER')
            ])

            datos_tabla = [encabezados] + list(self.mana.obtener_datos_desde_mysql("jornadas"))
            ancho_columnas = [0.5 * inch, 2.5 * inch, 1.75 * inch, 1.75 * inch]
            tabla = Table(datos_tabla, colWidths=ancho_columnas)
            tabla.setStyle(estilo_tabla)

            elementos.append(tabla)

            # Crear el archivo PDF
            pdf.build(elementos)

            # Abrir el archivo PDF generado
            # os.system(f"start {nombre_archivo}")

        except Exception as e:
            print(e)

    def generar_pdf_paciente(self):
        try:
            # Obtener la ruta del escritorio
            desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
            nombre_archivo = os.path.join(desktop, "Reporte-Pacientes-Existentes.pdf")

            encabezados = ["Nombre", "Apellido", "Telefono", "Dpi", "Cita", "Cumpleaños"]

            # Cambiar el tamaño de la página
            ancho_personalizado = 14 * inch  # Ancho personalizado (más ancho que letter)
            alto_personalizado = 8.5 * inch  # Altura estándar
            pdf = SimpleDocTemplate(
                nombre_archivo,
                pagesize=(ancho_personalizado, alto_personalizado),
                topMargin=0.15 * inch
            )

            elementos = []

            # logo = Image('logo.png')
            # logo.drawHeight = 1.5 * inch
            # logo.drawWidth = 1.5 * inch
            # elementos.append(logo)

            styles = getSampleStyleSheet()
            estilo_titulo = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                alignment=1,
                spaceAfter=30
            )

            titulo = Paragraph("Reporte de Pacientes Existentes", estilo_titulo)
            elementos.append(titulo)

            elementos.append(Spacer(1, 20))

            estilo_tabla = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER')
            ])

            datos_tabla = [encabezados] + list(self.mana.obtener_datos_desde_mysql_paciente("paciente"))
            ancho_columnas = [2 * inch, 2 * inch, 2 * inch, 2 * inch, 2 * inch, 2 * inch]
            tabla = Table(datos_tabla, colWidths=ancho_columnas)
            tabla.setStyle(estilo_tabla)

            elementos.append(tabla)
            pdf.build(elementos)

            # Abrir el archivo PDF generado
            # os.system(f"start {nombre_archivo}")

        except Exception as e:
            print(f"Error generar pdf paciente: {e}")

    def generar_pdf_combos(self):
        try:
            # Obtener la ruta del escritorio
            desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"
            nombre_archivo = os.path.join(desktop, "Reporte-Combos-y-Promociones-Existentes.pdf")

            encabezados = ["ID", "Terapia", "Tarjeta", "Efectivo"]

            ancho_personalizado, alto_personalizado = letter
            pdf = SimpleDocTemplate(nombre_archivo, pagesize=(ancho_personalizado, alto_personalizado),
                                    topMargin=0.15 * inch)

            elementos = []

            # logo = Image('logo.png')
            # logo.drawHeight = 1.5 * inch
            # logo.drawWidth = 1.5 * inch
            # elementos.append(logo)

            styles = getSampleStyleSheet()
            estilo_titulo = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                alignment=1,
                spaceAfter=30
            )

            titulo = Paragraph("Reporte de Combos y Promociones Existentes", estilo_titulo)
            elementos.append(titulo)

            elementos.append(Spacer(1, 20))

            estilo_tabla = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER')
            ])

            datos_tabla = [encabezados] + list(self.mana.obtener_datos_desde_mysql("promociones"))
            ancho_columnas = [0.5 * inch, 2.5 * inch, 1.75 * inch, 1.75 * inch, 2 * inch, 1 * inch, 1 * inch]
            tabla = Table(datos_tabla, colWidths=ancho_columnas)
            tabla.setStyle(estilo_tabla)

            elementos.append(tabla)
            pdf.build(elementos)

            # Abrir el archivo PDF generado
            # os.system(f"start {nombre_archivo}")

        except Exception as e:
            print(f"Error generar pdf combos: {e}")

    def generar_pdf_cierre(self):
        from datetime import datetime
        try:
            today = datetime.today()
            # Obtener la fecha actual en formato "YYYY-MM-DD"
            fecha_actual = f"{today.year}{today.month:02d}{today.day:02d}"

            # Nombre del archivo PDF usando la fecha
            nombre_pdf = f"Reporte_Cierre_Inventario_{fecha_actual}.pdf"

            # Obtener la ruta del escritorio
            desktop = "C:\\Users\\andre\\OneDrive\\Escritorio"

            # Ruta completa donde se guardará el archivo en el escritorio
            ruta_pdf = os.path.join(desktop, nombre_pdf)

            # Crear el PDF en la ruta especificada
            pdf = canvas.Canvas(ruta_pdf, pagesize=landscape(legal))
            width, height = landscape(legal)
            # logo_path = "logo.png"

            def dibujar_encabezado(pdf, width, height, headers):
                # Dibujar logo
                # image_width = 75
                # image_height = 75
                # pdf.drawImage(logo_path, 50, height - image_height - 10, width=image_width, height=image_height)

                # Título del reporte
                pdf.setFont("Helvetica-Bold", 20)
                pdf.drawCentredString(width / 2, height - 60, "Reporte de Datos: Cierre de Inventario")

                # Encabezados de la tabla
                x = 30
                y = height - 100
                row_height = 30
                col_width = 126

                pdf.setFillColorRGB(0.1, 0.4, 0.7)
                pdf.setStrokeColorRGB(0, 0, 0)
                pdf.rect(x, y - row_height, col_width * len(headers), row_height, fill=1)
                pdf.setFillColorRGB(1, 1, 1)

                # Reducir el tamaño de fuente de los encabezados a 10
                pdf.setFont("Helvetica-Bold", 10)

                for col, header in enumerate(headers):
                    pdf.drawCentredString(x + col * col_width + col_width / 2, y - row_height / 2, header)

                return x, y - 2 * row_height

            # Obtener encabezados
            headers = [self.bd_cierre.horizontalHeaderItem(i).text() for i in range(self.bd_cierre.columnCount())]

            # Dibujar encabezado en la primera página
            x, y = dibujar_encabezado(pdf, width, height, headers)

            # Configuraciones iniciales
            row_height = 30
            col_width = 126

            # Iterar por todas las filas
            for row in range(self.bd_cierre.rowCount()):
                # Verificar si necesitamos una nueva página
                if y - row_height < 50:  # Margen inferior
                    pdf.showPage()
                    # Redibujar encabezados en la nueva página
                    x, y = dibujar_encabezado(pdf, width, height, headers)

                # Dibujar fila
                pdf.setFillColorRGB(0, 0, 0)
                for col in range(self.bd_cierre.columnCount()):
                    # Dibujar rectángulo de celda
                    pdf.rect(x + col * col_width, y - row_height, col_width, row_height, stroke=1, fill=0)

                    # Obtener y dibujar texto
                    item = self.bd_cierre.item(row, col)
                    if item is not None:
                        text = item.text()
                        pdf.drawCentredString(x + col * col_width + col_width / 2, y - row_height / 2, text)

                # Mover coordenada Y hacia abajo
                y -= row_height

            num_columns = self.tabla_carrito().columnCount()

            # Fila de Ventas efectivo
            pdf.setFillColorRGB(0.8, 0.8, 1)  # Color de fondo
            pdf.rect(x, y - row_height, col_width * num_columns, row_height, stroke=1, fill=1)
            pdf.setFillColorRGB(0, 0, 0)  # Color de texto
            pdf.drawCentredString(x + col_width / 2, y - row_height / 2, "Ventas Efectivo")
            pdf.drawCentredString(x + col_width * 1.5, y - row_height / 2, f"Q{self.label_44.text()}")
            y -= row_height  # Mover coordenada Y hacia abajo

            # Fila de Ventas tarjeta
            pdf.setFillColorRGB(0.8, 0.8, 1)  # Color de fondo
            pdf.rect(x, y - row_height, col_width * num_columns, row_height, stroke=1, fill=1)
            pdf.setFillColorRGB(0, 0, 0)  # Color de texto
            pdf.drawCentredString(x + col_width / 2, y - row_height / 2, "Ventas Tarjeta")
            pdf.drawCentredString(x + col_width * 1.5, y - row_height / 2, f"Q{self.label_46.text()}")
            y -= row_height  # Mover coordenada Y hacia abajo

            # Fila de Ventas Brutas
            pdf.setFillColorRGB(0.8, 0.8, 1)  # Color de fondo
            pdf.rect(x, y - row_height, col_width * num_columns, row_height, stroke=1, fill=1)
            pdf.setFillColorRGB(0, 0, 0)  # Color de texto
            pdf.drawCentredString(x + col_width / 2, y - row_height / 2, "Ventas Brutas")
            pdf.drawCentredString(x + col_width * 1.5, y - row_height / 2, f"Q{self.label_47.text()}")
            y -= row_height  # Mover coordenada Y hacia abajo

            # Fila de Impuestos
            pdf.setFillColorRGB(0.8, 0.8, 1)  # Color de fondo
            pdf.rect(x, y - row_height, col_width * num_columns, row_height, stroke=1, fill=1)
            pdf.setFillColorRGB(0, 0, 0)  # Color de texto
            pdf.drawCentredString(x + col_width / 2, y - row_height / 2, "Impuestos 16%")
            pdf.drawCentredString(x + col_width * 1.5, y - row_height / 2, f"Q{self.label_59.text()}")
            y -= row_height  # Mover coordenada Y hacia abajo

            # Fila de Ventas Netas
            pdf.setFillColorRGB(0.8, 0.8, 1)  # Color de fondo
            pdf.rect(x, y - row_height, col_width * num_columns, row_height, stroke=1, fill=1)
            pdf.setFillColorRGB(0, 0, 0)  # Color de texto
            pdf.drawCentredString(x + col_width / 2, y - row_height / 2, "Ventas Netas")
            pdf.drawCentredString(x + col_width * 1.5, y - row_height / 2, f"Q{self.label_65.text()}")
            y -= row_height  # Mover coordenada Y hacia abajo

            # Guardar el PDF
            pdf.save()

            # Notificación de éxito
            QMessageBox.information(self, "Éxito", f"Reporte guardado en: {ruta_pdf}")

            # Abrir el archivo PDF generado
            # os.system(f"start {ruta_pdf}")

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Ocurrió un error: {e}")

    def tabla_carrito(self):
        return self.bd_carrito

    def tabla_inventario(self):
        return self.bd_cierre

    @classmethod
    def get_dinero_total(cls):
        return cls._dinero_total

    @classmethod
    def get_tip_pago(cls):
        print(f"dato: {cls._dato}")
        return cls._dato

    @classmethod
    def get_porcentaje(cls):
        return cls._contra

    @classmethod
    def get_diferencia_efectivo(cls):
        return cls._diferencia_efectivo

    @classmethod
    def get_contra(cls):
        return cls._contra

    @classmethod
    def enviar_usuario(cls):
        return cls._usuario

    @classmethod
    def enviar_detalle(cls):
        return cls._id_detalle

    def actualizar_totales_carrito(self):
        """Actualiza los totales del carrito basándose en los precios actuales mostrados"""
        try:
            from .descuentosMedi import DescuentoMedi

            # Verificar si hay productos en el carrito
            if self.bd_carrito.rowCount() == 0:
                # Si no hay productos, mantener descuentos pero poner totales en 0
                self.label_28.setText("0.00")
                # NO resetear descuentos aquí - mantener el valor actual
                try:
                    porcentaje = DescuentoMedi.get_porcentaje()
                    cantidad = DescuentoMedi.get_cantidad()
                    if porcentaje > 0 or cantidad > 0:
                        # Mostrar el descuento aunque no haya productos
                        descuento_actual = cantidad if cantidad > 0 else 0
                        self.label_29.setText(f"{descuento_actual:.2f}")
                    else:
                        self.label_29.setText("0.00")
                except:
                    self.label_29.setText("0.00")

                self.label_30.setText("0.00")
                return

            # Obtener valores de descuentos
            porcentaje = 0
            cantidad = 0

            try:
                porcentaje = DescuentoMedi.get_porcentaje()
            except:
                porcentaje = 0

            try:
                cantidad = DescuentoMedi.get_cantidad()
            except:
                cantidad = 0

            subtotal = 0.0

            # Calcular subtotal basándose en los valores actuales mostrados en la tabla
            for fila in range(self.bd_carrito.rowCount()):
                precio_total_item = self.bd_carrito.item(fila, 4)  # Columna de precio total
                if precio_total_item and precio_total_item.text():
                    try:
                        subtotal += float(precio_total_item.text())
                    except ValueError:
                        continue

            # Calcular descuentos
            descuento = 0.0
            if porcentaje > 0:
                descuento = subtotal * porcentaje
            elif cantidad > 0:
                descuento = min(cantidad, subtotal)  # No permitir descuento mayor al subtotal

            total = max(0, subtotal - descuento)  # No permitir totales negativos

            # Actualizar etiquetas
            self.label_28.setText(f"{subtotal:.2f}")
            self.label_29.setText(f"{descuento:.2f}")
            self.label_30.setText(f"{total:.2f}")

            print(f"Totales actualizados - Subtotal: {subtotal:.2f}, Descuento: {descuento:.2f}, Total: {total:.2f}")

        except Exception as e:
            print(f"Error al actualizar totales: {e}")
            # En caso de error, mantener los valores actuales sin resetear descuentos
            pass

    def setup_column_selector_button(self):
        """Configurar botón para abrir selector de columnas"""
        try:
            from PyQt5.QtWidgets import QPushButton, QHBoxLayout, QVBoxLayout, QFrame
            from PyQt5.QtCore import QSize

            # Crear el botón
            self.btn_configurar_columnas = QPushButton("⚙ Configurar Columnas")
            self.btn_configurar_columnas.setToolTip("Seleccionar qué columnas mostrar en la tabla de farmacia")
            self.btn_configurar_columnas.setMinimumSize(QSize(150, 35))
            self.btn_configurar_columnas.clicked.connect(self.abrir_selector_columnas)

            # Aplicar estilo
            self.btn_configurar_columnas.setStyleSheet("""
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 11px;
                }
                QPushButton:hover {
                    background-color: #1976D2;
                }
            """)

            # Encontrar el widget contenedor de la tabla de farmacia
            parent_widget = self.bd_farmacia.parent()
            if parent_widget:
                current_layout = parent_widget.layout()
                if current_layout:
                    button_layout = QHBoxLayout()
                    button_layout.addWidget(self.btn_configurar_columnas)
                    button_layout.addStretch()

                    button_frame = QFrame()
                    button_frame.setLayout(button_layout)
                    button_frame.setMaximumHeight(50)
                    current_layout.insertWidget(0, button_frame)

            print("✅ Botón 'Configurar Columnas' agregado exitosamente")

        except Exception as e:
            print(f"⚠️ Error configurando botón selector columnas: {e}")

    def abrir_selector_columnas(self):
        """Abrir el diálogo para seleccionar columnas"""
        try:
            self.farmacia_column_manager.open_column_selector()
        except Exception as e:
            print(f"⚠️ Error abriendo selector de columnas: {e}")

    # PASO 5D: AGREGAR ESTAS FUNCIONES AL FINAL DE LA CLASE VentanaFuncional

    def vincular_paciente_venta(self):
        """Abre el diálogo para vincular un paciente a la venta"""
        try:
            from .seleccion_paciente import SeleccionPacienteDialog

            dialog = SeleccionPacienteDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                self.paciente_actual = dialog.paciente_seleccionado
                self.actualizar_info_paciente_vinculado()

                # CAMBIO: Mensaje sutil en lugar de QMessageBox grande
                self.mostrar_notificacion_sutil(f"✅ {self.paciente_actual['nombre_completo']} vinculado")

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error al vincular paciente: {e}")
            print(f"Error en vincular_paciente_venta: {e}")

    def mostrar_notificacion_sutil(self, mensaje):
        """Muestra una notificación pequeña y temporal"""
        try:
            # Crear label temporal para notificación
            if not hasattr(self, 'notificacion_label'):
                self.notificacion_label = QLabel(self)

            # Estilo de la notificación
            estilo_notificacion = """
                QLabel {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #4CAF50, stop:1 #45a049);
                    border: 2px solid #2e7d32;
                    border-radius: 15px;
                    color: white;
                    font-weight: bold;
                    font-size: 12px;
                    padding: 8px 15px;
                }
            """

            self.notificacion_label.setStyleSheet(estilo_notificacion)
            self.notificacion_label.setText(mensaje)
            self.notificacion_label.adjustSize()

            # Posicionar en esquina superior derecha
            x = self.width() - self.notificacion_label.width() - 20
            y = 50
            self.notificacion_label.move(x, y)
            self.notificacion_label.show()

            # Timer para ocultar después de 3 segundos
            QTimer.singleShot(3000, self.notificacion_label.hide)

        except Exception as e:
            print(f"Error en notificación: {e}")

    def actualizar_info_paciente_vinculado(self):
        """Actualiza la interfaz para mostrar el paciente vinculado"""
        try:
            if self.paciente_actual:
                # CAMBIO: Formato en múltiples líneas para espacio estrecho
                nombre = self.paciente_actual['nombre_completo']
                dpi = self.paciente_actual['dpi']

                # Información en múltiples líneas
                info_text = f"👤 {nombre}\nDPI: {dpi}"

                # Intentar obtener última compra
                try:
                    manager = sql_structures.Manager()
                    historial = manager.obtener_historial_compras_paciente(self.paciente_actual['id'], 1)
                    if historial:
                        fecha_ultima = str(historial[0][6])[:10]  # Solo fecha, sin hora
                        info_text += f"\nÚltima: {fecha_ultima}"
                except:
                    pass

                self.label_paciente_actual.setText(info_text)

            else:
                self.label_paciente_actual.setText("👤 Sin paciente\nvinculado")
        except Exception as e:
            print(f"Error al actualizar info paciente: {e}")

    def obtener_items_carrito_para_historial(self):
        """Obtiene los items del carrito en formato para historial"""
        items = []
        try:
            # Usar tu manager existente para obtener datos del carrito
            manager = sql_structures.Manager()
            carrito_data = manager.obtener_carrito_completo()

            for item in carrito_data:
                # Adaptar según tu estructura de carrito
                items.append({
                    'medicamento_id': item[5] if len(item) > 5 else None,  # medicamentos_id
                    'nombre': item[1],  # nombre
                    'cantidad': item[2],  # cantidad
                    'precio_unitario': item[3],  # precio unitario
                    'precio_total': item[4]  # precio total
                })

            return items

        except Exception as e:
            print(f"Error al obtener items del carrito: {e}")
            return []

    def ver_historial_paciente(self):
        """Abre la ventana de historial del paciente actual"""
        if not self.paciente_actual:
            QMessageBox.warning(self, "Sin Paciente", "Primero debe vincular un paciente a la venta.")
            return

        try:
            from .historial_paciente import HistorialPacienteDialog

            dialog = HistorialPacienteDialog(self.paciente_actual, self)
            dialog.exec_()

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error al abrir historial: {e}")
            print(f"Error en ver_historial_paciente: {e}")

    def limpiar_paciente_vinculado(self):
        """Limpia el paciente vinculado"""
        try:
            self.paciente_actual = None
            self.actualizar_info_paciente_vinculado()
            self.mostrar_notificacion_sutil("🧹 Paciente desvinculado")

        except Exception as e:
            print(f"Error al limpiar paciente: {e}")












class RowHighlighter:
    def __init__(self, table):
        self.table = table
        self.highlighted_rows = {}  # {row: (color, tipo_alerta)}

        # Colores para diferentes tipos de alerta
        self.STOCK_COLOR = QColor(254, 44, 85)  # Rojo para stock bajo
        self.EXPIRY_COLOR = QColor(8, 101, 254)  # Azul para próximo a vencer

    def highlight_row(self, row, is_stock_alert):

        try:
            color = self.STOCK_COLOR if is_stock_alert else self.EXPIRY_COLOR
            tipo = "stock" if is_stock_alert else "expiry"

            # Guardar el estado de la fila
            self.highlighted_rows[row] = (color, tipo)

            # Aplicar el color a toda la fila
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    item.setBackground(color)
                    item.setForeground(Qt.white)

            # Forzar actualización de la tabla
            self.table.viewport().repaint()

        except Exception as e:
            print(f"Error al resaltar fila: {e}")

    def remove_highlight(self, row):

        if row in self.highlighted_rows:
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    item.setBackground(Qt.white)
                    item.setForeground(Qt.black)
            del self.highlighted_rows[row]
            self.table.viewport().repaint()

    def restore_highlights(self):

        for row, (color, _) in self.highlighted_rows.items():
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    item.setBackground(color)
                    item.setForeground(Qt.white)
        self.table.viewport().repaint()