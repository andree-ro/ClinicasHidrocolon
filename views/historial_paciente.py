# CREAR ARCHIVO: views/historial_paciente.py
# Guarda este archivo en tu carpeta views/
from datetime import datetime, date, timedelta
import datetime as dt
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QPixmap
import sql_structures
from datetime import datetime


class HistorialPacienteDialog(QDialog):
    def __init__(self, paciente_info, parent=None):
        super().__init__(parent)
        self.paciente_info = paciente_info
        self.manager = sql_structures.Manager()

        self.setWindowTitle(f"Historial de Compras - {paciente_info['nombre_completo']}")
        self.setModal(True)
        self.resize(900, 600)

        self.setupUI()
        self.cargar_historial()

    def setupUI(self):
        """Configurar la interfaz de usuario"""
        layout = QVBoxLayout()

        # ===== HEADER CON INFO DEL PACIENTE =====
        header_frame = QFrame()
        header_frame.setFrameStyle(QFrame.StyledPanel)
        header_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                           stop:0 #e3f2fd, stop:1 #bbdefb);
                border: 2px solid #1976d2;
                border-radius: 8px;
                margin: 5px;
            }
        """)

        header_layout = QHBoxLayout(header_frame)

        # Info del paciente
        info_layout = QVBoxLayout()

        nombre_label = QLabel(f"👤 {self.paciente_info['nombre_completo']}")
        nombre_label.setFont(QFont("Arial", 14, QFont.Bold))
        nombre_label.setStyleSheet("color: #1565c0; margin: 5px;")

        dpi_label = QLabel(f"📄 DPI: {self.paciente_info['dpi']}")
        dpi_label.setStyleSheet("color: #424242; font-size: 12px; margin: 2px;")

        info_layout.addWidget(nombre_label)
        info_layout.addWidget(dpi_label)

        # Estadísticas rápidas
        stats_layout = QVBoxLayout()
        self.lbl_total_compras = QLabel("📊 Calculando estadísticas...")
        self.lbl_total_compras.setStyleSheet("color: #2e7d32; font-weight: bold; margin: 2px;")
        self.lbl_ultima_compra = QLabel("📅 Última compra: Calculando...")
        self.lbl_ultima_compra.setStyleSheet("color: #2e7d32; font-weight: bold; margin: 2px;")

        stats_layout.addWidget(self.lbl_total_compras)
        stats_layout.addWidget(self.lbl_ultima_compra)

        header_layout.addLayout(info_layout)
        header_layout.addStretch()
        header_layout.addLayout(stats_layout)

        # ===== FILTROS =====
        filtros_frame = QFrame()
        filtros_layout = QHBoxLayout(filtros_frame)

        filtros_layout.addWidget(QLabel("🔍 Filtrar por:"))

        self.combo_filtro = QComboBox()
        self.combo_filtro.addItems([
            "Todas las compras",
            "Última semana",
            "Último mes",
            "Últimos 3 meses",
            "Solo medicamentos",
            "Solo servicios"
        ])
        self.combo_filtro.currentTextChanged.connect(self.aplicar_filtro)

        self.txt_buscar = QLineEdit()
        self.txt_buscar.setPlaceholderText("Buscar producto...")
        self.txt_buscar.textChanged.connect(self.buscar_en_historial)

        btn_exportar = QPushButton("📄 Exportar")
        btn_exportar.clicked.connect(self.exportar_historial)
        btn_exportar.setStyleSheet("""
            QPushButton {
                background: #4caf50;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #45a049;
            }
        """)

        filtros_layout.addWidget(self.combo_filtro)
        filtros_layout.addWidget(self.txt_buscar)
        filtros_layout.addStretch()
        filtros_layout.addWidget(btn_exportar)

        # ===== TABLA DE HISTORIAL =====
        self.tabla_historial = QTableWidget()
        self.tabla_historial.setColumnCount(8)
        self.tabla_historial.setHorizontalHeaderLabels([
            "Fecha", "Producto", "Cantidad", "Precio Unit.",
            "Total", "Pago", "Vendedor", "Observaciones"
        ])

        # Configurar tabla
        header = self.tabla_historial.horizontalHeader()
        header.setStretchLastSection(True)
        header.resizeSection(0, 120)  # Fecha
        header.resizeSection(1, 200)  # Producto
        header.resizeSection(2, 80)  # Cantidad
        header.resizeSection(3, 100)  # Precio Unit
        header.resizeSection(4, 100)  # Total
        header.resizeSection(5, 80)  # Pago
        header.resizeSection(6, 120)  # Vendedor

        self.tabla_historial.setAlternatingRowColors(True)
        self.tabla_historial.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla_historial.setStyleSheet("""
            QTableWidget {
                gridline-color: #e0e0e0;
                selection-background-color: #e3f2fd;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #f0f0f0;
            }
        """)

        # ===== BOTONES =====
        botones_layout = QHBoxLayout()

        btn_actualizar = QPushButton("🔄 Actualizar")
        btn_actualizar.clicked.connect(self.cargar_historial)

        btn_cerrar = QPushButton("❌ Cerrar")
        btn_cerrar.clicked.connect(self.accept)

        botones_layout.addStretch()
        botones_layout.addWidget(btn_actualizar)
        botones_layout.addWidget(btn_cerrar)

        # ===== LAYOUT PRINCIPAL =====
        layout.addWidget(header_frame)
        layout.addWidget(filtros_frame)
        layout.addWidget(self.tabla_historial)
        layout.addLayout(botones_layout)

        self.setLayout(layout)

    def cargar_historial(self):
        """Cargar el historial de compras del paciente"""
        try:
            # Obtener historial completo
            historial = self.manager.obtener_historial_compras_paciente(self.paciente_info['id'])

            if not historial:
                self.tabla_historial.setRowCount(0)
                QMessageBox.information(self, "Sin historial",
                                        "Este paciente no tiene compras registradas.")
                return

            # Guardar historial para filtros
            self.historial_completo = historial

            # Mostrar en tabla
            self.mostrar_historial_en_tabla(historial)

            # Actualizar estadísticas
            self.actualizar_estadisticas(historial)

        except Exception as e:
            print(f"Error al cargar historial: {e}")
            QMessageBox.critical(self, "Error", f"Error al cargar el historial: {e}")

    def mostrar_historial_en_tabla(self, historial):
        """Mostrar el historial en la tabla"""
        self.tabla_historial.setRowCount(len(historial))

        for fila, compra in enumerate(historial):
            try:
                # Formatear fecha
                fecha = compra[6]  # fecha_compra
                if isinstance(fecha, str):
                    fecha_formateada = fecha[:16]  # Solo fecha y hora
                else:
                    fecha_formateada = fecha.strftime("%Y-%m-%d %H:%M")

                # Llenar celdas
                self.tabla_historial.setItem(fila, 0, QTableWidgetItem(fecha_formateada))
                self.tabla_historial.setItem(fila, 1, QTableWidgetItem(str(compra[1])))  # producto_nombre
                self.tabla_historial.setItem(fila, 2, QTableWidgetItem(str(compra[2])))  # cantidad
                self.tabla_historial.setItem(fila, 3, QTableWidgetItem(f"Q{float(compra[3]):.2f}"))  # precio_unitario
                self.tabla_historial.setItem(fila, 4, QTableWidgetItem(f"Q{float(compra[4]):.2f}"))  # precio_total
                self.tabla_historial.setItem(fila, 5, QTableWidgetItem(str(compra[5])))  # tipo_pago
                self.tabla_historial.setItem(fila, 6, QTableWidgetItem(str(compra[7]) if compra[7] else ""))  # vendedor
                self.tabla_historial.setItem(fila, 7,
                                             QTableWidgetItem(str(compra[8]) if compra[8] else ""))  # observaciones

                # Centrar algunos elementos
                for col in [2, 3, 4, 5]:  # cantidad, precios, pago
                    if self.tabla_historial.item(fila, col):
                        self.tabla_historial.item(fila, col).setTextAlignment(Qt.AlignCenter)

            except Exception as e:
                print(f"Error en fila {fila}: {e}")
                continue

    def actualizar_estadisticas(self, historial):
        """Actualizar las estadísticas mostradas"""
        try:
            if not historial:
                self.lbl_total_compras.setText("📊 Sin compras registradas")
                self.lbl_ultima_compra.setText("📅 Sin compras")
                return

            # Total de compras y monto
            total_compras = len(historial)
            monto_total = sum(float(compra[4]) for compra in historial)

            self.lbl_total_compras.setText(f"📊 {total_compras} compras - Total: Q{monto_total:.2f}")

            # Última compra
            ultima_fecha = historial[0][6]  # El historial viene ordenado DESC
            if isinstance(ultima_fecha, str):
                fecha_str = ultima_fecha[:10]
            else:
                fecha_str = ultima_fecha.strftime("%Y-%m-%d")

            self.lbl_ultima_compra.setText(f"📅 Última compra: {fecha_str}")

        except Exception as e:
            print(f"Error al actualizar estadísticas: {e}")

    def aplicar_filtro(self):
        """Aplicar filtro seleccionado"""
        if not hasattr(self, 'historial_completo'):
            return

        filtro = self.combo_filtro.currentText()
        historial_filtrado = self.historial_completo.copy()

        try:
            if filtro == "Última semana":
                from datetime import datetime, timedelta
                hace_semana = datetime.now() - timedelta(days=7)
                historial_filtrado = [h for h in historial_filtrado
                                      if datetime.strptime(str(h[6])[:10], "%Y-%m-%d") >= hace_semana]

            elif filtro == "Último mes":
                from datetime import datetime, timedelta
                hace_mes = datetime.now() - timedelta(days=30)
                historial_filtrado = [h for h in historial_filtrado
                                      if datetime.strptime(str(h[6])[:10], "%Y-%m-%d") >= hace_mes]

            elif filtro == "Últimos 3 meses":
                from datetime import datetime, timedelta
                hace_tres_meses = datetime.now() - timedelta(days=90)
                historial_filtrado = [h for h in historial_filtrado
                                      if datetime.strptime(str(h[6])[:10], "%Y-%m-%d") >= hace_tres_meses]

            elif filtro == "Solo medicamentos":
                # Filtrar solo items que tienen medicamento_id (no None)
                # Esto depende de tu estructura de datos
                pass

            elif filtro == "Solo servicios":
                # Filtrar solo servicios/terapias
                pass

            self.mostrar_historial_en_tabla(historial_filtrado)
            self.actualizar_estadisticas(historial_filtrado)

        except Exception as e:
            print(f"Error al aplicar filtro: {e}")

    def buscar_en_historial(self):
        """Buscar en el historial por nombre de producto"""
        if not hasattr(self, 'historial_completo'):
            return

        texto_busqueda = self.txt_buscar.text().lower()

        if not texto_busqueda:
            self.mostrar_historial_en_tabla(self.historial_completo)
            return

        # Filtrar por nombre de producto
        historial_filtrado = [
            h for h in self.historial_completo
            if texto_busqueda in str(h[1]).lower()  # producto_nombre
        ]

        self.mostrar_historial_en_tabla(historial_filtrado)

    def exportar_historial(self):
        """Exportar historial a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import os

            wb = Workbook()
            ws = wb.active
            ws.title = "Historial de Compras"

            # Headers
            headers = ["Fecha", "Producto", "Cantidad", "Precio Unit.",
                       "Total", "Pago", "Vendedor", "Observaciones"]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Datos
            historial = getattr(self, 'historial_completo', [])
            for fila, compra in enumerate(historial, 2):
                ws.cell(row=fila, column=1, value=str(compra[6])[:16])
                ws.cell(row=fila, column=2, value=str(compra[1]))
                ws.cell(row=fila, column=3, value=compra[2])
                ws.cell(row=fila, column=4, value=float(compra[3]))
                ws.cell(row=fila, column=5, value=float(compra[4]))
                ws.cell(row=fila, column=6, value=str(compra[5]))
                ws.cell(row=fila, column=7, value=str(compra[7]) if compra[7] else "")
                ws.cell(row=fila, column=8, value=str(compra[8]) if compra[8] else "")

            # Guardar
            nombre_archivo = f"historial_{self.paciente_info['nombre_completo'].replace(' ', '_')}.xlsx"
            ruta_archivo = os.path.join(os.path.expanduser("~"), "Desktop", nombre_archivo)

            wb.save(ruta_archivo)

            QMessageBox.information(self, "Exportación exitosa",
                                    f"Historial exportado a:\n{ruta_archivo}")

        except ImportError:
            QMessageBox.warning(self, "Módulo faltante",
                                "Necesitas instalar openpyxl para exportar:\npip install openpyxl")
        except Exception as e:
            print(f"Error al exportar: {e}")
            QMessageBox.critical(self, "Error", f"Error al exportar: {e}")