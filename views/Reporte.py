import sys
from datetime import datetime, date, timedelta
import datetime as dt
import pymysql
from datetime import datetime

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QComboBox, QPushButton,
                             QTableWidget, QTableWidgetItem, QHeaderView,
                             QDateEdit, QMessageBox, QFrame, QCheckBox, QTextEdit,
                             QFileDialog)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QPainter, QPen, QColor
import pandas as pd

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class AplicacionReporteFinanciero(QMainWindow):
    def __init__(self):
        super().__init__()
        self.iniciar_ui()
        self.conectar_base_datos()

    def conectar_base_datos(self):
        try:
            self.conn = pymysql.connect(
                host="127.0.0.1",
                user="root",
                password="2332",  # Reemplaza con tu contraseña real
                database="bdhidrocolon"
            )
            print("Conexión a la base de datos exitosa")
            # Verificar si la tabla tiene la columna de justificación
            cursor = self.conn.cursor()
            cursor.execute("SHOW COLUMNS FROM cierre_neto LIKE 'justificacion'")
            result = cursor.fetchone()
            # Si no existe la columna, agregarla
            if not result:
                cursor.execute("ALTER TABLE cierre_neto ADD COLUMN justificacion TEXT")
                self.conn.commit()
                print("Columna de justificación agregada a la tabla cierre_neto")

            cursor.close()
        except Exception as e:
            print(f"Error de conexión a la base de datos: {e}")
            QMessageBox.critical(self, "Error de Conexión",
                                 f"No se pudo conectar a la base de datos: {e}")

    def iniciar_ui(self):
        self.setWindowTitle("Reporte Financiero Mensual")
        self.setGeometry(700, 100, 900, 900)  # Aumenté el tamaño para acomodar la justificación

        # Widget y diseño principal
        widget_principal = QWidget()
        diseno_principal = QVBoxLayout()

        # Controles para rango de fechas
        widget_fecha = QWidget()
        diseno_fecha = QHBoxLayout()

        diseno_fecha.addWidget(QLabel("Del:"))
        self.fecha_inicio = QDateEdit()
        self.fecha_inicio.setCalendarPopup(True)
        self.fecha_inicio.setDate(QDate.currentDate().addMonths(-1))
        diseno_fecha.addWidget(self.fecha_inicio)

        diseno_fecha.addWidget(QLabel("Al:"))
        self.fecha_fin = QDateEdit()
        self.fecha_fin.setCalendarPopup(True)
        self.fecha_fin.setDate(QDate.currentDate())
        diseno_fecha.addWidget(self.fecha_fin)

        boton_generar = QPushButton("Generar Reporte")
        boton_generar.clicked.connect(self.generar_reporte)
        diseno_fecha.addWidget(boton_generar)

        widget_fecha.setLayout(diseno_fecha)
        diseno_principal.addWidget(widget_fecha)

        botones_layout = QHBoxLayout()

        self.btn_exportar_pdf = QPushButton("Generar PDF")
        self.btn_exportar_pdf.clicked.connect(self.generar_pdf)
        botones_layout.addWidget(self.btn_exportar_pdf)

        self.btn_exportar_excel = QPushButton("Exportar a Excel")
        self.btn_exportar_excel.clicked.connect(self.exportar_a_excel)
        botones_layout.addWidget(self.btn_exportar_excel)

        botones_widget = QWidget()
        botones_widget.setLayout(botones_layout)
        diseno_principal.addWidget(botones_widget)

        # Agregar checkbox para justificación
        self.chk_justificacion = QCheckBox("Requiere justificación")
        self.chk_justificacion.stateChanged.connect(self.toggle_justificacion)
        diseno_principal.addWidget(self.chk_justificacion)

        # Área de justificación
        self.justificacion_widget = QWidget()
        justificacion_layout = QVBoxLayout(self.justificacion_widget)

        lbl_justificacion = QLabel("Justificación:")
        justificacion_layout.addWidget(lbl_justificacion)

        # Campo de texto para la justificación
        self.txt_justificacion = QTextEdit()
        self.txt_justificacion.setPlaceholderText("Ingrese la justificación aquí...")
        self.txt_justificacion.setMinimumHeight(100)
        justificacion_layout.addWidget(self.txt_justificacion)

        # Inicialmente oculto
        self.justificacion_widget.setVisible(False)
        diseno_principal.addWidget(self.justificacion_widget)

        # Botón para guardar justificación
        self.btn_guardar_justificacion = QPushButton("Guardar Justificación")
        self.btn_guardar_justificacion.clicked.connect(self.guardar_justificacion)
        self.btn_guardar_justificacion.setVisible(False)
        diseno_principal.addWidget(self.btn_guardar_justificacion)

        # Contenedor del reporte
        self.contenedor_reporte = QWidget()
        self.diseno_reporte = QVBoxLayout()
        self.contenedor_reporte.setLayout(self.diseno_reporte)
        diseno_principal.addWidget(self.contenedor_reporte)

        widget_principal.setLayout(diseno_principal)
        self.setCentralWidget(widget_principal)

    def toggle_justificacion(self, state):
        # Mostrar u ocultar el área de justificación según el estado del checkbox
        self.justificacion_widget.setVisible(state == Qt.Checked)
        self.btn_guardar_justificacion.setVisible(state == Qt.Checked)

    def guardar_justificacion(self):
        if not self.chk_justificacion.isChecked():
            return

        justificacion = self.txt_justificacion.toPlainText()
        if not justificacion.strip():
            QMessageBox.warning(self, "Advertencia", "Debe proporcionar una justificación.")
            return

        fecha_fin = self.fecha_fin.date().toString("yyyy-MM-dd")
        try:
            cursor = self.conn.cursor()

            # Verificar si ya existe un registro para la fecha
            consulta_verificacion = "SELECT id FROM cierre_neto WHERE fecha = %s ORDER BY id DESC LIMIT 1"
            cursor.execute(consulta_verificacion, (fecha_fin,))
            resultado = cursor.fetchone()

            if resultado:
                # Si existe, actualizar la justificación
                id_registro = resultado[0]
                consulta = "UPDATE cierre_neto SET justificacion = %s WHERE id = %s"
                cursor.execute(consulta, (justificacion, id_registro))
            else:
                # Si no existe, mostrar mensaje
                QMessageBox.warning(self, "Advertencia",
                                    "No hay registros para la fecha seleccionada. Primero genere un reporte.")
                cursor.close()
                return

            self.conn.commit()
            cursor.close()

            QMessageBox.information(self, "Éxito", "Justificación guardada correctamente en la base de datos.")

        except Exception as err:
            QMessageBox.critical(self, "Error de BD", f"Error al guardar justificación: {err}")

    def generar_reporte(self):
        # Limpiar reporte anterior si existe
        for i in reversed(range(self.diseno_reporte.count())):
            self.diseno_reporte.itemAt(i).widget().setParent(None)

        # Obtener rango de fechas
        fecha_inicio = self.fecha_inicio.date().toString("yyyy-MM-dd")
        fecha_fin = self.fecha_fin.date().toString("yyyy-MM-dd")

        # Crear título del reporte
        etiqueta_titulo = QLabel("Reporte Mensual")
        etiqueta_titulo.setAlignment(Qt.AlignCenter)
        etiqueta_titulo.setFont(QFont("Arial", 14, QFont.Bold))
        self.diseno_reporte.addWidget(etiqueta_titulo)

        # Mostrar rango de fechas
        rango_fechas = QLabel(
            f"DEL {self.fecha_inicio.date().toString('dd/MM/yyyy')} AL {self.fecha_fin.date().toString('dd/MM/yyyy')}")
        rango_fechas.setAlignment(Qt.AlignCenter)
        self.diseno_reporte.addWidget(rango_fechas)

        # Obtener justificación guardada si existe
        justificacion_guardada = self.obtener_justificacion(fecha_fin)
        if justificacion_guardada:
            # Si hay una justificación guardada, mostrarla y marcar el checkbox
            self.chk_justificacion.setChecked(True)
            self.txt_justificacion.setText(justificacion_guardada)

        try:
            # Crear tabla de resumen
            tabla_resumen = QTableWidget(3, 3)
            tabla_resumen.setHorizontalHeaderLabels(["", "DEL", "AL"])

            # Establecer elementos de la tabla
            saldo_anterior = self.obtener_saldo_anterior(fecha_inicio)

            total_ingresos, total_gastos, saldo_final = self.obtener_resumen_financiero(fecha_inicio, fecha_fin)
            tabla_resumen.setItem(0, 0, QTableWidgetItem("Saldo final mes anterior"))
            tabla_resumen.setItem(0, 1, QTableWidgetItem("Q"))
            tabla_resumen.setItem(0, 2, QTableWidgetItem(f"{saldo_anterior:,.2f}"))

            tabla_resumen.setItem(1, 0, QTableWidgetItem("Saldo final en libros"))
            tabla_resumen.setItem(1, 1, QTableWidgetItem("Q"))
            tabla_resumen.setItem(1, 2, QTableWidgetItem(f"{saldo_final:,.2f}"))

            # Formatear tabla
            tabla_resumen.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
            tabla_resumen.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
            tabla_resumen.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)

            self.diseno_reporte.addWidget(tabla_resumen)

            # Añadir tabla detallada de movimientos financieros
            # Obtenemos primero la lista de doctores para determinar cuántas filas necesitamos
            doctores = self.obtener_lista_doctores()
            num_filas = 6 + len(doctores)  # Filas base + una fila por cada doctor

            tabla_detalles = QTableWidget(num_filas, 3)
            tabla_detalles.setColumnCount(3)
            tabla_detalles.horizontalHeader().setVisible(False)

            # Establecer elementos de fila
            tabla_detalles.setItem(0, 0, QTableWidgetItem("(+) Saldo final mes anterior"))
            tabla_detalles.setItem(0, 1, QTableWidgetItem("Q"))
            tabla_detalles.setItem(0, 2, QTableWidgetItem(f"{saldo_anterior:,.2f}"))

            detalles_ingresos = self.obtener_detalles_ingresos(fecha_inicio, fecha_fin)
            tabla_detalles.setItem(1, 0, QTableWidgetItem("(+) Total de ingresos Cierre crudo Mensual"))
            tabla_detalles.setItem(1, 1, QTableWidgetItem("Q"))
            tabla_detalles.setItem(1, 2, QTableWidgetItem(f"{detalles_ingresos:,.2f}"))

            comisiones_bancarias = self.obtener_comisiones_bancarias(fecha_inicio, fecha_fin)
            tabla_detalles.setItem(2, 0, QTableWidgetItem("(-) Total de Comisiones bancarias"))
            tabla_detalles.setItem(2, 1, QTableWidgetItem("Q"))
            tabla_detalles.setItem(2, 2, QTableWidgetItem(f"{comisiones_bancarias:,.2f}"))

            impuestos = self.obtener_impuestos(fecha_inicio, fecha_fin)
            tabla_detalles.setItem(3, 0, QTableWidgetItem("(-) Total de impuestos del mes"))
            tabla_detalles.setItem(3, 1, QTableWidgetItem("Q"))
            tabla_detalles.setItem(3, 2, QTableWidgetItem(f"{impuestos:,.2f}"))

            # Añadir dinámicamente filas para cada doctor
            fila_actual = 4
            total_comisiones_doctores = 0

            for i, doctor in enumerate(doctores):
                id_doctor = doctor['id']
                nombre_doctor = doctor['doctor']
                comision = self.obtener_comision_doctor(fecha_inicio, fecha_fin, id_doctor)
                total_comisiones_doctores += comision

                tabla_detalles.setItem(fila_actual, 0, QTableWidgetItem(f"(-) Total de comisiones a {nombre_doctor}"))
                tabla_detalles.setItem(fila_actual, 1, QTableWidgetItem("Q"))
                tabla_detalles.setItem(fila_actual, 2, QTableWidgetItem(f"{comision:,.2f}"))
                fila_actual += 1

            gastos = self.obtener_gastos(fecha_inicio, fecha_fin)
            tabla_detalles.setItem(fila_actual, 0, QTableWidgetItem("(-) Total de gastos del mes"))
            tabla_detalles.setItem(fila_actual, 1, QTableWidgetItem("Q"))
            tabla_detalles.setItem(fila_actual, 2, QTableWidgetItem(f"{gastos:,.2f}"))
            fila_actual += 1

            # Añadir una línea
            linea = QFrame()
            linea.setFrameShape(QFrame.HLine)
            linea.setFrameShadow(QFrame.Sunken)

            # Fila de saldo final con fuente en negrita
            fuente = QFont()
            fuente.setBold(True)

            tabla_detalles.setItem(fila_actual, 0,
                                   QTableWidgetItem(f"Saldo final al {self.fecha_fin.date().toString('dd/MM/yyyy')}"))
            tabla_detalles.setItem(fila_actual, 1, QTableWidgetItem("Q"))
            tabla_detalles.setItem(fila_actual, 2, QTableWidgetItem(f"{saldo_final:,.2f}"))

            for col in range(3):
                if tabla_detalles.item(fila_actual, col):
                    tabla_detalles.item(fila_actual, col).setFont(fuente)

            # Formatear tabla de detalles
            tabla_detalles.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
            tabla_detalles.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
            tabla_detalles.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
            tabla_detalles.verticalHeader().setVisible(False)

            self.diseno_reporte.addWidget(tabla_detalles)
            self.mostrar_cheques_circulacion(fecha_fin)
            #self.mostrar_cheques_en_circulacion(fecha_fin)
            # SECCIÓN: Añadir tabla de productos del rango de fechas seleccionado
            self.agregar_tabla_productos_por_rango(fecha_inicio, fecha_fin)

            # Guardar el cierre en la base de datos si no existe
            self.guardar_cierre_en_bd(fecha_fin, saldo_final)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Ocurrió un error al generar el reporte: {str(e)}")

    def guardar_cierre_en_bd(self, fecha_fin, saldo_final):
        """Guarda el resultado del cierre en la tabla cierre_neto si no existe"""
        try:
            cursor = self.conn.cursor()

            # Verificar si ya existe un registro para esta fecha
            consulta_check = "SELECT id FROM cierre_neto WHERE fecha = %s"
            cursor.execute(consulta_check, (fecha_fin,))
            registro_existente = cursor.fetchone()

            if not registro_existente:
                # Si no existe, guardar el nuevo registro
                consulta_insert = """
                    INSERT INTO cierre_neto (fecha, resultado_neto, efectivo_neto) 
                    VALUES (%s, %s, %s)
                """
                # Usamos el saldo_final como resultado_neto y efectivo_neto por simplicidad
                # En un caso real, estos valores podrían ser diferentes
                cursor.execute(consulta_insert, (fecha_fin, saldo_final, saldo_final))
                self.conn.commit()
                print(f"Nuevo cierre guardado para la fecha {fecha_fin}")

            cursor.close()
        except Exception as e:
            print(f"Error al guardar cierre en BD: {e}")

    def agregar_tabla_productos_por_rango(self, fecha_inicio, fecha_fin):
        """Añade una tabla con los productos vendidos en el rango de fechas seleccionado"""
        # Crear un título para esta sección
        etiqueta_productos = QLabel("Productos Vendidos en el Período Seleccionado")
        etiqueta_productos.setAlignment(Qt.AlignCenter)
        etiqueta_productos.setFont(QFont("Arial", 12, QFont.Bold))
        self.diseno_reporte.addWidget(etiqueta_productos)

        # Mostrar el rango de fechas
        fecha_mostrar = QLabel(
            f"Del: {self.fecha_inicio.date().toString('dd/MM/yyyy')} al {self.fecha_fin.date().toString('dd/MM/yyyy')}")
        fecha_mostrar.setAlignment(Qt.AlignCenter)
        self.diseno_reporte.addWidget(fecha_mostrar)

        # Obtener los productos vendidos en el rango de fechas
        productos = self.obtener_productos_por_rango(fecha_inicio, fecha_fin)

        if not productos:
            # Si no hay productos, mostrar un mensaje
            sin_productos = QLabel("No hay productos vendidos en el rango de fechas seleccionado.")
            sin_productos.setAlignment(Qt.AlignCenter)
            self.diseno_reporte.addWidget(sin_productos)
            return

        # Crear una tabla para mostrar los productos
        tabla_productos = QTableWidget(len(productos), 8)  # Agregamos la columna de fecha
        tabla_productos.setHorizontalHeaderLabels(
            ["ID", "Nombre", "Cantidad", "Tarjeta", "Efectivo", "Monto", "Usuario", "Fecha"])

        # Llenar la tabla con los productos
        for i, producto in enumerate(productos):
            tabla_productos.setItem(i, 0, QTableWidgetItem(str(producto['id'])))
            tabla_productos.setItem(i, 1, QTableWidgetItem(producto['nombre']))
            tabla_productos.setItem(i, 2, QTableWidgetItem(str(producto['cantidad'])))
            tabla_productos.setItem(i, 3, QTableWidgetItem(str(producto['tarjeta'])))
            tabla_productos.setItem(i, 4, QTableWidgetItem(str(producto['efectivo'])))
            tabla_productos.setItem(i, 5, QTableWidgetItem(str(producto['monto'])))
            tabla_productos.setItem(i, 6, QTableWidgetItem(producto['usuario']))
            # Formatear la fecha para mostrarla
            fecha_formateada = datetime.strptime(str(producto['fecha']), "%Y-%m-%d").strftime("%d/%m/%Y")
            tabla_productos.setItem(i, 7, QTableWidgetItem(fecha_formateada))

        # Formatear tabla
        tabla_productos.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        tabla_productos.verticalHeader().setVisible(False)

        self.diseno_reporte.addWidget(tabla_productos)

        # Mostrar resumen de ventas del período
        total_efectivo, total_tarjeta, total_ventas = self.calcular_totales_ventas(productos)

        resumen_ventas = QLabel(f"Resumen de Ventas del Período:")
        resumen_ventas.setFont(QFont("Arial", 10, QFont.Bold))
        self.diseno_reporte.addWidget(resumen_ventas)

        detalle_ventas = QLabel(
            f"Total Efectivo: Q {total_efectivo:,.2f} | Total Tarjeta: Q {total_tarjeta:,.2f} | Total: Q {total_ventas:,.2f}")
        self.diseno_reporte.addWidget(detalle_ventas)

    def obtener_productos_por_rango(self, fecha_inicio, fecha_fin):
        """Obtiene los productos vendidos en el rango de fechas desde la base de datos"""
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            consulta = """
                SELECT id, nombre, cantidad, tarjeta, efectivo, monto, usuario, carrito_id, fecha
                FROM bdhidrocolon.cierre
                WHERE fecha BETWEEN %s AND %s
                ORDER BY fecha, id
            """
            cursor.execute(consulta, (fecha_inicio, fecha_fin))
            productos = cursor.fetchall()
            cursor.close()

            return productos
        except Exception as err:
            QMessageBox.warning(self, "Error de Base de Datos",
                                f"Error al recuperar productos del período: {err}")
            return []

    def calcular_totales_ventas(self, productos):
        """Calcula los totales de ventas del período seleccionado a partir de los productos"""
        total_efectivo = 0
        total_tarjeta = 0

        for producto in productos:
            total_efectivo += float(producto['efectivo'])
            total_tarjeta += float(producto['tarjeta'])

        total_ventas = total_efectivo + total_tarjeta

        return total_efectivo, total_tarjeta, total_ventas

    def obtener_justificacion(self, fecha_fin):
        """Obtiene la justificación guardada para la fecha seleccionada"""
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            consulta = """
                SELECT justificacion FROM cierre_neto
                WHERE fecha = %s
                ORDER BY id DESC LIMIT 1
            """
            cursor.execute(consulta, (fecha_fin,))
            resultado = cursor.fetchone()
            cursor.close()

            if resultado and resultado['justificacion']:
                return resultado['justificacion']
            return ""
        except Exception as err:
            print(f"Error al recuperar justificación: {err}")
            return ""

    def obtener_saldo_anterior(self, fecha_inicio):
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            # Obtenemos la fecha de inicio como objeto fecha
            fecha_inicio_obj = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()

            # Buscamos la última entrada anterior a la fecha de inicio
            consulta = """
                SELECT resultado_neto FROM cierre_neto
                WHERE fecha < %s 
                ORDER BY fecha DESC, id DESC
                LIMIT 1
            """
            cursor.execute(consulta, (fecha_inicio,))
            resultado = cursor.fetchone()
            cursor.close()

            if resultado and resultado['resultado_neto'] is not None:
                return float(resultado['resultado_neto'])
            return 0.00  # Saldo inicial predeterminado si no se encuentra registro previo

        except Exception as err:
            QMessageBox.warning(self, "Error de Base de Datos", f"Error al recuperar saldo anterior: {err}")
            return 0.00

    def obtener_lista_doctores(self):
        """Obtiene la lista de doctores desde la base de datos"""
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            consulta = "SELECT id, doctor FROM bdhidrocolon.pagos_comisiones GROUP BY id, doctor"
            cursor.execute(consulta)
            doctores = cursor.fetchall()
            cursor.close()

            if not doctores:
                # Si no hay doctores, devolvemos una lista predeterminada para evitar errores
                return [{'id': 1, 'doctor': 'Doctor Predeterminado'}]

            return doctores
        except Exception as err:
            QMessageBox.warning(self, "Error de Base de Datos", f"Error al recuperar lista de doctores: {err}")
            return [{'id': 1, 'doctor': 'Doctor Predeterminado'}]

    def obtener_comision_doctor(self, fecha_inicio, fecha_fin, id_doctor):
        """Obtiene la comisión para un doctor específico"""
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            consulta = """
                SELECT SUM(monto) as total_comision 
                FROM bdhidrocolon.pagos_comisiones
                WHERE id = %s AND fecha BETWEEN %s AND %s
            """
            cursor.execute(consulta, (id_doctor, fecha_inicio, fecha_fin))
            resultado = cursor.fetchone()
            cursor.close()

            if resultado and resultado['total_comision'] is not None:
                return float(resultado['total_comision'])
            return 0.00  # Si no hay comisiones, devolvemos 0
        except Exception as err:
            QMessageBox.warning(self, "Error de Base de Datos",
                                f"Error al recuperar comisión del doctor {id_doctor}: {err}")
            return 0.00

    def obtener_resumen_financiero(self, fecha_inicio, fecha_fin):
        # Calcular ingresos totales
        ingresos = self.obtener_detalles_ingresos(fecha_inicio, fecha_fin)

        # Calcular gastos totales
        comisiones_bancarias = self.obtener_comisiones_bancarias(fecha_inicio, fecha_fin)
        impuestos = self.obtener_impuestos(fecha_inicio, fecha_fin)

        # Obtener comisiones para todos los doctores
        doctores = self.obtener_lista_doctores()
        total_comisiones = 0
        for doctor in doctores:
            total_comisiones += self.obtener_comision_doctor(fecha_inicio, fecha_fin, doctor['id'])

        gastos = self.obtener_gastos(fecha_inicio, fecha_fin)

        total_gastos = comisiones_bancarias + impuestos + total_comisiones + gastos

        # Obtener saldo anterior
        saldo_anterior = self.obtener_saldo_anterior(fecha_inicio)

        # Calcular saldo final
        saldo_final = saldo_anterior + ingresos - total_gastos

        return ingresos, total_gastos, saldo_final

    def obtener_detalles_ingresos(self, fecha_inicio, fecha_fin):
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            consulta = "SELECT SUM(totales) as total FROM cierre_crudo WHERE fecha BETWEEN %s AND %s"

            cursor.execute(consulta, (fecha_inicio, fecha_fin))
            resultado = cursor.fetchone()
            cursor.close()

            if resultado and resultado['total'] is not None:
                return float(resultado['total'])
            return 0.00  # Valor predeterminado si no se encuentran registros
        except Exception as err:
            QMessageBox.warning(self, "Error de Base de Datos", f"Error al recuperar detalles de ingresos: {err}")
            return 0.00

    def obtener_comisiones_bancarias(self, fecha_inicio, fecha_fin):
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            co = """
                SELECT SUM(total_real) as total_r FROM ingreso_tarjeta
                WHERE fecha BETWEEN %s AND %s
            """
            cursor.execute(co, (fecha_inicio, fecha_fin))
            resu = cursor.fetchone()
            cursor.close()

            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            co = """
                SELECT SUM(total_bancos) as total_b FROM ingreso_tarjeta
                WHERE fecha BETWEEN %s AND %s
            """
            cursor.execute(co, (fecha_inicio, fecha_fin))
            resul = cursor.fetchone()
            cursor.close()

            if resu['total_r'] is None or resul['total_b'] is None:
                return 0.00

            resultado = float(resu['total_r']) - float(resul['total_b'])

            if resultado:
                return resultado
            return 0.00  # Valor predeterminado
        except Exception as e:
            QMessageBox.warning(self, "Error de Base de Datos", f"Error al calcular comisiones bancarias: {e}")
            return 0.00

    def obtener_impuestos(self, fecha_inicio, fecha_fin):
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            consulta = "SELECT SUM(venta_neta) as total FROM cierre_crudo WHERE fecha BETWEEN %s AND %s"
            cursor.execute(consulta, (fecha_inicio, fecha_fin))
            res = cursor.fetchone()
            cursor.close()

            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            consulta = "SELECT SUM(efectivo_neto) as total FROM cierre_neto WHERE fecha BETWEEN %s AND %s"
            cursor.execute(consulta, (fecha_inicio, fecha_fin))
            resu = cursor.fetchone()
            cursor.close()

            # Verificar si hay valores nulos
            if res['total'] is None or resu['total'] is None:
                return 0.00

            resultado = float(res['total']) - float(resu['total'])

            if resultado:
                return resultado
            return 0.00  # Valor predeterminado
        except Exception as e:
            QMessageBox.warning(self, "Error de Base de Datos", f"Error al calcular impuestos: {e}")
            return 0.00

    def obtener_gastos(self, fecha_inicio, fecha_fin):
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            consulta = """
                SELECT SUM(monto) as total FROM gastos
                WHERE fecha BETWEEN %s AND %s
            """
            cursor.execute(consulta, (fecha_inicio, fecha_fin))
            resultado = cursor.fetchone()
            cursor.close()

            if resultado and resultado['total']:
                return float(resultado['total'])
            return 1100.00  # Valor predeterminado
        except Exception as e:
            QMessageBox.warning(self, "Error de Base de Datos", f"Error al calcular impuestos: {e}")
            return 1100.00

    def obtener_cheques_en_circulacion(self, fecha_fin):
        try:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)
            consulta = """
                SELECT no_cheque, fecha, emitido_A, detalle, monto 
                FROM cheques
                WHERE fecha <= %s AND estado = "Circulación" ORDER BY fecha
            """
            cursor.execute(consulta, (fecha_fin,))
            resultados = cursor.fetchall()
            cursor.close()

            print(resultados)
            return resultados

        except Exception as e:
            QMessageBox.warning(self, "Error de Base de Datos", f"Error al recuperar cheques en circulación: {e}")
            return [{
                'numero_cheque': '1234',
                'fecha': datetime.strptime('2025-03-06', '%Y-%m-%d'),
                'beneficiario': 'pablo',
                'descripcion': 'Publicidad',
                'monto': 500.00
            }]

    def mostrar_cheques_circulacion(self, fecha_fin):
        """Muestra los cheques en circulación hasta la fecha seleccionada"""
        # Título para la sección de cheques
        etiqueta_cheques = QLabel("Cheques en Circulación")
        etiqueta_cheques.setAlignment(Qt.AlignCenter)
        etiqueta_cheques.setFont(QFont("Arial", 12, QFont.Bold))
        self.diseno_reporte.addWidget(etiqueta_cheques)

        # Obtener los cheques en circulación
        cheques = self.obtener_cheques_en_circulacion(fecha_fin)

        if not cheques:
            # Si no hay cheques, mostrar un mensaje
            sin_cheques = QLabel("No hay cheques en circulación para la fecha seleccionada.")
            sin_cheques.setAlignment(Qt.AlignCenter)
            self.diseno_reporte.addWidget(sin_cheques)
            return

        # Crear una tabla para mostrar los cheques
        tabla_cheques = QTableWidget(len(cheques), 5)
        tabla_cheques.setHorizontalHeaderLabels(
            ["No. Cheque", "Fecha", "Emitido a", "Detalle", "Monto"])

        # Llenar la tabla con los cheques
        total_cheques = 0
        for i, cheque in enumerate(cheques):
            tabla_cheques.setItem(i, 0, QTableWidgetItem(str(cheque['no_cheque'])))

            # Formatear la fecha para mostrarla
            fecha_formateada = datetime.strptime(str(cheque['fecha']), "%Y-%m-%d").strftime("%d/%m/%Y")
            tabla_cheques.setItem(i, 1, QTableWidgetItem(fecha_formateada))

            tabla_cheques.setItem(i, 2, QTableWidgetItem(cheque['emitido_A']))
            tabla_cheques.setItem(i, 3, QTableWidgetItem(cheque['detalle']))
            tabla_cheques.setItem(i, 4, QTableWidgetItem(f"Q {float(cheque['monto']):,.2f}"))

            # Sumar al total
            total_cheques += float(cheque['monto'])

        # Formatear tabla
        tabla_cheques.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        tabla_cheques.verticalHeader().setVisible(False)

        self.diseno_reporte.addWidget(tabla_cheques)

        # Mostrar el total de cheques en circulación
        total_label = QLabel(f"Total de Cheques en Circulación: Q {total_cheques:,.2f}")
        total_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.diseno_reporte.addWidget(total_label)

    def generar_pdf(self):
        """Genera un archivo PDF con el contenido del reporte actual"""
        # Primero verificamos si hay datos para generar el PDF
        if self.diseno_reporte.count() == 0:
            QMessageBox.warning(self, "Advertencia", "Primero genere un reporte para crear el PDF.")
            return

        # Solicitar al usuario la ubicación donde guardar el PDF
        ruta_pdf, _ = QFileDialog.getSaveFileName(self, "Guardar PDF", "", "Archivos PDF (*.pdf)")
        if not ruta_pdf:
            return  # El usuario canceló la operación

        try:
            # Añadir .pdf si el usuario no lo incluyó
            if not ruta_pdf.endswith('.pdf'):
                ruta_pdf += '.pdf'

            # Crear el documento PDF
            doc = SimpleDocTemplate(ruta_pdf, pagesize=letter)
            elementos = []
            styles = getSampleStyleSheet()

            # Agregar título
            elementos.append(Paragraph("Reporte Financiero Mensual", styles['Title']))
            elementos.append(Spacer(1, 12))

            # Agregar rango de fechas
            elementos.append(Paragraph(
                f"Del {self.fecha_inicio.date().toString('dd/MM/yyyy')} al {self.fecha_fin.date().toString('dd/MM/yyyy')}",
                styles['Heading2']))
            elementos.append(Spacer(1, 12))

            # Extraer datos para crear las tablas en el PDF
            fecha_inicio = self.fecha_inicio.date().toString("yyyy-MM-dd")
            fecha_fin = self.fecha_fin.date().toString("yyyy-MM-dd")

            # Obtener datos financieros
            saldo_anterior = self.obtener_saldo_anterior(fecha_inicio)
            total_ingresos, total_gastos, saldo_final = self.obtener_resumen_financiero(fecha_inicio, fecha_fin)

            # Crear tabla de resumen
            datos_resumen = [
                ["", "DEL", "AL"],
                ["Saldo final mes anterior", "Q", f"{saldo_anterior:,.2f}"],
                ["Saldo final en libros", "Q", f"{saldo_final:,.2f}"]
            ]

            tabla_resumen = Table(datos_resumen)
            tabla_resumen.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (2, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (2, 0), colors.black),
                ('ALIGN', (0, 0), (2, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (2, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (2, 0), 12),
                ('BACKGROUND', (0, 1), (2, 2), colors.white),
                ('GRID', (0, 0), (2, 2), 1, colors.black),
            ]))

            elementos.append(tabla_resumen)
            elementos.append(Spacer(1, 24))

            # Añadir detalles financieros
            elementos.append(Paragraph("Detalles Financieros", styles['Heading2']))
            elementos.append(Spacer(1, 12))

            # Construir datos para la tabla detallada
            doctores = self.obtener_lista_doctores()
            datos_detalles = [
                ["(+) Saldo final mes anterior", "Q", f"{saldo_anterior:,.2f}"],
                ["(+) Total de ingresos Cierre crudo Mensual", "Q",
                 f"{self.obtener_detalles_ingresos(fecha_inicio, fecha_fin):,.2f}"],
                ["(-) Total de Comisiones bancarias", "Q",
                 f"{self.obtener_comisiones_bancarias(fecha_inicio, fecha_fin):,.2f}"],
                ["(-) Total de impuestos del mes", "Q", f"{self.obtener_impuestos(fecha_inicio, fecha_fin):,.2f}"]
            ]

            # Añadir comisiones por doctor
            for doctor in doctores:
                comision = self.obtener_comision_doctor(fecha_inicio, fecha_fin, doctor['id'])
                datos_detalles.append([f"(-) Total de comisiones a {doctor['doctor']}", "Q", f"{comision:,.2f}"])

            # Añadir gastos
            gastos = self.obtener_gastos(fecha_inicio, fecha_fin)
            datos_detalles.append(["(-) Total de gastos del mes", "Q", f"{gastos:,.2f}"])

            # Añadir saldo final
            datos_detalles.append(
                [f"Saldo final al {self.fecha_fin.date().toString('dd/MM/yyyy')}", "Q", f"{saldo_final:,.2f}"])

            tabla_detalles = Table(datos_detalles)
            tabla_detalles.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))

            elementos.append(tabla_detalles)
            elementos.append(Spacer(1, 24))

            # Agregar productos vendidos en el período
            elementos.append(Paragraph("Productos Vendidos en el Período", styles['Heading2']))
            elementos.append(Spacer(1, 12))

            productos = self.obtener_productos_por_rango(fecha_inicio, fecha_fin)

            if productos:
                # Encabezado de la tabla de productos
                datos_productos = [["ID", "Nombre", "Cantidad", "Tarjeta", "Efectivo", "Monto", "Usuario", "Fecha"]]

                # Añadir datos de productos
                for producto in productos:
                    fecha_formateada = datetime.strptime(str(producto['fecha']), "%Y-%m-%d").strftime("%d/%m/%Y")
                    datos_productos.append([
                        str(producto['id']),
                        producto['nombre'],
                        str(producto['cantidad']),
                        str(producto['tarjeta']),
                        str(producto['efectivo']),
                        str(producto['monto']),
                        producto['usuario'],
                        fecha_formateada
                    ])

                tabla_productos = Table(datos_productos)
                tabla_productos.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))

                elementos.append(tabla_productos)
                elementos.append(Spacer(1, 12))

                # Añadir resumen de ventas
                total_efectivo, total_tarjeta, total_ventas = self.calcular_totales_ventas(productos)
                elementos.append(Paragraph(f"Resumen de Ventas:", styles['Heading3']))
                elementos.append(Paragraph(f"Total Efectivo: Q {total_efectivo:,.2f}", styles['Normal']))
                elementos.append(Paragraph(f"Total Tarjeta: Q {total_tarjeta:,.2f}", styles['Normal']))
                elementos.append(Paragraph(f"Total: Q {total_ventas:,.2f}", styles['Normal']))
            else:
                elementos.append(Paragraph("No hay productos vendidos en el período seleccionado.", styles['Normal']))

            # Agregar justificación si existe
            if self.chk_justificacion.isChecked() and self.txt_justificacion.toPlainText().strip():
                elementos.append(Spacer(1, 24))
                elementos.append(Paragraph("Justificación:", styles['Heading2']))
                elementos.append(Paragraph(self.txt_justificacion.toPlainText(), styles['Normal']))

            # Construir el PDF
            doc.build(elementos)

            QMessageBox.information(self, "Éxito", f"PDF generado exitosamente y guardado en {ruta_pdf}")

        except Exception as e:
            QMessageBox.critical(self, "Error al Generar PDF", f"Ocurrió un error al generar el PDF: {str(e)}")

    # MÉTODO PARA EXPORTAR A EXCEL
    # Agregar este método a la clase AplicacionReporteFinanciero
    def exportar_a_excel(self):
        """Exporta los datos del reporte actual a un archivo Excel"""
        # Primero verificamos si hay datos para generar el Excel
        if self.diseno_reporte.count() == 0:
            QMessageBox.warning(self, "Advertencia", "Primero genere un reporte para exportar a Excel.")
            return

        # Solicitar al usuario la ubicación donde guardar el archivo Excel
        ruta_excel, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", "", "Archivos Excel (*.xlsx)")
        if not ruta_excel:
            return  # El usuario canceló la operación

        try:
            # Añadir .xlsx si el usuario no lo incluyó
            if not ruta_excel.endswith('.xlsx'):
                ruta_excel += '.xlsx'

            # Crear un nuevo libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Financiero"

            # Establecer título
            ws['A1'] = "REPORTE FINANCIERO MENSUAL"
            ws.merge_cells('A1:G1')
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Añadir rango de fechas
            fecha_rango = f"Del {self.fecha_inicio.date().toString('dd/MM/yyyy')} al {self.fecha_fin.date().toString('dd/MM/yyyy')}"
            ws['A2'] = fecha_rango
            ws.merge_cells('A2:G2')
            ws['A2'].font = Font(size=12)
            ws['A2'].alignment = Alignment(horizontal='center')

            # Obtener datos para el reporte
            fecha_inicio = self.fecha_inicio.date().toString("yyyy-MM-dd")
            fecha_fin = self.fecha_fin.date().toString("yyyy-MM-dd")
            saldo_anterior = self.obtener_saldo_anterior(fecha_inicio)
            total_ingresos, total_gastos, saldo_final = self.obtener_resumen_financiero(fecha_inicio, fecha_fin)

            # Añadir tabla de resumen
            ws['A4'] = "RESUMEN"
            ws.merge_cells('A4:C4')
            ws['A4'].font = Font(bold=True)
            ws['A4'].alignment = Alignment(horizontal='center')

            encabezados_resumen = ["", "DEL", "AL"]
            for col, encabezado in enumerate(encabezados_resumen, 1):
                ws.cell(row=5, column=col).value = encabezado
                ws.cell(row=5, column=col).font = Font(bold=True)

            ws['A6'] = "Saldo final mes anterior"
            ws['B6'] = "Q"
            ws['C6'] = saldo_anterior

            ws['A7'] = "Saldo final en libros"
            ws['B7'] = "Q"
            ws['C7'] = saldo_final

            # Aplicar formato de números para los valores monetarios
            ws['C6'].number_format = '#,##0.00'
            ws['C7'].number_format = '#,##0.00'

            # Añadir tabla de detalles financieros
            ws['A9'] = "DETALLES FINANCIEROS"
            ws.merge_cells('A9:C9')
            ws['A9'].font = Font(bold=True)
            ws['A9'].alignment = Alignment(horizontal='center')

            row = 10
            # Detalles financieros - Ingresos y gastos
            ws.cell(row=row, column=1).value = "(+) Saldo final mes anterior"
            ws.cell(row=row, column=2).value = "Q"
            ws.cell(row=row, column=3).value = saldo_anterior
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1

            detalles_ingresos = self.obtener_detalles_ingresos(fecha_inicio, fecha_fin)
            ws.cell(row=row, column=1).value = "(+) Total de ingresos Cierre crudo Mensual"
            ws.cell(row=row, column=2).value = "Q"
            ws.cell(row=row, column=3).value = detalles_ingresos
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1

            comisiones_bancarias = self.obtener_comisiones_bancarias(fecha_inicio, fecha_fin)
            ws.cell(row=row, column=1).value = "(-) Total de Comisiones bancarias"
            ws.cell(row=row, column=2).value = "Q"
            ws.cell(row=row, column=3).value = comisiones_bancarias
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1

            impuestos = self.obtener_impuestos(fecha_inicio, fecha_fin)
            ws.cell(row=row, column=1).value = "(-) Total de impuestos del mes"
            ws.cell(row=row, column=2).value = "Q"
            ws.cell(row=row, column=3).value = impuestos
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1

            # Añadir comisiones por doctor
            doctores = self.obtener_lista_doctores()
            for doctor in doctores:
                comision = self.obtener_comision_doctor(fecha_inicio, fecha_fin, doctor['id'])
                ws.cell(row=row, column=1).value = f"(-) Total de comisiones a {doctor['doctor']}"
                ws.cell(row=row, column=2).value = "Q"
                ws.cell(row=row, column=3).value = comision
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                row += 1

            gastos = self.obtener_gastos(fecha_inicio, fecha_fin)
            ws.cell(row=row, column=1).value = "(-) Total de gastos del mes"
            ws.cell(row=row, column=2).value = "Q"
            ws.cell(row=row, column=3).value = gastos
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1

            # Saldo final con formato destacado
            ws.cell(row=row, column=1).value = f"Saldo final al {self.fecha_fin.date().toString('dd/MM/yyyy')}"
            ws.cell(row=row, column=2).value = "Q"
            ws.cell(row=row, column=3).value = saldo_final
            ws.cell(row=row, column=3).number_format = '#,##0.00'

            # Aplicar formato negrita al saldo final
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)

            # Añadir tabla de productos vendidos
            row += 2
            ws.cell(row=row, column=1).value = "PRODUCTOS VENDIDOS EN EL PERÍODO"
            ws.merge_cells(f'A{row}:H{row}')
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1

            # Encabezados de la tabla de productos
            encabezados_productos = ["ID", "Nombre", "Cantidad", "Tarjeta", "Efectivo", "Monto", "Usuario", "Fecha"]
            for col, encabezado in enumerate(encabezados_productos, 1):
                ws.cell(row=row, column=col).value = encabezado
                ws.cell(row=row, column=col).font = Font(bold=True)
                # Aplicar un fondo de color claro a los encabezados
                ws.cell(row=row, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9",
                                                                fill_type="solid")
            row += 1

            # Obtener y añadir datos de productos
            productos = self.obtener_productos_por_rango(fecha_inicio, fecha_fin)
            if productos:
                for producto in productos:
                    ws.cell(row=row, column=1).value = producto['id']
                    ws.cell(row=row, column=2).value = producto['nombre']
                    ws.cell(row=row, column=3).value = producto['cantidad']
                    ws.cell(row=row, column=4).value = producto['tarjeta']
                    ws.cell(row=row, column=5).value = producto['efectivo']
                    ws.cell(row=row, column=6).value = producto['monto']
                    ws.cell(row=row, column=7).value = producto['usuario']
                    # Formatear la fecha para mostrarla
                    fecha_formateada = datetime.strptime(str(producto['fecha']), "%Y-%m-%d").strftime("%d/%m/%Y")
                    ws.cell(row=row, column=8).value = fecha_formateada
                    row += 1

                # Calcular y añadir totales de ventas
                total_efectivo, total_tarjeta, total_ventas = self.calcular_totales_ventas(productos)

                row += 1
                ws.cell(row=row, column=1).value = "RESUMEN DE VENTAS:"
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1

                ws.cell(row=row, column=1).value = "Total Efectivo:"
                ws.cell(row=row, column=2).value = f"Q {total_efectivo:,.2f}"
                row += 1

                ws.cell(row=row, column=1).value = "Total Tarjeta:"
                ws.cell(row=row, column=2).value = f"Q {total_tarjeta:,.2f}"
                row += 1

                ws.cell(row=row, column=1).value = "Total Ventas:"
                ws.cell(row=row, column=2).value = f"Q {total_ventas:,.2f}"
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(bold=True)
            else:
                ws.cell(row=row, column=1).value = "No hay productos vendidos en el período seleccionado."
                ws.merge_cells(f'A{row}:H{row}')
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                row += 1

            # Añadir justificación si existe
            if self.chk_justificacion.isChecked() and self.txt_justificacion.toPlainText().strip():
                row += 2
                ws.cell(row=row, column=1).value = "JUSTIFICACIÓN:"
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1

                ws.cell(row=row, column=1).value = self.txt_justificacion.toPlainText()
                ws.merge_cells(f'A{row}:H{row}')
                # Ajustar el texto para que se vea bien
                ws.row_dimensions[row].height = max(50, len(self.txt_justificacion.toPlainText()) // 50 * 15)
                ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')

            # Ajustar ancho de columnas
            for col in range(1, 9):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 15

            # Guardar el archivo
            wb.save(ruta_excel)
            QMessageBox.information(self, "Éxito", f"Archivo Excel generado exitosamente y guardado en {ruta_excel}")

        except Exception as e:
            QMessageBox.critical(self, "Error al Exportar a Excel", f"Ocurrió un error al exportar a Excel: {str(e)}")
